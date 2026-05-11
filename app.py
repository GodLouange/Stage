import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import os
import glob
import base64
import urllib.request
import urllib.parse
import json

# ─────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="STADE AURILLACOIS",
    page_icon=" ",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
#  CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
  /* Background  */
  .stApp { background-color: #0e1117; }

  /* Sidebar */
  [data-testid="stSidebar"] {
    background: linear-gradient(180deg, #1a1f2e 0%, #0e1117 100%);
    border-right: 1px solid #2d3748;
  }

  /* KPI cards */
  .kpi-card {
    background: linear-gradient(135deg, #1a1f2e 0%, #2d3748 100%);
    border: 1px solid #4a5568;
    border-radius: 12px;
    padding: 20px;
    text-align: center;
    margin: 6px 0;
    transition: transform 0.2s;
  }
  .kpi-card:hover { transform: translateY(-2px); }
  .kpi-value { font-size: 2.2rem; font-weight: 800; margin: 4px 0; }
  .kpi-label { font-size: 0.82rem; color: #a0aec0; text-transform: uppercase; letter-spacing: 1px; }
  .kpi-delta { font-size: 0.78rem; margin-top: 4px; }
  .team-a-color { color: #63b3ed; }
  .team-b-color { color: #fc8181; }
  .win  { color: #68d391; }
  .lose { color: #fc8181; }

  /* Section headers */
  .section-title {
    font-size: 1.1rem; font-weight: 1000;
    padding-left: 12px; margin: 24px 0 16px 0;
  }

  /* Match banner  */
  .match-banner {
    background:linear-gradient(90deg, #1a365d 0%, #1a1f2e 50%, #3d1a1a 100%) ;
    border: 1px solid #4a5568; border-radius: 16px;
    padding: 28px 32px; text-align: center; margin-bottom: 28px;
  }
  .match-title { font-size: 2rem; font-weight: 900; color: #e2e8f0; }
  .match-sub { font-size: 0.9rem; color: #a0aec0; margin-top: 6px; }

  /* Pills */
  .pill {
    display: inline-block; padding: 3px 10px;
    border-radius: 999px; font-size: 0.75rem; font-weight: 600;
  }
  .pill-blue { background:#2b6cb0; color:#bee3f8; }
  .pill-red  { background:#742a2a; color:#fed7d7; }
  .pill-green{ background:#276749; color:#c6f6d5; }

  /* Divider */
  hr { border-color: #2d3748; }

  /* Plotly tweaks */
  .js-plotly-plot .plotly .modebar { background: transparent !important; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
#  DATA LOADING
# ─────────────────────────────────────────────
def _normalize_vertical(df: pd.DataFrame) -> pd.DataFrame:
    """
    Détecte et convertit le format VERTICAL (1 attribut par ligne) en format
    HORIZONTAL (1 action par ligne, tous attributs en colonnes).

    Format vertical identifié quand plusieurs lignes partagent le même
    'Début du clip' pour la même 'Nom de la ligne' (ratio > 2 lignes/action).

    Étapes :
      1. Supprime les colonnes horodatage internes (suffixe ': Temps')
      2. Groupe par (Nom de la ligne, Début du clip)
      3. Prend la première valeur non-nulle de chaque colonne dans le groupe
    """
    # Colonnes clé selon le format
    clip_col = None
    for candidate in ("Début du clip", "Temps de début"):
        if candidate in df.columns:
            clip_col = candidate
            break

    if clip_col is None:
        return df  # Pas de colonne de clip détectée → impossible de pivoter

    # Vérifier si le format est vraiment vertical
    # → plusieurs lignes par (action, clip) avec des NaN différents
    sample = df[["Nom de la ligne", clip_col]].dropna()
    if sample.empty:
        return df

    total_rows   = len(sample)
    unique_clips = sample.drop_duplicates().shape[0]
    ratio        = total_rows / max(unique_clips, 1)

    if ratio < 2.0:
        return df  # Déjà en format horizontal

    # ── Conversion verticale → horizontale ──────────────────────────
    # 1. Supprimer les colonnes horodatage internes (": Temps")
    cols_keep = [c for c in df.columns if not c.strip().endswith(": Temps")]
    df = df[cols_keep].copy()

    # 2. Remplacer les chaînes vides et "nan" par NaN
    for col in df.select_dtypes("object").columns:
        df[col] = df[col].replace({"nan": pd.NA, "": pd.NA})

    # 3. Grouper par (Nom de la ligne + clip_col) → first non-null par colonne
    key_cols  = ["Nom de la ligne", clip_col]
    attr_cols = [c for c in df.columns if c not in key_cols]

    df_h = (df.groupby(key_cols, sort=False, dropna=False)[attr_cols]
              .first()
              .reset_index())

    # 4. Renommer "Début du clip" → "Temps de début" pour compatibilité
    if clip_col == "Début du clip" and "Temps de début" not in df_h.columns:
        df_h = df_h.rename(columns={"Début du clip": "Temps de début"})

    return df_h


@st.cache_data(show_spinner="Chargement des données…")
def load_csv(filepath: str) -> pd.DataFrame:
    """
    Charge un CSV Dartfish/Stats Perform en gérant :
      - Plusieurs encodages (utf-16, utf-8-sig, latin-1…)
      - Format HORIZONTAL (1 ligne/action) et VERTICAL (N lignes/action)
    """
    for enc in ["utf-16", "utf-16-le", "utf-8-sig", "latin-1"]:
        try:
            df = pd.read_csv(
                filepath, encoding=enc, sep=",",
                low_memory=False, on_bad_lines="skip",
            )
            # Nettoie les noms de colonnes (BOM + espaces)
            df.columns = [c.strip().lstrip("\ufeff") for c in df.columns]
            if "Nom de la ligne" not in df.columns:
                continue

            # Normalise le format vertical → horizontal si nécessaire
            df = _normalize_vertical(df)

            # Normalise les valeurs texte
            for col in df.select_dtypes("object").columns:
                df[col] = df[col].astype(str).str.strip().replace("nan", pd.NA)

            return df
        except Exception:
            continue
    raise ValueError("Format de fichier non supporté.")


KNOWN_ACTIONS = [
    "Contacts", "Rucks", "Passes", "Soutiens Offensifs", "Franchissements",
    "Defenseurs battus", "Jeux a la main", "Essais", "ESSAI",
    "Plaquages", "Assistant plaqueur", "Contre Ruck", "Contest",
    "Turn over et contre attaque", "Ballons perdus", "Bras casses",
    "Jeux au pied", "Coups d'envoi", "Receptions aeriennes",
    "Renvois ligne de but", "Botteur", "Buteur",
    "Melees", "MELEE", "Touches", "TOUCHE", "Lancements sur melees",
    "Lancements sur touches", "Maul dans le jeu courant",
    "Penalites", "PENALITE", "Cartons", "FAUTE", "TMO", "Avantage",
    "Possession", "22m", "22m - 50m", "22m adverse", "22m - 50m adverse",
    "Remplacement", "CE", "CPP", "JOP", "R0", "RUCK", "TRANS",
]


def detect_teams(df: pd.DataFrame):
    """Détecte les équipes quel que soit leur nom (Stade X, SU X, RC X, US X…)."""
    teams = set()
    for line in df["Nom de la ligne"].dropna().unique():
        line = str(line).strip()
        if " - " not in line:
            continue
        parts = line.split(" - ", 1)
        team_candidate = parts[0].strip()
        action_candidate = parts[1].strip()
        # Ignore les codes courts : SA, ADV, CE…
        if len(team_candidate) <= 3:
            continue
        if any(action_candidate.startswith(a) for a in KNOWN_ACTIONS):
            teams.add(team_candidate)
    return sorted(teams)


def team_stats(df: pd.DataFrame, team: str) -> dict:
    prefix = f"{team} - "
    rows = df[df["Nom de la ligne"].str.startswith(prefix)]
    actions = rows["Nom de la ligne"].str.replace(prefix, "", regex=False)
    return actions.value_counts().to_dict()


def get_stat(stats: dict, key: str) -> int:
    return stats.get(key, 0)


def compute_score(df: pd.DataFrame, team: str) -> dict:
    essais    = int((df["Nom de la ligne"] == f"{team} - Essais").sum())
    buteur_df = df[df["Nom de la ligne"] == f"{team} - Buteur"]

    def _count(type_kw: str) -> int:
        if buteur_df.empty or "Type de tir au but" not in df.columns or "Resultat" not in df.columns:
            return 0
        return int(
            (buteur_df["Type de tir au but"].str.contains(type_kw, case=False, na=False) &
             (buteur_df["Resultat"].str.strip().str.lower() == "positif"))
            .sum()
        )

    transfo  = _count("Transformation")
    penalite = _count("nalit|enalt|[Pp]enalty")
    drop     = _count("[Dd]rop")
    score    = essais * 5 + transfo * 2 + penalite * 3 + drop * 3

    return {
        "score":    score,
        "essais":   essais,
        "transfo":  transfo,
        "penalite": penalite,
        "drop":     drop,
    }


# ─────────────────────────────────────────────
#  LOGO — Wikipedia / Wikimedia Commons API
# ─────────────────────────────────────────────

# Mots dans le nom de fichier qui indiquent un logo/blason (pas un stade, photo…)
_LOGO_KW    = ("logo", "blason", "ecusson", "écusson", "badge",
               "crest", "shield", "emblem", "insigne", "armoirie")
# Fichiers à ignorer (icônes génériques Wikipedia)
_SKIP_KW    = ("picto", "commons-logo", "flag_of", "red_question",
               "replace_this", "wikidata", "stub", "ambox", "portal",
               "edit-clear", "nuvola", "gnome", "crystal", "icon")


def _wiki_get(url: str, timeout: int = 5) -> dict:
    """Requête HTTP vers l'API Wikimedia."""
    req = urllib.request.Request(
        url, headers={"User-Agent": "RugbyDashboard/1.0"}
    )
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return json.loads(r.read())


def _file_score(filename: str) -> int:
    """Score de pertinence d'un fichier image : plus c'est élevé, plus c'est un logo."""
    low = filename.lower()
    if any(s in low for s in _SKIP_KW):
        return -1                          # à ignorer
    score = sum(3 for kw in _LOGO_KW if kw in low)   # +3 par mot-clé logo
    if low.endswith((".svg", ".png", ".jpg", ".jpeg", ".svg.png")):
        score += 1
    return score


def _file_url(filename: str, size: int = 300) -> str:
    """
    Récupère l'URL directe d'un fichier image (thumbnail PNG).
    Cherche dans l'ordre :
      1. Wikimedia Commons  (File:)
      2. fr.wikipedia.org   (Fichier:)
      3. en.wikipedia.org   (File:)
    Certains logos de clubs français sont sur fr.wikipedia et non sur Commons.
    """
    name = (filename
            .replace("File:", "").replace("Fichier:", "")
            .replace("Image:", "").strip())

    sources = [
        ("commons.wikimedia.org", "File"),
        ("fr.wikipedia.org",      "Fichier"),
        ("en.wikipedia.org",      "File"),
    ]
    for wiki, prefix in sources:
        try:
            url = (
                f"https://{wiki}/w/api.php"
                f"?action=query&titles={prefix}:{urllib.parse.quote(name)}"
                f"&prop=imageinfo&iiprop=url&iiurlwidth={size}"
                f"&format=json&origin=*"
            )
            data = _wiki_get(url)
            for page in data.get("query", {}).get("pages", {}).values():
                if int(page.get("pageid", -1)) < 0:
                    continue                    # fichier inexistant sur ce wiki
                for info in page.get("imageinfo", []):
                    u = info.get("thumburl") or info.get("url", "")
                    if u:
                        return u
        except Exception:
            continue
    return ""


def _logo_from_wiki_page(title: str, lang: str = "fr") -> str:
    """
    Ouvre une page Wikipedia, liste toutes ses images,
    trie par score (logo > blason > autres) et retourne
    l'URL de la meilleure image de type logo.
    """
    url = (
        f"https://{lang}.wikipedia.org/w/api.php"
        f"?action=query&titles={urllib.parse.quote(title)}"
        f"&prop=images&imlimit=50&format=json&origin=*"
    )
    data = _wiki_get(url)
    for page in data.get("query", {}).get("pages", {}).values():
        if int(page.get("pageid", -1)) < 0:     # page inexistante
            continue
        images = page.get("images", [])
        if not images:
            continue

        # Trier par score décroissant (logos en tête)
        scored = sorted(
            [(img["title"], _file_score(img["title"])) for img in images],
            key=lambda x: -x[1]
        )
        for filename, score in scored:
            if score < 0:      # fichier blacklisté
                continue
            try:
                u = _file_url(filename)
                if u:
                    return u
            except Exception:
                continue
    return ""


def _logo_from_commons_search(query: str) -> str:
    """
    Cherche directement sur Wikimedia Commons un fichier logo
    correspondant au nom du club (namespace 6 = Files).
    """
    url = (
        "https://commons.wikimedia.org/w/api.php"
        f"?action=query&list=search"
        f"&srsearch={urllib.parse.quote(query)}&srnamespace=6"
        f"&srlimit=8&format=json&origin=*"
    )
    data = _wiki_get(url)
    results = data.get("query", {}).get("search", [])

    # Trier les résultats par score de pertinence logo
    scored = sorted(
        [(r["title"], _file_score(r["title"])) for r in results],
        key=lambda x: -x[1]
    )
    for filename, score in scored:
        if score < 0:
            continue
        try:
            u = _file_url(filename)
            if u:
                return u
        except Exception:
            continue
    return ""


def _logo_from_wiki_search(query: str, lang: str = "fr") -> str:
    """Recherche Wikipedia full-text → visite chaque page trouvée → extrait le logo."""
    url = (
        f"https://{lang}.wikipedia.org/w/api.php"
        f"?action=query&list=search"
        f"&srsearch={urllib.parse.quote(query)}"
        f"&srlimit=4&format=json&origin=*"
    )
    data = _wiki_get(url)
    for result in data.get("query", {}).get("search", []):
        logo = _logo_from_wiki_page(result["title"], lang)
        if logo:
            return logo
    return ""


# ── Logos locaux (SVG/PNG copiés dans le dossier logos/) ──────────────────────
# Chargés en base64 pour être intégrés directement dans le HTML (pas de serveur)
_LOCAL_LOGOS_DIR = os.path.join(os.path.dirname(__file__), "logos")

_LOCAL_OVERRIDES = {
    # ── Stade Aurillacois ─────────────────────────────────────────────
    "aurillacois":          "logo SA.png",
    "stade aurillacois":    "logo SA.png",
    "sa":                   "logo SA.png",
    "auri":                 "logo SA.png",

    # ── Stade Niçois ──────────────────────────────────────────────────
    "niçois":               "Logo_Stade_Nicois_Rugby_2020.svg",
    "nicois":               "Logo_Stade_Nicois_Rugby_2020.svg",
    "stade niçois":         "Logo_Stade_Nicois_Rugby_2020.svg",
    "stade nicois":         "Logo_Stade_Nicois_Rugby_2020.svg",

    # ── RCNM — RC Narbonne Méditerranée ──────────────────────────────
    "rcnm":                 "1116px-Logo_RC_Narbonne_Méditerranée_(orange).png",
    "narbonne méditerranée":"1116px-Logo_RC_Narbonne_Méditerranée_(orange).png",
    "narbonne mediterranee":"1116px-Logo_RC_Narbonne_Méditerranée_(orange).png",

    # ── ASM Clermont Auvergne ─────────────────────────────────────────
    "clermont":             "1200px-Logo_ASM_Clermont_Auvergne_2019.png",
    "asm":                  "1200px-Logo_ASM_Clermont_Auvergne_2019.png",
    "asm clermont":         "1200px-Logo_ASM_Clermont_Auvergne_2019.png",

    # ── ASBH — AS Béziers Hérault (logo 2023) ────────────────────────
    "asbh":                 "Logo_Association_sportive_de_Béziers_Hérault_2023.png",
    "béziers":              "Logo_Association_sportive_de_Béziers_Hérault_2023.png",
    "beziers":              "Logo_Association_sportive_de_Béziers_Hérault_2023.png",

    # ── Aviron Bayonnais ──────────────────────────────────────────────
    "bayonne":              "1200px-Logo_Aviron_bayonnais_rugby_2010.png",
    "aviron bayonnais":     "1200px-Logo_Aviron_bayonnais_rugby_2010.png",

    # ── Biarritz Olympique Pays Basque ────────────────────────────────
    "biarritz":             "1200px-Logo_Biarritz_Olympique_Pays_Basque_-_2016.png",
    "bopb":                 "1200px-Logo_Biarritz_Olympique_Pays_Basque_-_2016.png",

    # ── Castres Olympique ─────────────────────────────────────────────
    "castres":              "1200px-Logo_Castres_olympique_2018.png",

    # ── Montpellier Hérault Rugby ─────────────────────────────────────
    "montpellier":          "1200px-Logo_Montpellier_Hérault_rugby_2013.png",
    "mhr":                  "1200px-Logo_Montpellier_Hérault_rugby_2013.png",

    # ── RC Massy Essonne ──────────────────────────────────────────────
    "massy":                "Logo_Rugby_Club_Massy_Essonne_-_2020.svg.png",
    "rc massy":             "Logo_Rugby_Club_Massy_Essonne_-_2020.svg.png",

    # ── RC Toulon ─────────────────────────────────────────────────────
    "toulon":               "1200px-Logo_RC_Toulon_2015.png",
    "rct":                  "1200px-Logo_RC_Toulon_2015.png",

    # ── Racing 92 ─────────────────────────────────────────────────────
    "racing 92":            "1200px-Logo_Racing_92_2015.png",
    "racing":               "1200px-Logo_Racing_92_2015.png",

    # ── Soyaux Angoulême XV Charente (SA XV) ──────────────────────────
    "sa xv":                "1200px-Logo_Soyaux_Angoulême_XV_Charente_-_2017.svg.png",
    "charente":             "1200px-Logo_Soyaux_Angoulême_XV_Charente_-_2017.svg.png",
    "angoulême":            "1200px-Logo_Soyaux_Angoulême_XV_Charente_-_2017.svg.png",
    "angouleme":            "1200px-Logo_Soyaux_Angoulême_XV_Charente_-_2017.svg.png",
    "soyaux":               "1200px-Logo_Soyaux_Angoulême_XV_Charente_-_2017.svg.png",
    "logo sa":              "logo SA.png",

    # ── Stade Français Paris ──────────────────────────────────────────
    "stade français":       "1200px-Logo_Stade_français_PR_2018.png",
    "stade francais":       "1200px-Logo_Stade_français_PR_2018.png",
    "sfp":                  "1200px-Logo_Stade_français_PR_2018.png",

    # ── USON Nevers ───────────────────────────────────────────────────
    "nevers":               "1200px-Logo_USON_Nevers_2016.png",
    "uson":                 "1200px-Logo_USON_Nevers_2016.png",
    "uson nevers":          "1200px-Logo_USON_Nevers_2016.png",

    # ── USBPA — US Bressane / Bourg-en-Bresse ────────────────────────
    "usbpa":                "1200px-Logo_US_Bressane_PA_2015.png",
    "bourg-en-bresse":      "1200px-Logo_US_Bressane_PA_2015.png",
    "bourg en bresse":      "1200px-Logo_US_Bressane_PA_2015.png",
    "bressane":             "1200px-Logo_US_Bressane_PA_2015.png",

    # ── UBB — Union Bordeaux Bègles ───────────────────────────────────
    "bordeaux":             "1200px-Logo_Union_Bordeaux_Bègles_2018.png",
    "ubb":                  "1200px-Logo_Union_Bordeaux_Bègles_2018.png",
    "union bordeaux":       "1200px-Logo_Union_Bordeaux_Bègles_2018.png",
    "bègles":               "1200px-Logo_Union_Bordeaux_Bègles_2018.png",

    # ── RCHCC — RC Hyères-Carqueiranne-La Crau ───────────────────────
    "rchcc":                "1200px-Logo_du_Rugby_club_Hyères-Carqueiranne-La_crau.svg.png",
    "hyères":               "1200px-Logo_du_Rugby_club_Hyères-Carqueiranne-La_crau.svg.png",
    "hyeres":               "1200px-Logo_du_Rugby_club_Hyères-Carqueiranne-La_crau.svg.png",

    # ── Stade Montois ─────────────────────────────────────────────────
    "stade montois":        "1200px-Stade_Montois_Rugby_Pro.png",

    # ── CSBJ — CS Bourgoin-Jallieu ────────────────────────────────────
    "csbj":                 "2560px-Logo_Club_sportif_Bourgoin-Jallieu_rugby.svg.png",
    "bourgoin":             "2560px-Logo_Club_sportif_Bourgoin-Jallieu_rugby.svg.png",

    # ── USAP — Perpignan ──────────────────────────────────────────────
    "usap":                 "790px-Logo_USA_Perpignan_2017.png",
    "perpignan":            "790px-Logo_USA_Perpignan_2017.png",

    # ── Stade Rochelais ───────────────────────────────────────────────
    "rochelais":            "800px-Logo_Stade_rochelais_2016.png",
    "la rochelle":          "800px-Logo_Stade_rochelais_2016.png",

    # ── USC Carcassonne ───────────────────────────────────────────────
    "carcassonne":          "839px-Logo_US_Carcassonne_XV.png",
    "usc carcassonne":      "839px-Logo_US_Carcassonne_XV.png",

    # ── Oyonnax Rugby ─────────────────────────────────────────────────
    "oyonnax":              "899px-Logo_Oyonnax_rugby_2018.png",

    # ── CA Brive Corrèze Limousin ─────────────────────────────────────
    "brive":                "Club_athlétique_Brive_Corrèze_Limousin_logo.svg.png",
    "corrèze":              "Club_athlétique_Brive_Corrèze_Limousin_logo.svg.png",

    # ── CAP Périgueux ─────────────────────────────────────────────────
    "périgueux":            "Logo_CA_Périgueux_2014.png",
    "perigueux":            "Logo_CA_Périgueux_2014.png",
    "cap périgueux":        "Logo_CA_Périgueux_2014.png",

    # ── US Colomiers Rugby ────────────────────────────────────────────
    "colomiers":            "Logo_Colomiers_Rugby_2008.png",

    # ── FC Grenoble Rugby ─────────────────────────────────────────────
    "grenoble":             "Logo_FC_Grenoble_Rugby.png",

    # ── RC Suresnes ───────────────────────────────────────────────────
    "suresnes":             "Logo_RC_Suresnes_-_2018.svg.png",
    "rc suresnes":          "Logo_RC_Suresnes_-_2018.svg.png",

    # ── RC Vannes ─────────────────────────────────────────────────────
    "vannes":               "Logo_RC_Vannes_-_2013.png",
    "rc vannes":            "Logo_RC_Vannes_-_2013.png",

    # ── Rouen Normandie Rugby ─────────────────────────────────────────
    "rouen":                "Logo_Rouen_Normandie_Rugby_-_2019.png",

    # ── SC Albi ───────────────────────────────────────────────────────
    "albi":                 "Logo_SC_Albi_-_Ancien.svg.png",
    "sca":                  "Logo_SC_Albi_-_Ancien.svg.png",

    # ── Section Paloise (Pau) ─────────────────────────────────────────
    "pau":                  "Logo_Section_Paloise.jpeg",
    "section paloise":      "Logo_Section_Paloise.jpeg",

    # ── SOC Rugby — Savoie Mont Blanc (Chambéry) ──────────────────────
    "soc rugby":            "Logo_Stade_Olympique_Chambéry_Rugby_-_2019.svg.png",
    "chambéry":             "Logo_Stade_Olympique_Chambéry_Rugby_-_2019.svg.png",
    "chambery":             "Logo_Stade_Olympique_Chambéry_Rugby_-_2019.svg.png",
    "savoie mont blanc":    "Logo_Stade_Olympique_Chambéry_Rugby_-_2019.svg.png",

    # ── US Dax Rugby Landes ───────────────────────────────────────────
    "dax":                  "Logo_US_Dax_Rugby_Landes_-_2018.svg.png",
    "us dax":               "Logo_US_Dax_Rugby_Landes_-_2018.svg.png",

    # ── US Montauban ──────────────────────────────────────────────────
    "montauban":            "Logo_US_Montalbanaise_-_2017.png",
    "montalbanaise":        "Logo_US_Montalbanaise_-_2017.png",

    # ── Valence Romans Drôme Rugby ────────────────────────────────────
    "valence romans":       "Logo_Valence_Romans_Drôme_Rugby_-_2019.svg.png",
    "valence":              "Logo_Valence_Romans_Drôme_Rugby_-_2019.svg.png",

    # ── LOU Rugby Lyon ────────────────────────────────────────────────
    "lyon":                 "Lyon_olympique_universitaire.png",
    "lou":                  "Lyon_olympique_universitaire.png",

    # ── Provence Rugby ────────────────────────────────────────────────
    "provence":             "Provence.jpg",

    # ── SU Agen ───────────────────────────────────────────────────────
    "agen":                 "SU_Agen.png",
    "su agen":              "SU_Agen.png",

    # ── Stade Toulousain ──────────────────────────────────────────────
    "toulouse":             "Stade_toulousain.png",
    "stade toulousain":     "Stade_toulousain.png",

    # ── Gloucester Rugby ─────────────────────────────────────────────
    "gloucester":           "gloucester.png",

    # ── SUA L&G — Sporting Union Agenais ─────────────────────────────
    "sua":                  "cG9dHaFgtr-MisGG9pCnrdbz-qQ.png",
    "sua l&g":              "cG9dHaFgtr-MisGG9pCnrdbz-qQ.png",
    "sporting union agenais": "cG9dHaFgtr-MisGG9pCnrdbz-qQ.png",
}


def _local_logo_url(filename: str) -> str:
    """
    Charge un fichier logo depuis le dossier logos/ local et le retourne
    sous forme d'une data URI base64 (compatible avec les balises <img>).
    """
    path = os.path.join(_LOCAL_LOGOS_DIR, filename)
    if not os.path.isfile(path):
        return ""
    ext = os.path.splitext(filename)[1].lower()
    mime = {
        ".svg": "image/svg+xml",
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
    }.get(ext, "image/png")
    with open(path, "rb") as f:
        data = base64.b64encode(f.read()).decode("utf-8")
    return f"data:{mime};base64,{data}"


# ── Logos fixés manuellement (fichier Wikimedia Commons / fr.wikipedia) ──────
# Format : "mot-clé du nom du club en minuscules" → "nom du fichier Wikimedia"
# Cherche sur Commons puis sur fr.wikipedia via _file_url()
_LOGO_OVERRIDES = {

    # ── Stade Aurillacois ──────────────────────────────────────────────
    "aurillacois":          "Logo_Stade_aurillacois_Cantal_Auvergne_2018.svg",
    "stade aurillacois":    "Logo_Stade_aurillacois_Cantal_Auvergne_2018.svg",

    # ── ASBH — AS Béziers Hérault ─────────────────────────────────────
    "asbh":                 "Logo_ASBH_Rugby.svg",
    "béziers":              "Logo_ASBH_Rugby.svg",
    "beziers":              "Logo_ASBH_Rugby.svg",

    # ── CAP Rugby Périgueux Dordogne ───────────────────────────────────
    "périgueux":            "Logo_CA_Périgueux_rugby.png",
    "perigueux":            "Logo_CA_Périgueux_rugby.png",
    "cap périgueux":        "Logo_CA_Périgueux_rugby.png",

    # ── SOC Rugby — Savoie Mont Blanc (Chambéry) ───────────────────────
    "soc rugby":            "Logo_SOC_Rugby.svg",
    "chambéry":             "Logo_SOC_Rugby.svg",
    "chambery":             "Logo_SOC_Rugby.svg",
    "savoie mont blanc":    "Logo_SOC_Rugby.svg",

    # ── CSBJ — CS Bourgoin-Jallieu ────────────────────────────────────
    "csbj":                 "Logo_CS_Bourgoin-Jallieu_Rugby.svg",
    "bourgoin":             "Logo_CS_Bourgoin-Jallieu_Rugby.svg",

    # ── Stade Niçois Rugby (logo local via _LOCAL_OVERRIDES) ─────────

    # ── USAP — Union Sportive Arlequins Perpignanais ──────────────────
    "usap":                 "Logo_USAP.svg",
    "perpignan":            "Logo_USAP.svg",

    # ── Stade Rochelais ───────────────────────────────────────────────
    "rochelais":            "Logo_Stade_rochelais_2021.svg",
    "la rochelle":          "Logo_Stade_rochelais_2021.svg",

    # ── USC XV Carcassonne ────────────────────────────────────────────
    "carcassonne":          "Logo_USC_Carcassonne_XV.svg",
    "usc carcassonne":      "Logo_USC_Carcassonne_XV.svg",

    # ── Oyonnax Rugby ─────────────────────────────────────────────────
    "oyonnax":              "Logo_USO_Oyonnax.svg",

    # ── Gloucester Rugby (club anglais) ───────────────────────────────
    "gloucester":           "Gloucester_Rugby_logo.svg",

    # ── RCHCC — RC Hyères-Carqueiranne-La Crau ────────────────────────
    "rchcc":                "Logo_RC_Hyeres-Carqueiranne-La_Crau.png",
    "hyères":               "Logo_RC_Hyeres-Carqueiranne-La_Crau.png",
    "hyeres":               "Logo_RC_Hyeres-Carqueiranne-La_Crau.png",

    # ── Top 14 / autres Pro D2 ────────────────────────────────────────
    "toulouse":             "Logo_Stade_toulousain.svg",
    "stade toulousain":     "Logo_Stade_toulousain.svg",
    "bordeaux":             "Logo_Union_Bordeaux_Bègle.svg",
    "agen":                 "Logo_SU_Agen_Rugby.svg",
    "su agen":              "Logo_SU_Agen_Rugby.svg",
    "biarritz":             "Logo_Biarritz_olympique.svg",
    "grenoble":             "Logo_FC_Grenoble_Rugby.svg",
    "vannes":               "Logo_RC_Vannes.svg",
    "rc vannes":            "Logo_RC_Vannes.svg",
    "stade montois":        "Logo_Stade_montois_rugby.svg",
    "colomiers":            "Logo_US_Colomiers_rugby.svg",
    "nevers":               "Logo_US_Nevers_rugby.svg",
    "albi":                 "Logo_SC_Albi_rugby.svg",
    "narbonne":             "Logo_RC_Narbonne.svg",
    "montauban":            "Logo_US_Montauban_rugby.svg",
    "rouen":                "Logo_Rouen_Normandie_Rugby.svg",
    "provence":             "Logo_Provence_Rugby.svg",
    "valence romans":       "Logo_Valence_Romans_Drôme_Rugby.svg",
    "massy":                "Logo_Massy_Essonne_Rugby.svg",
    "toulon":               "Logo_RC_Toulon.svg",
    "clermont":             "Logo_ASM_Clermont_Auvergne.svg",
    "asm":                  "Logo_ASM_Clermont_Auvergne.svg",
    "racing 92":            "Logo_Racing_92.svg",
    "montpellier":          "Logo_Montpellier_HR.svg",
    "castres":              "Logo_CO_Castres.svg",
    "bayonne":              "Logo_Aviron_Bayonnais_Rugby_Pro.svg",
    "aviron bayonnais":     "Logo_Aviron_Bayonnais_Rugby_Pro.svg",
    "lyon":                 "Logo_LOU_Rugby.svg",
    "lou":                  "Logo_LOU_Rugby.svg",

    # ── UBB — Union Bordeaux Bègles ───────────────────────────────────
    "ubb":                  "Logo_Union_Bordeaux_Bègle.svg",
    "union bordeaux":       "Logo_Union_Bordeaux_Bègle.svg",
    "bègles":               "Logo_Union_Bordeaux_Bègle.svg",
    "begles":               "Logo_Union_Bordeaux_Bègle.svg",

    # ── RCNM — RC Narbonne Méditerranée ──────────────────────────────
    "rcnm":                 "Logo_RC_Narbonne.svg",
    "narbonne méditerranée":"Logo_RC_Narbonne.svg",
    "narbonne mediterranee":"Logo_RC_Narbonne.svg",

    # ── Stade Français Paris ──────────────────────────────────────────
    "stade français":       "Logo_Stade_français_Paris.svg",
    "stade francais":       "Logo_Stade_français_Paris.svg",
    "sfp":                  "Logo_Stade_français_Paris.svg",

    # ── SA XV Charente Rugby (Angoulême) ──────────────────────────────
    "sa xv":                "Logo_SA_XV_Charente_Rugby.svg",
    "charente":             "Logo_SA_XV_Charente_Rugby.svg",
    "angoulême":            "Logo_SA_XV_Charente_Rugby.svg",
    "angouleme":            "Logo_SA_XV_Charente_Rugby.svg",
    "sa xv charente":       "Logo_SA_XV_Charente_Rugby.svg",

    # ── USBPA — US Bourg-en-Bresse Péronnas Athlétisme ───────────────
    "usbpa":                "Logo_USBPA_Rugby.svg",
    "bourg-en-bresse":      "Logo_USBPA_Rugby.svg",
    "bourg en bresse":      "Logo_USBPA_Rugby.svg",

    # ── USON Nevers ───────────────────────────────────────────────────
    "uson":                 "Logo_US_Nevers_rugby.svg",
    "uson nevers":          "Logo_US_Nevers_rugby.svg",

    # ── SUA L&G — Sporting Union Agenais (Lot-et-Garonne, 1908) ─────
    "sua":                  "Logo_SUA_Rugby.png",
    "sua l&g":              "Logo_SUA_Rugby.png",
    "sporting union agenais": "Logo_SUA_Rugby.png",

    # ── Pau Section Paloise ───────────────────────────────────────────
    "pau":                  "Logo_Section_paloise.svg",
    "section paloise":      "Logo_Section_paloise.svg",

    # ── Brive CA ──────────────────────────────────────────────────────
    "brive":                "Logo_CA_Brive.svg",
    "corrèze":              "Logo_CA_Brive.svg",

    # ── Aurillac (alias court) ────────────────────────────────────────
    "auri":                 "Logo_Stade_aurillacois_Cantal_Auvergne_2018.svg",

    # ── US Dax Rugby Landes ───────────────────────────────────────────
    "dax":                  "Logo_US_Dax_rugby_landes.svg",
    "us dax":               "Logo_US_Dax_rugby_landes.svg",

    # ── RC Suresnes ───────────────────────────────────────────────────
    "suresnes":             "Logo_Rugby_Club_Suresnes.svg",
    "rc suresnes":          "Logo_Rugby_Club_Suresnes.svg",

    # ── SC Albi (alias SCA) ───────────────────────────────────────────
    "sca":                  "Logo_SC_Albi_rugby.svg",
}


def _name_variants(team_name: str) -> list:
    """Génère les variantes du nom d'un club pour la recherche."""
    raw = team_name.strip()
    prefixes = ["Stade ", "SU ", "US ", "RC ", "SC ", "FC ",
                "ASM ", "LOU ", "UBB ", "CA ", "SA "]
    variants = [raw]
    for p in prefixes:
        if raw.startswith(p):
            variants.append(raw[len(p):])
    variants += [
        " ".join(raw.split()[:2]),   # 2 premiers mots
        " ".join(raw.split()[:3]),   # 3 premiers mots
        raw.split()[-1],             # dernier mot = souvent la ville
    ]
    return list(dict.fromkeys(v.strip() for v in variants if v.strip() and len(v.strip()) > 2))


@st.cache_data(show_spinner=False, ttl=86400)
def _get_team_logo_network(team_name: str) -> str:
    """Cherche le logo via Wikimedia/Wikipedia uniquement (mis en cache 24h)."""
    variants = _name_variants(team_name)
    key = team_name.strip().lower()

    # ── Override manuel Wikimedia ─────────────────────────────────────
    for pattern, filename in _LOGO_OVERRIDES.items():
        if pattern in key or key in pattern:
            try:
                u = _file_url(filename)
                if u:
                    return u
            except Exception:
                pass

    # ── 1. Wikimedia Commons : recherche directe de fichiers logo ────
    for v in variants[:4]:
        for q in (f"logo {v} rugby", f"blason {v} rugby", f"logo {v}"):
            try:
                u = _logo_from_commons_search(q)
                if u:
                    return u
            except Exception:
                continue

    # ── 2 & 3. Page Wikipedia directe (fr puis en) ──────────────────
    for lang in ("fr", "en"):
        for v in variants[:5]:
            for suffix in ("", " rugby"):
                try:
                    u = _logo_from_wiki_page(f"{v}{suffix}", lang)
                    if u:
                        return u
                except Exception:
                    continue

    # ── 4. Recherche full-text Wikipedia FR ─────────────────────────
    for v in variants[:3]:
        try:
            u = _logo_from_wiki_search(f"{v} rugby", "fr")
            if u:
                return u
        except Exception:
            continue

    return ""


def get_team_logo(team_name: str) -> str:
    """
    Point d'entrée public.
    1. Vérifie d'abord le dossier logos/ local (PAS de cache → toujours à jour).
    2. Si rien en local, appelle _get_team_logo_network() qui est mise en cache 24h.
    """
    # ── Local en priorité absolue (pas de cache, juste une lecture fichier) ──
    key = team_name.strip().lower()
    for pattern, filename in _LOCAL_OVERRIDES.items():
        if pattern in key or key in pattern:
            u = _local_logo_url(filename)
            if u:
                return u
    # ── Réseau en fallback (résultat mis en cache 24h) ──────────────────────
    return _get_team_logo_network(team_name)


# ─────────────────────────────────────────────
#  TERRAIN DE RUGBY — ZONE D'ACTIVITÉ JOUEUR
# ─────────────────────────────────────────────

# Mapping action → zone typique sur le terrain (x: 0‑100m, y: 0‑70m)
# x=0 = en-but propre, x=100 = en-but adverse
_ACTION_ZONES = {
    # En-but adverse
    "Essais":                       (93, 35),
    "ESSAI":                        (93, 35),
    "TRANS":                        (93, 35),
    # 22m adverse
    "Franchissements":              (82, 35),
    "Defenseurs battus":            (76, 35),
    "Contacts":                     (72, 35),
    "Soutiens Offensifs":           (66, 35),
    "Maul dans le jeu courant":     (62, 35),
    # Milieu de terrain (côté attaque)
    "Rucks":                        (56, 35),
    "RUCK":                         (56, 35),
    "Jeux a la main":               (55, 35),
    "Passes":                       (52, 35),
    "Bras casses":                  (52, 35),
    "Ballons perdus":               (50, 35),
    # Milieu de terrain (côté défense)
    "Plaquages":                    (44, 35),
    "Assistant plaqueur":           (42, 35),
    "Contest":                      (46, 35),
    "Contre Ruck":                  (38, 35),
    "Turn over et contre attaque":  (34, 35),
    # Propre 22m
    "Jeux au pied":                 (28, 35),
    "Botteur":                      (28, 35),
    "Buteur":                       (24, 35),
    "Penalites":                    (30, 35),
    "PENALITE":                     (30, 35),
    "Receptions aeriennes":         (22, 35),
    "Renvois ligne de but":         (8,  35),
    # Set-pieces (touches sur les côtés)
    "Touches":                      (50, 5),
    "TOUCHE":                       (50, 5),
    "Lancements sur touches":       (50, 5),
    # Set-pieces (mêlées au centre)
    "Melees":                       (50, 35),
    "MELEE":                        (50, 35),
    "Lancements sur melees":        (50, 35),
    # Coups d'envoi
    "Coups d'envoi":                (50, 35),
    # Zones textuelles (si présentes comme lignes)
    "22m adverse":                  (89, 35),
    "22m - 50m adverse":            (64, 35),
    "22m - 50m":                    (36, 35),
    "22m":                          (11, 35),
}


def _draw_rugby_field() -> go.Figure:
    """Crée le fond de terrain de rugby avec les lignes réglementaires (Plotly)."""
    fig = go.Figure()

    W, H = 100, 70          # largeur et hauteur du terrain en mètres
    IG = 10                 # profondeur des en-buts

    def rect(x0, y0, x1, y1, fill, line_color="white", lw=1, dash="solid"):
        fig.add_shape(type="rect", x0=x0, y0=y0, x1=x1, y1=y1,
                      fillcolor=fill, line=dict(color=line_color, width=lw, dash=dash))

    def vline(x, dash="solid", lw=1.5, color="white"):
        fig.add_shape(type="line", x0=x, y0=0, x1=x, y1=H,
                      line=dict(color=color, width=lw, dash=dash))

    # Pelouse principale
    rect(0, 0, W, H, "#2d6a1f")
    # En-buts (plus foncés)
    rect(0,    0, IG,    H, "#1e4a14")
    rect(W-IG, 0, W,     H, "#1e4a14")
    # Lignes de but
    vline(IG,   lw=2)
    vline(W-IG, lw=2)
    # Lignes des 22m
    vline(IG + 22, dash="dash")
    vline(W - IG - 22, dash="dash")
    # Ligne des 10m (depuis chaque ligne de but)
    vline(IG + 10, dash="dot", lw=1, color="rgba(255,255,255,0.4)")
    vline(W - IG - 10, dash="dot", lw=1, color="rgba(255,255,255,0.4)")
    # Ligne médiane
    vline(W / 2, lw=2.5)
    # Bords du terrain
    rect(0, 0, W, H, "rgba(0,0,0,0)", line_color="white", lw=2)

    # Poteaux (symboles sur les lignes de but)
    for gx in [IG, W - IG]:
        cy = H / 2
        fig.add_shape(type="line", x0=gx, y0=cy - 5.6, x1=gx, y1=cy + 5.6,
                      line=dict(color="#ffe066", width=3))

    # Étiquettes de zones
    for label, xpos in [("EN-BUT", IG / 2), ("22m", IG + 11), ("DÉFENSE", IG + 28),
                         ("50m", W / 2), ("ATTAQUE", W - IG - 28),
                         ("22m ADV", W - IG - 11), ("EN-BUT", W - IG / 2)]:
        fig.add_annotation(x=xpos, y=H - 3, text=label,
                           showarrow=False, font=dict(color="rgba(255,255,255,0.4)", size=9),
                           xanchor="center")

    fig.update_layout(
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="#1a1f2e",
        font=dict(color="#e2e8f0"),
        xaxis=dict(range=[-2, W + 2], showgrid=False, zeroline=False,
                   showticklabels=False, fixedrange=True),
        yaxis=dict(range=[-2, H + 2], showgrid=False, zeroline=False,
                   showticklabels=False, fixedrange=True, scaleanchor="x", scaleratio=1),
        margin=dict(l=10, r=10, t=40, b=10),
        height=420,
        showlegend=True,
        legend=dict(bgcolor="#1a1f2e", bordercolor="#4a5568",
                    borderwidth=1, font=dict(size=11)),
    )
    return fig


def rugby_field_activity(player_df: pd.DataFrame, player_name: str) -> go.Figure:
    """
    Génère le terrain de rugby avec les bulles de zones d'activité du joueur.

    Stratégie de détection (par priorité) :
      1. Colonnes X / Y numériques dans le CSV  → coordonnées réelles
      2. Colonne Zone / Localisation textuelle  → zones prédéfinies
      3. Mapping par nom d'action (KNOWN_ACTIONS) → zones typiques
    """
    fig = _draw_rugby_field()

    # ── Détection des colonnes de position ──────────────────────────
    x_col = y_col = zone_col = None
    for col in player_df.columns:
        cl = col.strip().lower()
        if cl in ("x", "pos_x", "coord_x", "position_x"):
            x_col = col
        if cl in ("y", "pos_y", "coord_y", "position_y"):
            y_col = col
        if any(k in cl for k in ("zone", "localisation", "secteur", "terrain")):
            zone_col = col

    # ── Cas 1 : coordonnées X/Y disponibles ────────────────────────
    if x_col and y_col:
        coords = player_df[[x_col, y_col]].replace("nan", pd.NA).dropna()
        coords = coords.apply(pd.to_numeric, errors="coerce").dropna()
        if not coords.empty:
            fig.add_trace(go.Histogram2dContour(
                x=coords[x_col], y=coords[y_col],
                colorscale=[[0, "rgba(255,200,0,0)"], [1, "rgba(255,120,0,0.85)"]],
                ncontours=12, showscale=False, name="Densité",
                contours=dict(coloring="fill"),
            ))
            fig.add_trace(go.Scatter(
                x=coords[x_col], y=coords[y_col],
                mode="markers",
                marker=dict(color="#ffd700", size=6, opacity=0.6,
                            line=dict(color="white", width=0.5)),
                name=player_name,
            ))
            fig.update_layout(title=f"Zone d'activité — {player_name}")
            return fig

    # ── Cas 2 : colonne Zone textuelle ─────────────────────────────
    zone_points: dict[str, list] = {}   # zone_label → [x, y, count]

    if zone_col:
        zone_vals = player_df[zone_col].replace("nan", pd.NA).dropna().astype(str)
        for zone_val in zone_vals:
            key = zone_val.strip()
            xy = _ACTION_ZONES.get(key)
            if xy:
                zone_points.setdefault(key, [xy[0], xy[1], 0])
                zone_points[key][2] += 1

    # ── Cas 3 : mapping par nom d'action ───────────────────────────
    if not zone_points:
        action_col = "Nom de la ligne"
        if action_col in player_df.columns:
            for raw in player_df[action_col].replace("nan", pd.NA).dropna().astype(str):
                # cherche une correspondance dans _ACTION_ZONES
                for key, xy in _ACTION_ZONES.items():
                    if key.lower() in raw.lower():
                        zone_points.setdefault(key, [xy[0], xy[1], 0])
                        zone_points[key][2] += 1
                        break

        # Fallback sur la colonne "code"
        if not zone_points and "code" in player_df.columns:
            for raw in player_df["code"].replace("nan", pd.NA).dropna().astype(str):
                for key, xy in _ACTION_ZONES.items():
                    if key.lower() in raw.lower():
                        zone_points.setdefault(key, [xy[0], xy[1], 0])
                        zone_points[key][2] += 1
                        break

    # ── Rendu des bulles ───────────────────────────────────────────
    if zone_points:
        total = sum(v[2] for v in zone_points.values())
        xs, ys, sizes, texts, colors = [], [], [], [], []
        for label, (x, y, cnt) in zone_points.items():
            xs.append(x)
            ys.append(y)
            sizes.append(max(18, min(60, cnt / max(total, 1) * 400)))
            pct = cnt / total * 100 if total > 0 else 0
            texts.append(f"<b>{label}</b><br>{cnt} actions<br>{pct:.0f}%")
            # couleur selon zone : défense (bleu) → attaque (rouge/orange)
            t = x / 100.0
            r = int(50 + 200 * t)
            g = int(150 - 80 * t)
            b = int(230 - 180 * t)
            colors.append(f"rgba({r},{g},{b},0.82)")

        fig.add_trace(go.Scatter(
            x=xs, y=ys, mode="markers+text",
            marker=dict(size=sizes, color=colors,
                        line=dict(color="white", width=1.5)),
            text=[str(v[2]) for v in zone_points.values()],
            textposition="middle center",
            textfont=dict(color="white", size=11, family="Arial Black"),
            customdata=texts,
            hovertemplate="%{customdata}<extra></extra>",
            name="Actions",
        ))
    else:
        fig.add_annotation(
            x=50, y=35, text="Aucune donnée de zone disponible",
            showarrow=False, font=dict(color="#a0aec0", size=14),
        )

    fig.update_layout(title=f"Zone d'activité — {player_name}")
    return fig


def team_logo_html(team_name: str, color: str, label: str) -> str:
    """Retourne le bloc HTML logo : image si trouvée, sinon initiales stylisées."""
    # SA toujours bleu, adversaire toujours rouge — peu importe le match ou l'ordre
    border_color = "#1d4ed8" if _is_aurillacois(team_name) else "#ef4444"
    logo_url = get_team_logo(team_name)
    if logo_url:
        logo_block = f'<img src="{logo_url}" style="height:70px;width:auto;object-fit:contain;margin-bottom:6px;" />'
    else:
        initials = "".join(w[0].upper() for w in team_name.split()[:3] if w)
        logo_block = (
            f'<div style="width:64px;height:64px;border-radius:50%;'
            f'background:{border_color}22;border:2px solid {border_color};'
            f'display:flex;align-items:center;justify-content:center;'
            f'margin:0 auto 6px;font-size:1.1rem;font-weight:900;color:{border_color};">'
            f'{initials}</div>'
        )
    return (
        f'<div style="background:#000;border-radius:20px;padding:24px 16px;'
        f'text-align:center;min-height:140px;display:flex;flex-direction:column;'
        f'justify-content:center;align-items:center;border:2px solid {border_color};">'
        f'  {logo_block}'
        f'  <div style="font-size:0.7rem;color:#a0aec0;text-transform:uppercase;'
        f'letter-spacing:2px;margin-top:2px;">{label}</div>'
        f'  <div style="font-size:1.05rem;font-weight:900;color:#FFFFFF;'
        f'margin-top:6px;line-height:1.2;">{team_name}</div>'
        f'</div>'
    )


# ─────────────────────────────────────────────
#  DOSSIER DATA/ PAR DÉFAUT (Google Drive sync)
# ─────────────────────────────────────────────
_DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
os.makedirs(_DATA_DIR, exist_ok=True)

def _data_dir_files() -> dict:
    """Retourne les CSV présents dans le dossier data/, sans doublons.
    Les fichiers avec suffixe de hash (ex: 'ASBH-AURI-a1b2c3d4.csv')
    sont ignorés si la version sans hash existe déjà.
    Le nom affiché est nettoyé : 'Clip CSV' et suffixes retirés.
    """
    import re
    # 1. Collecter tous les fichiers
    all_files = sorted(glob.glob(os.path.join(_DATA_DIR, "*.csv")))

    # 2. Retirer le suffixe hash (-xxxxxxxx) pour trouver le nom "propre"
    _hash_re = re.compile(r'-[0-9a-f]{8}(?=\.csv$)', re.IGNORECASE)

    # Grouper par nom propre → garder le fichier sans hash si possible
    clean_to_fp = {}
    for fp in all_files:
        name = os.path.basename(fp)
        clean = _hash_re.sub("", name)          # nom sans hash
        if clean not in clean_to_fp:
            clean_to_fp[clean] = fp             # premier trouvé
        elif clean == name:                     # version sans hash → priorité
            clean_to_fp[clean] = fp

    # 3. Construire le dict affiché → chemin
    result = {}
    for clean_name, fp in sorted(clean_to_fp.items()):
        # Nom d'affichage : retirer " Clip CSV" et " CSV" pour plus de lisibilité
        display = re.sub(r'\s+Clip\s+CSV', '', clean_name, flags=re.IGNORECASE)
        display = re.sub(r'\s+CSV', '', display, flags=re.IGNORECASE)
        display = display.replace(".csv", "").strip()
        result[display] = fp
    return result


# ─────────────────────────────────────────────
#  SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("## Rugby Analytics")
    st.markdown("---")

    # Source des données
    _source = st.radio(
        "Source des données",
        ["Dossier data/ (Google Drive)", "Importer un fichier"],
        label_visibility="visible",
    )

    uploaded_files = []
    folder_path = ""

    if _source == "Importer un fichier":
        uploaded_files = st.file_uploader(
            "Charger des fichiers CSV",
            type=["csv"],
            accept_multiple_files=True,
            help="Fichiers d'analyse vidéo (format Dartfish / Stats Perform)",
        )
    else:
        # Liste déroulante des matchs disponibles
        _local_files = _data_dir_files()
        if _local_files:
            _match_names = list(_local_files.keys())
            _selected_match = st.selectbox(
                "🏉 Sélectionner un match",
                _match_names,
                key="sidebar_match_select",
            )
            st.session_state["selected_data_file"] = _local_files[_selected_match]
            st.markdown(f"<small style='color:#68d391'>✓ {len(_match_names)} match(s) disponible(s)</small>",
                        unsafe_allow_html=True)
        else:
            st.info("Aucun fichier dans data/. Lance `sync_gdrive.py` pour synchroniser.")

        # ── Bouton sync ──────────────────────────────────────
        if st.button("Synchroniser depuis Google Drive"):
            _sync_script = os.path.join(os.path.dirname(__file__), "sync_gdrive.py")
            if os.path.exists(_sync_script):
                import subprocess
                with st.spinner("Synchronisation en cours…"):
                    result = subprocess.run(
                        ["python3", _sync_script],
                        capture_output=True, text=True, timeout=120
                    )
                if result.returncode == 0:
                    st.success("Synchronisation réussie !")
                    st.rerun()
                else:
                    st.error(f"Erreur : {result.stderr[:300]}")
            else:
                st.warning("Script sync_gdrive.py introuvable.")

        st.markdown("---")

        # ── Upload vers Google Drive ──────────────────────────
        st.markdown("**Ajouter un fichier dans Drive**")
        _file_to_upload = st.file_uploader(
            "Sélectionner un CSV à envoyer",
            type=["csv"],
            accept_multiple_files=False,
            key="gdrive_uploader",
            help="Le fichier sera ajouté au dossier Google Drive partagé",
        )
        if _file_to_upload is not None:
            if st.button("Envoyer vers Google Drive"):
                try:
                    from googleapiclient.discovery import build
                    from googleapiclient.http import MediaInMemoryUpload
                    from google.oauth2.credentials import Credentials
                    from google.auth.transport.requests import Request

                    _token_path = os.path.join(os.path.dirname(__file__), "token.json")
                    _creds_path = os.path.join(os.path.dirname(__file__), "credentials.json")

                    if not os.path.exists(_token_path):
                        st.error("Authentification Google non configurée. Lance sync_gdrive.py une première fois pour générer token.json.")
                    else:
                        _creds = Credentials.from_authorized_user_file(
                            _token_path,
                            ["https://www.googleapis.com/auth/drive"]
                        )
                        if _creds.expired and _creds.refresh_token:
                            _creds.refresh(Request())

                        _drive = build("drive", "v3", credentials=_creds)

                        _file_bytes = _file_to_upload.read()
                        _media = MediaInMemoryUpload(
                            _file_bytes,
                            mimetype="text/csv",
                            resumable=False,
                        )
                        _meta = {
                            "name": _file_to_upload.name,
                            "parents": ["1fqHroE4sUsSVjgV6fkWA8uF54ksQ4mX-"],
                        }
                        with st.spinner(f"Envoi de {_file_to_upload.name}…"):
                            _result = _drive.files().create(
                                body=_meta,
                                media_body=_media,
                                fields="id, name",
                            ).execute()

                        # Sauvegarde aussi en local dans data/
                        _local_dest = os.path.join(_DATA_DIR, _file_to_upload.name)
                        with open(_local_dest, "wb") as _lf:
                            _lf.write(_file_bytes)

                        st.success(f"Fichier **{_result['name']}** envoyé dans Google Drive et disponible localement.")
                        st.rerun()

                except ImportError:
                    st.error("Dépendances manquantes : pip install google-auth google-auth-oauthlib google-api-python-client")
                except Exception as _e:
                    st.error(f"Erreur lors de l'envoi : {_e}")

        folder_path = _DATA_DIR

    st.markdown("---")

    # ── GPS Excel uploader ──────────────────────────────────────────────────
    st.markdown("**Données GPS (HUB DATAS)**")
    _gps_file = st.file_uploader(
        "Charger le fichier GPS (.xlsx)",
        type=["xlsx"],
        accept_multiple_files=False,
        key="gps_uploader",
        help="Fichier HUB DATAS exporté depuis Catapult",
    )
    if _gps_file is not None:
        import tempfile as _tmpmod
        # Ne recréer le fichier temporaire que si le fichier GPS a changé.
        # Sans cette garde, chaque changement de match dans la sidebar
        # génère un nouveau chemin temporaire → invalide le cache @st.cache_data
        # → recharge le fichier GPS inutilement.
        if st.session_state.get("gps_file_name") != _gps_file.name:
            _gps_tmp = _tmpmod.NamedTemporaryFile(delete=False, suffix=".xlsx")
            _gps_tmp.write(_gps_file.read())
            _gps_tmp.flush()
            st.session_state["gps_path"] = _gps_tmp.name
            st.session_state["gps_file_name"] = _gps_file.name
        st.success(f"GPS chargé : {_gps_file.name}")
    elif "gps_path" not in st.session_state:
        st.session_state["gps_path"] = None

    st.markdown("---")
    page = st.radio(
        "Navigation",
        ["Vue d'ensemble", "Comparaison équipes", "Analyse joueurs",
         "Rucks & Contacts", "Jeu au pied & Passes", "Par quart-temps",
         "Suivi de Saison", "Analyse GPS", "Fusion GPS×Vidéo"],
        label_visibility="collapsed",
    )
    st.markdown("---")
    st.markdown("<small style='color:#4a5568'>Rugby Analytics Dashboard v1.0</small>",
                unsafe_allow_html=True)


# ─────────────────────────────────────────────
#  LOAD FILES
# ─────────────────────────────────────────────
files_to_load = {}

if uploaded_files:
    import tempfile, shutil
    for f in uploaded_files:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".csv") as tmp:
            f.seek(0)
            shutil.copyfileobj(f, tmp)
            files_to_load[f.name] = tmp.name

elif folder_path and os.path.isdir(folder_path):
    # Si un match a été sélectionné via le menu déroulant, charger uniquement celui-là
    if "selected_data_file" in st.session_state and st.session_state["selected_data_file"]:
        _sel_fp = st.session_state["selected_data_file"]
        files_to_load[os.path.basename(_sel_fp)] = _sel_fp
    else:
        for fp in sorted(glob.glob(os.path.join(folder_path, "*.csv"))):
            files_to_load[os.path.basename(fp)] = fp

if not files_to_load:
    st.markdown("""
    <div style='text-align:center; padding:80px 0;'>
      <div style='font-size:4rem;'></div>
      <h2 style='color:#e2e8f0;'>Rugby Analytics Dashboard</h2>
      <p style='color:#a0aec0; max-width:500px; margin:12px auto;'>
        Les fichiers CSV sont automatiquement chargés depuis le dossier <strong>data/</strong>
        synchronisé avec Google Drive.<br>
        Tu peux aussi importer manuellement via le panneau latéral.
      </p>
      <p style='color:#4a5568; font-size:0.85rem;'>
        Compatible : encodage UTF-16-LE · Plusieurs matchs simultanés
      </p>
    </div>
    """, unsafe_allow_html=True)
    st.stop()


# ─────────────────────────────────────────────
#  MATCH SELECTOR
# ─────────────────────────────────────────────
match_name = list(files_to_load.keys())[0]
df = load_csv(files_to_load[match_name])
teams = detect_teams(df)

if len(teams) < 2:
    st.error("Impossible de détecter deux équipes dans ce fichier.")
    st.warning(f"Équipes trouvées : **{teams if teams else 'aucune'}**")
    with st.expander("Diagnostic — ouvrir pour voir le contenu du fichier"):
        st.write("**Top 20 valeurs de 'Nom de la ligne' :**")
        st.dataframe(df["Nom de la ligne"].value_counts().head(20).reset_index(),
                     use_container_width=True)
        mask = df["Nom de la ligne"].str.contains(" - ", na=False)
        exemples = df[mask]["Nom de la ligne"].unique()[:15]
        st.write("**Exemples avec ' - ' :**")
        for e in exemples:
            parts = e.split(" - ", 1)
            st.write(f"→ team=`{parts[0]}` (len={len(parts[0])}) | action=`{parts[1]}`")
    st.info("**Solution :** Arrêtez Streamlit (Ctrl+C) puis relancez avec `streamlit run app.py`")
    st.stop()

# ── Détection Aurillacois ─────────────────────────────────────────────────
def _is_aurillacois(name: str) -> bool:
    n = name.lower()
    return "aurillac" in n or n in ("sa", "stade aurillacois")

def _code_is_auri(code: str) -> bool:
    """Vérifie si un code de filename (ex: 'AURI', 'SA') représente Aurillacois."""
    c = code.upper().strip()
    return c in ("AURI", "SA") or "AURI" in c

# ── Ordre gauche/droite déterminé par le nom du fichier CSV ───────────────
# Format attendu : "EQUIPE1-EQUIPE2.csv" → EQUIPE1 = gauche, EQUIPE2 = droite
# Ex: AURI-BO → SA à gauche | PRO-AURI → SA à droite
_fname_base = os.path.splitext(match_name)[0]
_fname_parts = _fname_base.split("-")
_left_code  = _fname_parts[0].upper().strip() if len(_fname_parts) >= 1 else ""
_right_code = _fname_parts[1].upper().strip() if len(_fname_parts) >= 2 else ""

_auri_team = next((t for t in teams if _is_aurillacois(t)), None)
_adv_team  = next((t for t in teams if not _is_aurillacois(t)), None)

if len(teams) >= 2 and _auri_team and _adv_team:
    if _code_is_auri(_left_code):
        # AURI est à gauche dans le fichier → team_a = SA (gauche), team_b = adv (droite)
        team_a, team_b = _auri_team, _adv_team
    elif _code_is_auri(_right_code):
        # AURI est à droite dans le fichier → team_a = adv (gauche), team_b = SA (droite)
        team_a, team_b = _adv_team, _auri_team
    else:
        # Pas de code AURI détecté dans le nom → ordre alphabétique par défaut
        team_a, team_b = teams[0], teams[1]
elif len(teams) >= 2:
    team_a, team_b = teams[0], teams[1]
else:
    team_a = teams[0] if teams else "?"
    team_b = teams[0] if teams else "?"

stats_a = team_stats(df, team_a)
stats_b = team_stats(df, team_b)

COLOR_SA  = "#1d4ed8"   # Stade Aurillacois — toujours bleu
COLOR_ADV = "#ef4444"   # Adversaire        — toujours rouge

# Couleur de chaque équipe selon sa position dans le fichier CSV
# SA = bleu, adversaire = rouge, peu importe qui est team_a ou team_b
COLOR_TA = COLOR_SA  if _is_aurillacois(team_a) else COLOR_ADV
COLOR_TB = COLOR_SA  if _is_aurillacois(team_b) else COLOR_ADV

# Alias rétrocompatibles (utilisés dans bar_comparison, radar, kpi…)
COLOR_A = COLOR_TA
COLOR_B = COLOR_TB


# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────
def kpi(label, val_a, val_b, higher_is_better=True):
    # Couleurs basées sur l'identité de l'équipe : SA=bleu, ADV=rouge, peu importe la position
    col1, col2, col3 = st.columns([2, 1, 2])
    with col1:
        st.markdown(f"""
        <div class='kpi-card'>
          <div class='kpi-value' style='color:{COLOR_TA};'>{val_a}</div>
          <div class='kpi-label'>{team_a.replace("Stade ", "")}</div>
        </div>""", unsafe_allow_html=True)
    with col2:
        st.markdown(f"<div style='text-align:center;padding-top:28px;color:#a0aec0;font-size:0.8rem;font-weight:600;'>{label}</div>", unsafe_allow_html=True)
    with col3:
        st.markdown(f"""
        <div class='kpi-card'>
          <div class='kpi-value' style='color:{COLOR_TB};'>{val_b}</div>
          <div class='kpi-label'>{team_b.replace("Stade ", "")}</div>
        </div>""", unsafe_allow_html=True)


def bar_comparison(labels, values_a, values_b, title=""):
    # Calcule les pourcentages pour chaque paire (A vs B)
    def pct_texts(va_list, vb_list):
        texts_a, texts_b = [], []
        for va, vb in zip(va_list, vb_list):
            total = va + vb
            if total > 0:
                texts_a.append(f"{va}<br>({va/total*100:.0f}%)")
                texts_b.append(f"{vb}<br>({vb/total*100:.0f}%)")
            else:
                texts_a.append("0")
                texts_b.append("0")
        return texts_a, texts_b

    texts_a, texts_b = pct_texts(values_a, values_b)
    max_val = max(max(values_a, default=0), max(values_b, default=0), 1)

    fig = go.Figure()
    fig.add_trace(go.Bar(
        name=team_a, x=labels, y=values_a,
        marker_color=COLOR_A, marker_line_width=0,
        text=texts_a, textposition="outside",
        textfont=dict(size=11, color="#e2e8f0"),
        cliponaxis=False,
    ))
    fig.add_trace(go.Bar(
        name=team_b, x=labels, y=values_b,
        marker_color=COLOR_B, marker_line_width=0,
        text=texts_b, textposition="outside",
        textfont=dict(size=11, color="#e2e8f0"),
        cliponaxis=False,
    ))
    fig.update_layout(
        title=title, barmode="group",
        plot_bgcolor="#1a1f2e", paper_bgcolor="#1a1f2e",
        font=dict(color="#e2e8f0", size=12),
        legend=dict(bgcolor="#2d3748", bordercolor="#4a5568", borderwidth=1),
        margin=dict(l=40, r=20, t=60, b=40),
        xaxis=dict(gridcolor="#2d3748"),
        yaxis=dict(gridcolor="#2d3748", range=[0, max_val * 1.45]),
    )
    return fig


def radar_chart(categories, values_a, values_b):
    cats = categories + [categories[0]]
    # Normalize
    maxs = [max(a, b, 1) for a, b in zip(values_a, values_b)]
    norm_a = [v / m * 100 for v, m in zip(values_a, maxs)] + [values_a[0] / maxs[0] * 100]
    norm_b = [v / m * 100 for v, m in zip(values_b, maxs)] + [values_b[0] / maxs[0] * 100]

    fig = go.Figure()
    _fill_a = "rgba(29,78,216,0.15)"  if _is_aurillacois(team_a) else "rgba(239,68,68,0.15)"
    _fill_b = "rgba(29,78,216,0.15)"  if _is_aurillacois(team_b) else "rgba(239,68,68,0.15)"
    fig.add_trace(go.Scatterpolar(r=norm_a, theta=cats, fill="toself",
                                   name=team_a, line_color=COLOR_TA,
                                   fillcolor=_fill_a))
    fig.add_trace(go.Scatterpolar(r=norm_b, theta=cats, fill="toself",
                                   name=team_b, line_color=COLOR_TB,
                                   fillcolor=_fill_b))
    fig.update_layout(
        polar=dict(
            bgcolor="#1a1f2e",
            radialaxis=dict(visible=True, range=[0, 100], gridcolor="#2d3748",
                            tickcolor="#4a5568", tickfont=dict(color="#a0aec0", size=9)),
            angularaxis=dict(gridcolor="#2d3748", tickfont=dict(color="#e2e8f0", size=11)),
        ),
        plot_bgcolor="#1a1f2e", paper_bgcolor="#1a1f2e",
        font=dict(color="#e2e8f0"),
        legend=dict(bgcolor="#2d3748", bordercolor="#4a5568", borderwidth=1),
        margin=dict(l=60, r=60, t=40, b=40),
    )
    return fig


def pie_chart(labels, values, title, colors):
    fig = go.Figure(go.Pie(
        labels=labels, values=values,
        marker=dict(colors=colors, line=dict(color="#0e1117", width=2)),
        textinfo="label+percent",
        textfont=dict(color="#e2e8f0", size=12),
        hole=0.45,
    ))
    fig.update_layout(
        title=title, plot_bgcolor="#1a1f2e", paper_bgcolor="#1a1f2e",
        font=dict(color="#e2e8f0"),
        legend=dict(bgcolor="#2d3748", bordercolor="#4a5568"),
        margin=dict(l=20, r=20, t=40, b=20),
        showlegend=False,
    )
    return fig


def hbar(df_data, x_col, y_col, color, title):
    total = df_data[x_col].sum()
    texts = [
        f"{v} ({v/total*100:.0f}%)" if total > 0 else str(v)
        for v in df_data[x_col]
    ]
    fig = px.bar(df_data, x=x_col, y=y_col, orientation="h",
                 color_discrete_sequence=[color], title=title,
                 text=texts)
    fig.update_traces(textposition="outside", textfont=dict(size=11, color="#e2e8f0"),
                      cliponaxis=False)
    max_val = df_data[x_col].max() if not df_data.empty else 1
    fig.update_layout(
        plot_bgcolor="#1a1f2e", paper_bgcolor="#1a1f2e",
        font=dict(color="#e2e8f0"),
        yaxis=dict(autorange="reversed", gridcolor="#2d3748"),
        xaxis=dict(gridcolor="#2d3748", range=[0, max_val * 1.35]),
        margin=dict(l=20, r=80, t=40, b=20),
    )
    return fig


# ─────────────────────────────────────────────
#  MATCH BANNER — layout scoreboard
# ─────────────────────────────────────────────
sc_a          = compute_score(df, team_a)
sc_b          = compute_score(df, team_b)
score_a       = sc_a["score"]
score_b       = sc_b["score"]
essais_a      = sc_a["essais"]
essais_b      = sc_b["essais"]
transfo_a     = sc_a["transfo"]
transfo_b     = sc_b["transfo"]
penalite_a    = sc_a["penalite"]
penalite_b    = sc_b["penalite"]
drop_a        = sc_a["drop"]
drop_b        = sc_b["drop"]
faute_a       = get_stat(stats_a, "FAUTE")
faute_b       = get_stat(stats_b, "FAUTE")
total_actions = len(df)

# ── Score mi-temps (Q1 + Q2) ──────────────────────────────────────────
_periode_col  = "Periode de jeu"
_mt1_periods  = ["1er quart-temps", "2nd quart-temps"]
if _periode_col in df.columns:
    df_mt1   = df[df[_periode_col].isin(_mt1_periods)]
    sc_mt1_a = compute_score(df_mt1, team_a)["score"]
    sc_mt1_b = compute_score(df_mt1, team_b)["score"]
else:
    sc_mt1_a = sc_mt1_b = 0

col_logo_a, col_score, col_logo_b = st.columns([1, 2, 1])

with col_logo_a:
    st.markdown(team_logo_html(team_a, COLOR_A, " "), unsafe_allow_html=True)

with col_score:
    st.markdown(f"""
    <div style="
        background:#000; border-radius:20px; padding:24px 20px;
        text-align:center; border: 1px solid #FFFFFF;">
      <div style="font-size:3rem; font-weight:900; color:#FFFFFF; letter-spacing:4px;">
        <span style="color:#FFFFFF;">{score_a}</span>
        &nbsp;–&nbsp;
        <span style="color:#FFFFFF;">{score_b}</span>
      </div>
      <div style="font-size:1.5rem; color:#FFFFFF; font-weight:700; text-transform:uppercase;
                  letter-spacing:1.5px; margin-top:5px;">Mi-temps</div>
      <div style="font-size:1.2rem; font-weight:700; color:#a0aec0; margin-top:2px;">
        <span style="color:#FFFFFF;">{sc_mt1_a}</span>
        &nbsp;–&nbsp;
        <span style="color:#FFFFFF;">{sc_mt1_b}</span>
      </div>
      <div style="font-size:0.72rem; color:#4a5568; margin-top:6px;">
        {match_name} &nbsp;·&nbsp; {total_actions:,} actions
      </div>
    </div>
    """, unsafe_allow_html=True)

with col_logo_b:
    st.markdown(team_logo_html(team_b, COLOR_B, " "), unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)


# ══════════════════════════════════════════════
#  FRAGMENT — Tableau joueurs + carte terrain
# ══════════════════════════════════════════════
@st.fragment
def _player_map_fragment(label, action_key, df_match, t_a, t_b, c_a, c_b):
    """Tableau joueurs cliquables + carte terrain interactive.
    @st.fragment : les clics ne ferment PAS la popup parente."""

    _pk = f"fmap_{action_key}"   # préfixe unique par action

    def _find_col(df, candidates):
        for c in candidates:
            if c in df.columns:
                return c
        for c in df.columns:
            for cand in candidates:
                if cand.lower() in c.lower():
                    return c
        return None

    # ── Tableau par équipe ────────────────────────────────────────────────
    def _general_table(team, color):
        mask  = df_match["Nom de la ligne"].str.strip() == f"{team} - {action_key}"
        rows  = df_match[mask].copy()
        total = len(rows)
        st.markdown(
            f"<div style='color:{color};font-weight:800;font-size:0.85rem;"
            f"text-transform:uppercase;letter-spacing:1px;margin-bottom:6px;'>"
            f"{team} — {total} {label.lower()}</div>",
            unsafe_allow_html=True,
        )
        if rows.empty:
            st.info("Aucune action trouvée.")
            return
        player_col = _find_col(rows, ["Joueur", "joueur", "Nom", "nom", "Player",
                                       f"{action_key} - Joueur"])
        result_col = _find_col(rows, ["Resultat", "résultat", "Résultat", "Result"])
        if player_col:
            rows[player_col] = rows[player_col].astype(str).str.strip()
            rows = rows[~rows[player_col].isin(["nan", "", "None"])]
        if player_col and not rows.empty:
            tbl = rows.groupby(player_col).size().reset_index(name="Nombre")
            tbl = tbl.sort_values("Nombre", ascending=False).reset_index(drop=True)
            if result_col:
                pos = (rows[rows[result_col].astype(str).str.lower().str.strip() == "positif"]
                       .groupby(player_col).size().reset_index(name="Reussis"))
                tbl = tbl.merge(pos, on=player_col, how="left")
                tbl["Reussis"] = tbl["Reussis"].fillna(0).astype(int)
                tbl["Taux %"] = (tbl["Reussis"] / tbl["Nombre"] * 100).round(1).astype(str) + " %"
                _tot_nb  = tbl["Nombre"].sum()
                _tot_ok  = tbl["Reussis"].sum()
                _tot_pct = round(_tot_ok / _tot_nb * 100, 1) if _tot_nb > 0 else 0
                _pct_clr = "#68d391" if _tot_pct >= 70 else ("#fbd38d" if _tot_pct >= 50 else "#fc8181")
                c1, c2, c3 = st.columns(3)
                for _cc, _val, _lbl, _clr in [
                    (c1, _tot_nb, "TOTAL", "#63b3ed"),
                    (c2, _tot_ok, "RÉUSSIES", "#68d391"),
                    (c3, f"{_tot_pct} %", "TAUX", _pct_clr),
                ]:
                    with _cc:
                        st.markdown(
                            f"<div style='background:#111827;border-radius:10px;padding:12px 8px;"
                            f"text-align:center;border:1px solid #1f2937;margin-bottom:10px;'>"
                            f"<div style='font-size:0.6rem;color:#6b7280;text-transform:uppercase;"
                            f"letter-spacing:1px;margin-bottom:4px;'>{_lbl}</div>"
                            f"<div style='font-size:1.6rem;font-weight:900;color:{_clr};'>{_val}</div>"
                            f"</div>", unsafe_allow_html=True)
                _total_row = pd.DataFrame([{
                    player_col: "🏁 TOTAL ÉQUIPE",
                    "Nombre": _tot_nb, "Reussis": _tot_ok,
                    "Taux %": f"{_tot_pct} %",
                }])
                tbl = pd.concat([tbl, _total_row], ignore_index=True)
            tbl = tbl.rename(columns={player_col: "Joueur"})
            _has_map = bool(_find_col(df_match, ["Coordonnee X", "Coordonnée X"]))
            _th_cols = st.columns([3, 1.2, 1.2, 1.6])
            for _hc, _hl in zip(_th_cols, ["Joueur", "Nb", "Réussies", "Taux %"]):
                _hc.markdown(
                    f"<div style='font-size:0.72rem;color:#9ca3af;font-weight:700;"
                    f"text-transform:uppercase;padding:4px 0 2px;"
                    f"border-bottom:1px solid #374151;'>{_hl}</div>",
                    unsafe_allow_html=True)
            for _, _tr in tbl.iterrows():
                _jname    = str(_tr["Joueur"])
                _is_total = _jname.startswith("🏁")
                _rc = st.columns([3, 1.2, 1.2, 1.6])
                with _rc[0]:
                    if _is_total:
                        st.markdown(
                            f"<div style='background:#1f2937;padding:5px 4px;"
                            f"font-size:0.82rem;font-weight:800;color:#e2e8f0;"
                            f"border-top:1px solid #374151;'>{_jname}</div>",
                            unsafe_allow_html=True)
                    elif _has_map:
                        _is_sel = (st.session_state.get(f"{_pk}_player") == _jname and
                                   st.session_state.get(f"{_pk}_team") == team)
                        if st.button(
                            f"▶ {_jname}" if _is_sel else _jname,
                            key=f"{_pk}_{team}_{_jname}",
                            use_container_width=True,
                            type="primary" if _is_sel else "secondary"
                        ):
                            if _is_sel:
                                st.session_state[f"{_pk}_player"] = None
                                st.session_state[f"{_pk}_team"]   = None
                            else:
                                st.session_state[f"{_pk}_player"] = _jname
                                st.session_state[f"{_pk}_team"]   = team
                            # PAS de st.rerun() — le fragment se rafraîchit seul
                    else:
                        st.markdown(
                            f"<div style='padding:5px 4px;font-size:0.83rem;"
                            f"color:#e2e8f0;'>{_jname}</div>", unsafe_allow_html=True)
                for _ci, _ckey in enumerate(["Nombre", "Reussis", "Taux %"], 1):
                    _val = _tr.get(_ckey, "")
                    _clr = "#e2e8f0" if _is_total else "#cbd5e0"
                    if _ckey == "Taux %" and not _is_total:
                        try:
                            _pv = float(str(_val).replace("%", "").strip())
                            _clr = "#4ade80" if _pv >= 70 else ("#fbbf24" if _pv >= 50 else "#f87171")
                        except Exception:
                            pass
                    _rc[_ci].markdown(
                        f"<div style='padding:5px 2px;font-size:0.83rem;"
                        f"font-weight:{'800' if _is_total else 'normal'};color:{_clr};"
                        f"text-align:center;'>{_val}</div>", unsafe_allow_html=True)
            if _has_map:
                st.markdown(
                    "<div style='font-size:0.7rem;color:#4b5563;margin-top:4px;'>"
                    "🗺️ Cliquer sur un joueur pour afficher sa carte terrain</div>",
                    unsafe_allow_html=True)
        elif result_col:
            summary = rows[result_col].astype(str).str.strip().value_counts().reset_index()
            summary.columns = ["Résultat", "Nb"]
            st.dataframe(summary, hide_index=True, use_container_width=True)
        else:
            st.metric("Total actions", total)

    col_l, col_r = st.columns(2)
    with col_l:
        _general_table(t_a, c_a)
    with col_r:
        _general_table(t_b, c_b)

    # ── Carte terrain ──────────────────────────────────────────────────────
    _cx_col = _find_col(df_match, ["Coordonnee X", "Coordonnée X"])
    _cy_col = _find_col(df_match, ["Coordonnee Y", "Coordonnée Y"])
    if not (_cx_col and _cy_col):
        return

    st.markdown("---")
    st.markdown(
        "<div style='font-size:1rem;font-weight:800;color:#e2e8f0;margin:16px 0 12px;"
        "border-left:4px solid #48bb78;padding-left:10px;'>🗺️ CARTE TERRAIN</div>",
        unsafe_allow_html=True)

    _sel_player = st.session_state.get(f"{_pk}_player")
    _map_team   = st.session_state.get(f"{_pk}_team", t_a)

    if not _sel_player:
        st.markdown(
            "<div style='text-align:center;padding:32px 0;color:#6b7280;font-size:0.9rem;'>"
            "🗺️ Cliquez sur le nom d'un joueur dans le tableau pour afficher sa carte terrain"
            "</div>", unsafe_allow_html=True)
        return

    _map_dot_color = c_a if _map_team == t_a else c_b
    _map_rows_all  = df_match[
        df_match["Nom de la ligne"].str.strip() == f"{_map_team} - {action_key}"
    ].copy()
    _map_pcol = _find_col(_map_rows_all, ["Joueur", "joueur", "Player"])

    _close_col, _title_col = st.columns([1, 8])
    with _title_col:
        st.markdown(
            f"<div style='padding:8px 12px;background:#1f2937;border-radius:8px;"
            f"border-left:4px solid {_map_dot_color};font-weight:700;color:#e2e8f0;"
            f"font-size:0.95rem;'>👤 {_sel_player} — {_map_team}</div>",
            unsafe_allow_html=True)
    with _close_col:
        if st.button("✕", key=f"{_pk}_close", help="Fermer la carte"):
            st.session_state[f"{_pk}_player"] = None
            st.session_state[f"{_pk}_team"]   = None

    _plot_rows = _map_rows_all.copy()
    if _map_pcol:
        _map_rows_all[_map_pcol] = _map_rows_all[_map_pcol].astype(str).str.strip()
        _plot_rows = _map_rows_all[_map_rows_all[_map_pcol] == _sel_player].copy()

    # Filtres
    _filter_defs = [
        ("Résultat",     ["Resultat"]),
        ("Efficacité",   ["Efficacite"]),
        ("Hauteur",      ["Hauteur"]),
        ("Période",      ["Periode de jeu"]),
        ("Côté terrain", ["Cote terrain"]),
        ("Zone terrain", ["Zone terrain"]),
        ("Type défense", ["Type de defense"]),
        ("Enchainement", ["Enchainement"]),
    ]
    _avail = {}
    for _fn, _fc_cands in _filter_defs:
        _fc = _find_col(_plot_rows, _fc_cands)
        if _fc:
            _uv = [str(v).strip() for v in _plot_rows[_fc].dropna().unique()
                   if str(v).strip().lower() not in ("nan", "none", "", "nat")]
            if len(_uv) >= 2:
                _avail[_fn] = (_fc, sorted(_uv))

    if _avail:
        st.markdown("**🔍 Filtres**")
        _fcols = st.columns(min(len(_avail), 4))
        for _fi, (_fn, (_fc, _uv)) in enumerate(_avail.items()):
            with _fcols[_fi % len(_fcols)]:
                _sel = st.multiselect(_fn, _uv, default=_uv, key=f"{_pk}_mf_{_fn}_{_map_team}")
                if _sel:
                    _plot_rows = _plot_rows[_plot_rows[_fc].astype(str).str.strip().isin(_sel)]

    _vc1, _vc2, _vc3 = st.columns(3)
    with _vc1:
        _color_opts    = list(_avail.keys())
        _color_by_name = st.selectbox(
            "🎨 Colorier par", _color_opts,
            index=_color_opts.index("Résultat") if "Résultat" in _color_opts else 0,
            key=f"{_pk}_color_by") if _color_opts else None
        _color_fc = _avail[_color_by_name][0] if _color_by_name else None
    with _vc2:
        _period_fc = _avail.get("Période", (None,))[0]
        _animate = st.toggle("▶️ Animer par période", value=False, key=f"{_pk}_anim") if _period_fc else False
    with _vc3:
        _show_density = st.toggle("🔥 Densité", value=False, key=f"{_pk}_density")

    _px = pd.to_numeric(_plot_rows[_cy_col], errors="coerce")
    _py = pd.to_numeric(_plot_rows[_cx_col], errors="coerce")
    _vmask       = _px.notna() & _py.notna()
    _px          = _px[_vmask];  _py = _py[_vmask]
    _plot_valid  = _plot_rows[_vmask].copy()

    import plotly.graph_objects as _pgo
    _fig = _pgo.Figure()
    _fig.add_shape(type="rect", x0=0, y0=0, x1=100, y1=68,
                   fillcolor="#2d5016", line=dict(color="white", width=0), layer="below")
    for _ex0, _ex1 in [(0, 10), (90, 100)]:
        _fig.add_shape(type="rect", x0=_ex0, y0=0, x1=_ex1, y1=68,
                       fillcolor="#1e3a0c", line=dict(color="white", width=2), layer="below")
    for _lx, _lw, _dash in [(10,1.5,"solid"),(22,2,"solid"),(50,2.5,"solid"),
                              (78,2,"solid"),(88,1.5,"solid"),(5,1,"dot"),(95,1,"dot")]:
        _fig.add_shape(type="line", x0=_lx, y0=0, x1=_lx, y1=68,
                       line=dict(color="white", width=_lw, dash=_dash), layer="below")
    _fig.add_shape(type="line", x0=0, y0=34, x1=100, y1=34,
                   line=dict(color="rgba(255,255,255,0.2)", width=1, dash="dot"), layer="below")
    for _lx, _lbl in [(5,"EN-BUT"),(16,"10m"),(36,"22m→50"),(50,"50m"),
                       (64,"50→22m"),(84,"10m"),(95,"EN-BUT")]:
        _fig.add_annotation(x=_lx, y=66, text=_lbl, showarrow=False,
                            font=dict(color="rgba(255,255,255,0.35)", size=8))
    _fig.add_shape(type="rect", x0=0, y0=0, x1=100, y1=68,
                   fillcolor="rgba(0,0,0,0)", line=dict(color="white", width=2))

    # ── Logos des équipes de part et d'autre du terrain ───────────────
    # Sens du jeu : équipe sélectionnée défend à gauche, attaque vers la droite
    _opp_team    = t_b if _map_team == t_a else t_a
    try:
        _logo_own = get_team_logo(_map_team)
        _logo_opp = get_team_logo(_opp_team)
    except Exception:
        _logo_own, _logo_opp = "", ""
    if _logo_own:
        _fig.add_layout_image(dict(
            source=_logo_own, xref="x", yref="y",
            x=5, y=34, sizex=8, sizey=28,
            xanchor="center", yanchor="middle",
            layer="above", opacity=0.85))
        _fig.add_annotation(x=5, y=6, text=f"<b>{_map_team}</b>",
                            showarrow=False, xanchor="center",
                            font=dict(color="rgba(255,255,255,0.85)", size=9),
                            bgcolor="rgba(0,0,0,0.55)", borderpad=2)
    if _logo_opp:
        _fig.add_layout_image(dict(
            source=_logo_opp, xref="x", yref="y",
            x=95, y=34, sizex=8, sizey=28,
            xanchor="center", yanchor="middle",
            layer="above", opacity=0.85))
        _fig.add_annotation(x=95, y=6, text=f"<b>{_opp_team}</b>",
                            showarrow=False, xanchor="center",
                            font=dict(color="rgba(255,255,255,0.85)", size=9),
                            bgcolor="rgba(0,0,0,0.55)", borderpad=2)
    # Flèche du sens du jeu (équipe → adversaire)
    _fig.add_annotation(x=50, y=4, ax=35, ay=4,
                        xref="x", yref="y", axref="x", ayref="y",
                        text="", showarrow=True, arrowhead=3,
                        arrowsize=1.4, arrowwidth=2,
                        arrowcolor="rgba(255,255,255,0.55)")
    _fig.add_annotation(x=50, y=4, text="sens du jeu",
                        showarrow=False, xanchor="center",
                        font=dict(color="rgba(255,255,255,0.55)", size=8))

    _CMAP = {
        "Positif":"#68d391","Negatif":"#fc8181","Neutre":"#fbd38d",
        "EFFICACE":"#68d391","NON EFFICACE":"#fc8181",
        "Avance":"#68d391","Ligne":"#fbd38d","Subi":"#fc8181",
        "1er quart-temps":"#63b3ed","2nd quart-temps":"#f6ad55",
        "3e quart-temps":"#fc8181","4e quart-temps":"#b794f4",
        "Haut":"#63b3ed","Jambe":"#fbd38d","Moyen":"#68d391",
    }

    if _show_density and len(_px) > 0:
        _fig.add_trace(_pgo.Histogram2dContour(
            x=_px, y=_py,
            colorscale=[[0,"rgba(0,0,0,0)"],[0.3,"rgba(255,200,0,0.2)"],
                        [0.7,"rgba(255,100,0,0.5)"],[1,"rgba(220,0,0,0.8)"]],
            showscale=False, ncontours=12,
            line=dict(width=0), contours=dict(showlabels=False), hoverinfo="skip"))

    if not _animate and len(_px) > 0:
        if _color_fc and _color_fc in _plot_valid.columns:
            _cvals = _plot_valid[_color_fc].astype(str).str.strip()
            for _cat in sorted(_cvals.unique()):
                _m = _cvals == _cat
                _fig.add_trace(_pgo.Scatter(
                    x=_px[_m], y=_py[_m], mode="markers", name=_cat,
                    marker=dict(size=11, color=_CMAP.get(_cat,"#63b3ed"),
                                opacity=0.88, line=dict(color="white",width=1.2)),
                    hovertemplate=f"<b>{_cat}</b><br>X: %{{x:.0f}} | Y: %{{y:.0f}}<extra></extra>"))
        else:
            _fig.add_trace(_pgo.Scatter(
                x=_px, y=_py, mode="markers", name=_sel_player,
                marker=dict(size=11, color=_map_dot_color, opacity=0.85,
                            line=dict(color="white",width=1.2)),
                hovertemplate="X: %{x:.0f} | Y: %{y:.0f}<extra></extra>"))

    elif _animate and _period_fc and len(_px) > 0:
        import plotly.express as _pex
        _anim_df = _plot_valid.copy()
        _anim_df["_px"] = _px.values
        _anim_df["_py"] = _py.values
        _anim_df["_color"] = _anim_df[_color_fc].astype(str).str.strip() if _color_fc else _map_team
        _p_order = ["1er quart-temps","2nd quart-temps","3e quart-temps","4e quart-temps"]
        _anim_df[_period_fc] = pd.Categorical(_anim_df[_period_fc], categories=_p_order, ordered=True)
        _anim_df = _anim_df.sort_values(_period_fc)
        _fig2 = _pex.scatter(
            _anim_df, x="_px", y="_py", color="_color",
            animation_frame=_period_fc, color_discrete_map=_CMAP,
            range_x=[0,100], range_y=[0,68],
            labels={"_px":"","_py":"","_color": _color_by_name or ""},
            title=f"Animation — {label} de {_sel_player}")
        _fig2.update_traces(marker=dict(size=12,opacity=0.88,line=dict(color="white",width=1.2)))
        _fig2.update_layout(
            paper_bgcolor="#0e1117", plot_bgcolor="#2d5016", font=dict(color="#e2e8f0"),
            xaxis=dict(showgrid=False,zeroline=False,showticklabels=False,fixedrange=True),
            yaxis=dict(showgrid=False,zeroline=False,showticklabels=False,
                       fixedrange=True,scaleanchor="x",scaleratio=0.68),
            legend=dict(bgcolor="rgba(0,0,0,0.6)",bordercolor="#374151"),
            margin=dict(l=5,r=5,t=40,b=5), height=420)
        # Logos des deux équipes — sens du jeu
        if _logo_own:
            _fig2.add_layout_image(dict(
                source=_logo_own, xref="x", yref="y",
                x=5, y=34, sizex=8, sizey=28,
                xanchor="center", yanchor="middle",
                layer="above", opacity=0.85))
        if _logo_opp:
            _fig2.add_layout_image(dict(
                source=_logo_opp, xref="x", yref="y",
                x=95, y=34, sizex=8, sizey=28,
                xanchor="center", yanchor="middle",
                layer="above", opacity=0.85))
        st.plotly_chart(_fig2, use_container_width=True)
        return

    _n_pts = len(_px)
    _fig.update_layout(
        paper_bgcolor="#0e1117", plot_bgcolor="#2d5016", font=dict(color="#e2e8f0"),
        xaxis=dict(range=[0,100],showgrid=False,zeroline=False,showticklabels=False,fixedrange=True),
        yaxis=dict(range=[0,68],showgrid=False,zeroline=False,showticklabels=False,
                   fixedrange=True,scaleanchor="x",scaleratio=0.68),
        legend=dict(bgcolor="rgba(0,0,0,0.6)",bordercolor="#374151",
                    font=dict(color="#e2e8f0",size=11),
                    orientation="h",yanchor="bottom",y=1.01,xanchor="left",x=0),
        margin=dict(l=5,r=5,t=40,b=5), height=420,
        title=dict(text=f"📍 {_n_pts} actions — {_sel_player}",
                   font=dict(color="#e2e8f0",size=13), x=0.5),
        hovermode="closest")
    if _n_pts > 0:
        st.plotly_chart(_fig, use_container_width=True)
    else:
        st.info("Aucun point à afficher avec les filtres sélectionnés.")


# ══════════════════════════════════════════════
#  POPUP — Détail joueurs par KPI
# ══════════════════════════════════════════════
@st.dialog("Détail", width="large")
def _kpi_popup(label, action_key, df_match, t_a, t_b, c_a, c_b):
    """Popup de détail par joueur / résultat pour chaque carte KPI."""
    st.markdown(
        f"<div style='font-size:1.05rem;font-weight:900;color:#e2e8f0;"
        f"text-transform:uppercase;letter-spacing:2px;margin-bottom:16px;"
        f"border-left:4px solid #63b3ed;padding-left:10px;'>{label}</div>",
        unsafe_allow_html=True,
    )

    def _find_col(df, candidates):
        for c in candidates:
            if c in df.columns:
                return c
        for c in df.columns:
            for cand in candidates:
                if cand.lower() in c.lower():
                    return c
        return None

    # ══════════════════════════════════════════════════════════════════════
    #  CAS SPÉCIAL : RUCKS
    # ══════════════════════════════════════════════════════════════════════
    if action_key == "Rucks":

        _card_r = ("background:#111827;border-radius:10px;padding:14px 8px;"
                   "text-align:center;border:1px solid #1f2937;margin-bottom:8px;")
        _lbl_r  = ("font-size:0.75rem;color:#cbd5e1;text-transform:uppercase;"
                   "letter-spacing:1px;margin-bottom:4px;")
        _val_r  = "font-size:1.8rem;font-weight:900;"
        _sub_r  = "font-size:0.82rem;"

        def _ruck_kpi_block(rows_r, color):
            """Affiche GAGNÉ / PERDU / NEUTRE + vitesse de libération."""
            total_r = len(rows_r)
            if rows_r.empty or total_r == 0:
                st.info("Aucune donnée ruck.")
                return

            _res_col = _find_col(rows_r, ["Resultat", "résultat", "Résultat"])
            _vit_col = _find_col(rows_r, ["Vitesse de liberation", "Vitesse liberation"])
            _off_col = _find_col(rows_r, ["Nombre de joueurs offensifs consommes",
                                           "Nb joueurs offensifs"])
            _def_col = _find_col(rows_r, ["Nombre de joueurs defensifs consommes",
                                           "Nb joueurs defensifs"])

            # ── KPI : GAGNÉ / PERDU / NEUTRE ─────────────────────────
            _res_colors = {"Positif": "#3b82f6", "Negatif": "#ef4444", "Neutre": "#f1f5f9"}
            _res_labels = {"Positif": "GAGNÉ", "Negatif": "PERDU", "Neutre": "NEUTRE"}

            if _res_col:
                _res_s = rows_r[_res_col].fillna("").astype(str).str.strip()
                _res_counts = _res_s.value_counts()
                _res_order = [r for r in ["Positif", "Negatif", "Neutre"]
                              if _res_counts.get(r, 0) > 0]
                _rcols = st.columns(max(len(_res_order), 1))
                for _ci, _rv in enumerate(_res_order):
                    _cnt = int(_res_counts.get(_rv, 0))
                    _pct = round(_cnt / total_r * 100, 1)
                    _clr = _res_colors.get(_rv, "#cbd5e1")
                    _lbl = _res_labels.get(_rv, _rv)
                    with _rcols[_ci]:
                        st.markdown(
                            f"<div style='{_card_r}'>"
                            f"<div style='{_lbl_r}'>{_lbl}</div>"
                            f"<div style='{_val_r}color:{_clr};'>{_cnt}</div>"
                            f"<div style='{_sub_r}color:{_clr};'>{_pct} %</div>"
                            f"</div>", unsafe_allow_html=True)

            # ── Vitesse de libération ─────────────────────────────────
            if _vit_col:
                _vit_s = rows_r[_vit_col].fillna("Non renseigné").astype(str).str.strip()
                _vit_counts = _vit_s[_vit_s != "Non renseigné"].value_counts()
                _vit_order  = ["0 - 3 sec", "3 - 6 sec", "+ 6 sec", "Arret de l'action"]
                _vit_present = [v for v in _vit_order if _vit_counts.get(v, 0) > 0]
                _vit_clrs    = ["#3b82f6", "#93c5fd", "#ef4444", "#f1f5f9"]

                if _vit_present:
                    st.markdown(
                        "<div style='font-size:0.82rem;color:#cbd5e1;text-transform:uppercase;"
                        "letter-spacing:1px;margin:14px 0 6px;'>Vitesse de libération</div>",
                        unsafe_allow_html=True)
                    _vcols = st.columns(len(_vit_present))
                    for _vi, _vv in enumerate(_vit_present):
                        _vc = int(_vit_counts.get(_vv, 0))
                        _vp = round(_vc / total_r * 100, 1)
                        _vclr = _vit_clrs[_vi % len(_vit_clrs)]
                        with _vcols[_vi]:
                            st.markdown(
                                f"<div style='{_card_r}'>"
                                f"<div style='{_lbl_r}'>{_vv}</div>"
                                f"<div style='{_val_r}color:{_vclr};'>{_vc}</div>"
                                f"<div style='{_sub_r}color:{_vclr};'>{_vp} %</div>"
                                f"</div>", unsafe_allow_html=True)

        # ── Récupérer lignes SA Rucks ─────────────────────────────────────
        _ndl = df_match["Nom de la ligne"].astype(str).str.strip()

        # Aurillacois : "Stade Aurillacois - Rucks" (données pleines)
        _sa_ruck_patterns = {
            f"{t_a} - Rucks", f"{t_a} - Ruck", f"{t_a} - rucks",
            "SA - RUCK", "SA - Rucks",
        }
        _rows_sa = df_match[_ndl.isin(_sa_ruck_patterns)].dropna(subset=["Resultat"]) \
            if "Resultat" in df_match.columns else df_match[_ndl.isin(_sa_ruck_patterns)]

        # Adversaire : lignes ADV
        _adv_ruck_patterns = {
            f"{t_b} - Rucks", f"{t_b} - Ruck", f"{t_b} - rucks",
            "ADV - RUCK", "ADV - Rucks",
        }
        _rows_adv = df_match[_ndl.isin(_adv_ruck_patterns)].dropna(subset=["Resultat"]) \
            if "Resultat" in df_match.columns else df_match[_ndl.isin(_adv_ruck_patterns)]

        # ── Section 2 colonnes : KPIs des deux équipes ────────────────────
        _col_sa, _col_adv = st.columns(2)
        with _col_sa:
            st.markdown(
                f"<div style='color:{c_a};font-weight:800;font-size:0.9rem;"
                f"text-transform:uppercase;letter-spacing:1px;margin-bottom:10px;'>"
                f"{t_a} — {len(_rows_sa)} RUCK(S)</div>",
                unsafe_allow_html=True)
            _ruck_kpi_block(_rows_sa, c_a)

        with _col_adv:
            st.markdown(
                f"<div style='color:{c_b};font-weight:800;font-size:0.9rem;"
                f"text-transform:uppercase;letter-spacing:1px;margin-bottom:10px;'>"
                f"{t_b} — {len(_rows_adv)} RUCK(S)</div>",
                unsafe_allow_html=True)
            _ruck_kpi_block(_rows_adv, c_b)

        # ── Carte terrain : zones d'activité SA ───────────────────────────
        _cx_col = _find_col(df_match, ["Coordonnee X", "Coordonnée X"])
        _cy_col = _find_col(df_match, ["Coordonnee Y", "Coordonnée Y"])
        _res_col_m = _find_col(_rows_sa, ["Resultat"])
        _jou_col_m = _find_col(_rows_sa, ["Joueur"])

        if _cx_col and _cy_col and not _rows_sa.empty:
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown(
                f"<div style='font-size:0.95rem;font-weight:700;color:#f1f5f9;"
                f"border-left:3px solid {c_a};padding-left:10px;margin-bottom:12px;'>"
                f"🗺 Zone d'activité des rucks — {t_a}</div>",
                unsafe_allow_html=True)

            # ── Filtre joueur ─────────────────────────────────────────
            _all_players = []
            if _jou_col_m and _jou_col_m in _rows_sa.columns:
                _all_players = sorted(
                    _rows_sa[_jou_col_m].dropna().astype(str).str.strip()
                    .replace("", pd.NA).dropna().unique().tolist()
                )
            _sel_players = st.multiselect(
                "Filtrer par joueur(s)",
                options=_all_players,
                default=[],
                placeholder="Tous les joueurs",
                key="ruck_map_player_filter",
            )

            # Coordonnées : Y→axe horizontal (longueur), X→axe vertical (largeur)
            _map_rows = _rows_sa.copy()
            _map_rows["_px"] = pd.to_numeric(_map_rows[_cy_col], errors="coerce")
            _map_rows["_py"] = pd.to_numeric(_map_rows[_cx_col], errors="coerce")
            _map_rows = _map_rows.dropna(subset=["_px", "_py"])

            # Appliquer le filtre joueur si sélection
            if _sel_players and _jou_col_m and _jou_col_m in _map_rows.columns:
                _map_rows = _map_rows[
                    _map_rows[_jou_col_m].astype(str).str.strip().isin(_sel_players)
                ]

            if not _map_rows.empty:
                # Couleur selon résultat
                _res_col_vals = _map_rows[_res_col_m].fillna("").astype(str).str.strip() \
                    if _res_col_m else pd.Series([""] * len(_map_rows))
                _pt_colors = _res_col_vals.map(
                    {"Positif": "#3b82f6", "Negatif": "#ef4444", "Neutre": "#f1f5f9"}
                ).fillna("#cbd5e1").tolist()

                _hover_texts = []
                for _, _row in _map_rows.iterrows():
                    _j = str(_row[_jou_col_m]) if _jou_col_m else "—"
                    _rv = str(_row[_res_col_m]).strip() if _res_col_m else "—"
                    _vit = _find_col(_rows_sa, ["Vitesse de liberation"])
                    _v  = str(_row[_vit]).strip() if _vit and _vit in _row.index else "—"
                    _hover_texts.append(f"<b>{_j}</b><br>Résultat : {_rv}<br>Vitesse : {_v}")

                import plotly.graph_objects as _pgo
                _fig_r = _pgo.Figure()

                # Terrain de base
                _fig_r.add_shape(type="rect", x0=0, y0=0, x1=100, y1=68,
                                 fillcolor="#2d5016", line=dict(color="white", width=0), layer="below")
                for _ex0, _ex1 in [(0, 10), (90, 100)]:
                    _fig_r.add_shape(type="rect", x0=_ex0, y0=0, x1=_ex1, y1=68,
                                     fillcolor="#1e3a0c", line=dict(color="white", width=2), layer="below")
                for _lx in [10, 22, 50, 78, 90]:
                    _fig_r.add_shape(type="line", x0=_lx, y0=0, x1=_lx, y1=68,
                                     line=dict(color="white", width=1.5), layer="below")
                _fig_r.add_shape(type="rect", x0=0, y0=0, x1=100, y1=68,
                                 fillcolor="rgba(0,0,0,0)", line=dict(color="white", width=2))
                for _lx, _lbl in [(5,"EN-BUT"),(16,"10m"),(36,"22m→50"),
                                   (50,"50m"),(64,"50→22m"),(84,"10m"),(95,"EN-BUT")]:
                    _fig_r.add_annotation(x=_lx, y=65, text=_lbl, showarrow=False,
                                          font=dict(color="rgba(255,255,255,0.4)", size=9))

                # ── Logos des équipes sur le terrain ─────────────────
                try:
                    _logo_sa  = get_team_logo(t_a)
                    _logo_adv = get_team_logo(t_b)
                except Exception:
                    _logo_sa, _logo_adv = "", ""

                if _logo_sa:
                    _fig_r.add_layout_image(dict(
                        source=_logo_sa, xref="x", yref="y",
                        x=5, y=34, sizex=10, sizey=30,
                        xanchor="center", yanchor="middle",
                        layer="above", opacity=0.85))
                _fig_r.add_annotation(
                    x=5, y=5, text=f"<b>{t_a}</b>",
                    showarrow=False, xanchor="center",
                    font=dict(color="rgba(255,255,255,0.9)", size=9),
                    bgcolor="rgba(0,0,0,0.55)", borderpad=2)

                if _logo_adv:
                    _fig_r.add_layout_image(dict(
                        source=_logo_adv, xref="x", yref="y",
                        x=95, y=34, sizex=10, sizey=30,
                        xanchor="center", yanchor="middle",
                        layer="above", opacity=0.85))
                _fig_r.add_annotation(
                    x=95, y=5, text=f"<b>{t_b}</b>",
                    showarrow=False, xanchor="center",
                    font=dict(color="rgba(255,255,255,0.9)", size=9),
                    bgcolor="rgba(0,0,0,0.55)", borderpad=2)

                # Flèche sens du jeu (SA attaque vers la droite)
                _fig_r.add_annotation(
                    x=58, y=4, ax=42, ay=4,
                    xref="x", yref="y", axref="x", ayref="y",
                    text="", showarrow=True, arrowhead=3,
                    arrowsize=1.4, arrowwidth=2,
                    arrowcolor="rgba(255,255,255,0.5)")
                _fig_r.add_annotation(
                    x=50, y=4, text="sens du jeu →",
                    showarrow=False, xanchor="center",
                    font=dict(color="rgba(255,255,255,0.5)", size=8))

                # Points par résultat
                for _rv, _rclr, _rlbl in [("Positif","#3b82f6","Gagné"),
                                            ("Negatif","#ef4444","Perdu"),
                                            ("Neutre", "#f1f5f9","Neutre")]:
                    _mask_rv = _res_col_vals == _rv if _res_col_m else pd.Series([False]*len(_map_rows))
                    _sub = _map_rows[_mask_rv.values]
                    if _sub.empty:
                        continue
                    _sub_hover = [_hover_texts[i] for i in _sub.index
                                  if i < len(_hover_texts)] if len(_hover_texts) == len(_map_rows) else []
                    _fig_r.add_trace(_pgo.Scatter(
                        x=_sub["_px"], y=_sub["_py"],
                        mode="markers",
                        name=_rlbl,
                        marker=dict(color=_rclr, size=11, opacity=0.85,
                                    line=dict(color="white", width=1.2)),
                        hovertemplate="%{customdata}<extra></extra>",
                        customdata=[_hover_texts[i] for i in range(len(_map_rows))
                                    if list(_map_rows.index)[i] in list(_sub.index)]
                                   if _sub_hover else None,
                        showlegend=True,
                    ))

                _fig_r.update_layout(
                    template="plotly_dark",
                    paper_bgcolor="#1a1f2e", plot_bgcolor="rgba(0,0,0,0)",
                    height=380,
                    xaxis=dict(range=[0, 100], showgrid=False, zeroline=False,
                               showticklabels=False, fixedrange=True),
                    yaxis=dict(range=[0, 68], showgrid=False, zeroline=False,
                               showticklabels=False, scaleanchor="x", scaleratio=0.68,
                               fixedrange=True),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02,
                                xanchor="left", x=0, font=dict(size=12)),
                    margin=dict(l=10, r=10, t=40, b=10),
                )
                st.plotly_chart(_fig_r, use_container_width=True)
            else:
                st.info("Pas de coordonnées X/Y disponibles pour ce match.")

        return

    # ══════════════════════════════════════════════════════════════════════
    #  CAS SPÉCIAL : MÊLÉES
    # ══════════════════════════════════════════════════════════════════════
    if action_key in ("Melees", "MELEE"):
        def _melee_team(team, color):
            # ── Construire la liste exhaustive des noms de lignes pour cette équipe ──
            _is_auri = "aurillac" in team.lower()

            # On n'utilise QUE les formes abrégées (SA - MELEE / ADV - MELEE).
            # La ligne "{Nom complet équipe} - Melees" est volontairement exclue
            # car elle peut dupliquer ou mélanger des données avec la ligne abrégée.
            if _is_auri:
                _melee_patterns = {"SA - MELEE", "SA - Melees", "SA - MELEES"}
            else:
                _melee_patterns = {"ADV - MELEE", "ADV - Melees", "ADV - MELEES"}

            # Masque : ligne dans les patterns ET pas "lancement" dans le nom
            _ndl = df_match["Nom de la ligne"].astype(str).str.strip()
            mask = _ndl.isin(_melee_patterns) & ~_ndl.str.lower().str.contains("lancement", na=False)
            rows  = df_match[mask].copy()
            total = len(rows)

            st.markdown(
                f"<div style='color:{color};font-weight:800;font-size:0.88rem;"
                f"text-transform:uppercase;letter-spacing:1px;margin-bottom:12px;'>"
                f"{team} — {total} mêlée(s)</div>",
                unsafe_allow_html=True,
            )
            if rows.empty:
                st.info("Aucune mêlée trouvée.")
                return

            # ── Mêlées à refaire ────────────────────────────────────────
            _refaire_col = _find_col(rows, ["Melee a refaire", "Mêlée à refaire"])
            _nb_refaire = 0
            if _refaire_col:
                _nb_refaire = rows[_refaire_col].notna().sum()
                _nb_refaire = int(rows[_refaire_col].dropna().astype(str).str.strip()
                                  .str.lower().ne("").sum())

            # ── Efficacité depuis colonne EFFICACITE (uppercase) ────────
            _eff_col = _find_col(rows, ["EFFICACITE"])

            def _parse_efficacite(v):
                try:
                    s = str(v).strip().upper()
                except Exception:
                    return "Non renseigné"
                if s in ("NAN", "", "NONE", "NAT"):
                    return "Non renseigné"
                # NON EFFICACE en premier (sinon "EFFICACE" matcherait aussi)
                if s.startswith("NON EFFICACE"):
                    return "Non efficace"
                if s.startswith("EFFICACE"):
                    return "Efficace"
                if "NEUTRE" in s:
                    return "Neutre"
                return "Autre"

            # ── Styles partagés ────────────────────────────────────────
            _card_style = (
                "background:#111827;border-radius:10px;padding:14px 8px;"
                "text-align:center;border:1px solid #1f2937;margin-bottom:10px;"
            )
            _lbl_style = "font-size:0.75rem;color:#cbd5e1;text-transform:uppercase;letter-spacing:1px;margin-bottom:4px;"
            _val_style = "font-size:1.6rem;font-weight:900;"
            _sub_style = "font-size:0.8rem;"

            # ── Cartes KPI : TOTAL et À REFAIRE ───────────────────────
            _kc1, _kc2, _spacer = st.columns([1, 1, 1])
            with _kc1:
                st.markdown(
                    f"<div style='{_card_style}'>"
                    f"<div style='{_lbl_style}'>TOTAL</div>"
                    f"<div style='{_val_style}color:{color};'>{total}</div>"
                    f"</div>", unsafe_allow_html=True)
            with _kc2:
                st.markdown(
                    f"<div style='{_card_style}'>"
                    f"<div style='{_lbl_style}'>À REFAIRE</div>"
                    f"<div style='{_val_style}color:#ef4444;'>{_nb_refaire}</div>"
                    f"<div style='{_sub_style}color:#ef4444;'>"
                    f"{round(_nb_refaire/total*100,1) if total else 0} %</div>"
                    f"</div>", unsafe_allow_html=True)

            # ── Résultats EFFICACITÉ (sans "Non renseigné") ────────────
            if _eff_col:
                rows["_eff_cat"] = rows[_eff_col].apply(_parse_efficacite)
                _eff_order  = ["Efficace", "Non efficace", "Neutre"]   # Non renseigné exclu
                _eff_colors = {
                    "Efficace":     "#3b82f6",
                    "Non efficace": "#ef4444",
                    "Neutre":       "#f1f5f9",
                }
                _eff_counts = rows["_eff_cat"].value_counts()
                _eff_cats   = [c for c in _eff_order if _eff_counts.get(c, 0) > 0]

                st.markdown(
                    "<div style='font-size:0.8rem;color:#cbd5e1;text-transform:uppercase;"
                    "letter-spacing:1px;margin:14px 0 6px;'>Résultats (EFFICACITÉ)</div>",
                    unsafe_allow_html=True)
                _rcols = st.columns(max(len(_eff_cats), 1))
                for _ci, _cat in enumerate(_eff_cats):
                    _cnt = int(_eff_counts.get(_cat, 0))
                    _pct = round(_cnt / total * 100, 1)
                    _clr = _eff_colors.get(_cat, "#cbd5e1")
                    with _rcols[_ci]:
                        st.markdown(
                            f"<div style='{_card_style}'>"
                            f"<div style='{_lbl_style}'>{_cat.upper()}</div>"
                            f"<div style='{_val_style}color:{_clr};'>{_cnt}</div>"
                            f"<div style='{_sub_style}color:{_clr};'>{_pct} %</div>"
                            f"</div>", unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # ── Tableau détaillé : LANCEMENT, LARGEUR TERRAIN, ZONE LANCEMENT, ZONE TERRAIN ──
            _tbl_cols_wanted = {
                "LANCEMENT":      _find_col(rows, ["LANCEMENT"]),
                "LARGEUR TERRAIN": _find_col(rows, ["LARGEUR TERRAIN"]),
                "ZONE LANCEMENT": _find_col(rows, ["ZONE LANCEMENT", "ZONE LANCE"]),
                "ZONE TERRAIN":   _find_col(rows, ["ZONE TERRAIN"]),
            }
            _tbl_cols_ok = {k: v for k, v in _tbl_cols_wanted.items() if v is not None}

            if _tbl_cols_ok:
                st.markdown(
                    "<div style='font-size:0.75rem;color:#9ca3af;text-transform:uppercase;"
                    "letter-spacing:1px;margin-bottom:6px;'>Détail des mêlées</div>",
                    unsafe_allow_html=True)
                _detail_tbl = rows[[v for v in _tbl_cols_ok.values()]].copy()
                _detail_tbl.columns = list(_tbl_cols_ok.keys())

                # Ajouter EFFICACITE si disponible
                if _eff_col:
                    _detail_tbl.insert(0, "EFFICACITÉ", rows["_eff_cat"].values)

                # Nettoyer
                for _c in _detail_tbl.columns:
                    _detail_tbl[_c] = (_detail_tbl[_c].astype(str).str.strip()
                                       .replace({"nan": "—", "None": "—", "": "—"}))

                # Supprimer les lignes entièrement vides
                _detail_tbl = _detail_tbl[
                    ~(_detail_tbl == "—").all(axis=1)
                ].reset_index(drop=True)

                st.dataframe(_detail_tbl, hide_index=True, use_container_width=True)

        col_l, col_r = st.columns(2)
        with col_l:
            _melee_team(t_a, c_a)
        with col_r:
            _melee_team(t_b, c_b)
        return


    # ══════════════════════════════════════════════════════════════════════
    #  CAS GÉNÉRAL : tableau joueurs + carte terrain (via fragment)
    # ══════════════════════════════════════════════════════════════════════
    _player_map_fragment(label, action_key, df_match, t_a, t_b, c_a, c_b)



# ══════════════════════════════════════════════
#  PAGE : VUE D'ENSEMBLE
# ══════════════════════════════════════════════
if page == "Vue d'ensemble":

    # ── Blocs stats clés style scoreboard ──
    col_stats_a, col_mid, col_stats_b = st.columns([2, 1, 2])

    def _stat_row(label, val_a, val_b, higher_is_better=True):
        if val_a + val_b == 0:
            pct_a = pct_b = 50
        else:
            pct_a = round(val_a / (val_a + val_b) * 100)
            pct_b = 100 - pct_a
        if higher_is_better:
            win_a = val_a >= val_b
        else:
            win_a = val_a <= val_b
        # Barres : couleur fixe et toujours visible — la largeur indique qui mène
        bar_a = COLOR_TA
        bar_b = COLOR_TB
        return (
            f"<div style='display:flex;justify-content:space-between;align-items:center;"
            f"padding:10px 0;border-bottom:1px solid #2d3748;'>"
            f"  <div style='font-size:2rem;font-weight:900;color:#FFFFFF;min-width:48px;text-align:center; margin-top:5px'>{val_a}</div>"
            f"  <div style='text-align:center;flex:1;'>"
            f"    <div style='font-size:1.6rem;color:#e2e8f0;font-weight:700;text-transform:uppercase;letter-spacing:1px;'>{label}</div>"
            f"    <div style='display:flex;height:15px;border-radius:3px;margin:2px 12px 0;overflow:hidden;'>"
            f"      <div style='flex:{pct_a};background:{bar_a};border-radius:3px 0 0 3px;'></div>"
            f"      <div style='flex:{pct_b};background:{bar_b};border-radius:0 3px 3px 0;'></div>"
            f"    </div>"
            f"  </div>"
            f"  <div style='font-size:2rem;font-weight:900;color:#FFFFFF;min-width:48px;text-align:center;'>{val_b}</div>"
            f"</div>"
        )

    SCORE_STATS = [
        ("ESSAI",          essais_a,   essais_b,   True),
        ("TRANSFORMATION", transfo_a,  transfo_b,  True),
        ("PÉNALITÉ",       penalite_a, penalite_b, True),
        ("DROP",           drop_a,     drop_b,     True),
        ("FAUTE",          faute_a,    faute_b,    False),
    ]

    rows_html = "".join(_stat_row(lbl, va, vb, hib) for lbl, va, vb, hib in SCORE_STATS)

    # ── BIP (Ball In Play) = somme des durées de toutes les séquences ──
    def _find_col_bip(dataframe, *candidates):
        for c in candidates:
            for col in dataframe.columns:
                if col.strip().lower() == c.strip().lower():
                    return col
        return None

    _seq_bip = df[df["Nom de la ligne"].str.strip().str.upper() == "SEQUENCE"].copy()
    _col_duree_bip = _find_col_bip(_seq_bip, "Durée", "Duree", "Durée ")
    if not _seq_bip.empty and _col_duree_bip:
        _bip_seconds = pd.to_numeric(_seq_bip[_col_duree_bip], errors="coerce").sum()
        _bip_min  = int(_bip_seconds // 60)
        _bip_sec  = int(_bip_seconds % 60)
        _bip_str  = f"{_bip_min}min {_bip_sec:02d}s"
        _bip_n    = len(_seq_bip.dropna(subset=[_col_duree_bip]))
    else:
        _bip_str  = "—"
        _bip_n    = 0

    _bip_row_html = f"""
    <div style="display:flex;justify-content:center;align-items:center;
                padding:12px 0 2px;border-top:1px solid #2d3748;margin-top:4px;">
      <div style="text-align:center;">
        <div style="font-size:1.6rem;font-weight:700;color:;text-transform:uppercase;
                    letter-spacing:1px;margin-bottom:4px;">BIP</div>
        <div style="font-size:1.5rem;font-weight:900;color:#FFFFFF;">{_bip_str}</div>
        <div style="font-size:0.9rem;color:#718096;font-weight:700;margin-top:2px;"> {_bip_n} séquences</div>
      </div>
    </div>"""

    st.markdown(f"""
    <div style="background:#11151E;border-radius:20px;padding:24px 28px;border:1px solid #204CD2;">
      <div style="display:flex;justify-content:space-between;margin-bottom:8px;">
        <div style="font-size:1.5rem;font-weight:800;color:#FFFFFF;text-transform:uppercase;">{team_a}</div>
        <div style="font-size:1.5rem;font-weight:800;color:#FFFFFF;text-transform:uppercase;">{team_b}</div>
      </div>
      {rows_html}
      {_bip_row_html}
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ═══════════════════════════════════════════════════════
    #  GRILLE KPI — 4 × 2 avec popup au clic
    # ═══════════════════════════════════════════════════════
    STATS_KEYS = [
        ("Plaquages",       True),
        ("Contacts",        True),
        ("Rucks",           True),
        ("Passes",          True),
        ("Franchissements", True),
        ("Ballons perdus",  False),
        ("Melees",          True),
        ("Touches",         True),
    ]

    def _kpi_card(label, va, vb, higher_is_better):
        # Barres : couleur fixe par équipe (SA=bleu, ADV=rouge)
        tot = va + vb
        pct_a = round(va / tot * 100) if tot else 50
        pct_b = 100 - pct_a
        return f"""<div style="background:#111827;border-radius:12px;padding:16px 14px 10px;
                    border:1px solid #1f2937;margin-bottom:4px;">
          <div style="font-size:1.2rem;font-weight:700; color:#FFFFFF;text-transform:uppercase;
                      letter-spacing:1.5px;text-align:center;margin-bottom:10px;">{label}</div>
          <div style="display:flex;justify-content:space-between;align-items:baseline;margin-bottom:10px;">
            <div style="font-size:1.8rem;font-weight:1000;color:#FFFFFF;line-height:1;">{va}</div>
            <div style="font-size:0.9rem;color:#FfffFF;font-weight:600;">vs</div>
            <div style="font-size:1.8rem;font-weight:900;color:#FFFFFF;line-height:1;">{vb}</div>
          </div>
          <div style="display:flex;height:4px;border-radius:2px;overflow:hidden;background:#1f2937;">
            <div style="flex:{pct_a};background:{COLOR_TA};"></div>
            <div style="flex:{pct_b};background:{COLOR_TB};"></div>
          </div>
        </div>"""

    # Ligne 1 : 4 premières KPIs
    _kpi_cols1 = st.columns(4)
    for _ki, (_key, _hib) in enumerate(STATS_KEYS[:4]):
        with _kpi_cols1[_ki]:
            _va = get_stat(stats_a, _key)
            _vb = get_stat(stats_b, _key)
            st.markdown(_kpi_card(_key, _va, _vb, _hib), unsafe_allow_html=True)
            if st.button("Voir détail", key=f"kpi_btn_{_key}", use_container_width=True):
                _kpi_popup(_key, _key, df, team_a, team_b, COLOR_TA, COLOR_TB)

    st.markdown("<div style='margin-bottom:6px'></div>", unsafe_allow_html=True)

    # Ligne 2 : 4 dernières KPIs
    _kpi_cols2 = st.columns(4)
    for _ki, (_key, _hib) in enumerate(STATS_KEYS[4:]):
        with _kpi_cols2[_ki]:
            _va = get_stat(stats_a, _key)
            _vb = get_stat(stats_b, _key)
            st.markdown(_kpi_card(_key, _va, _vb, _hib), unsafe_allow_html=True)
            if st.button("Voir détail", key=f"kpi_btn_{_key}", use_container_width=True):
                _kpi_popup(_key, _key, df, team_a, team_b, COLOR_TA, COLOR_TB)

    st.markdown("<div style='margin-bottom:20px'></div>", unsafe_allow_html=True)

    # ═══════════════════════════════════════════════════════
    #  RANGÉE 2 : Radar | Possession donut | Rucks donut
    # ═══════════════════════════════════════════════════════
    radar_cats = ["Essais", "Contacts", "Rucks", "Passes", "Franchissements",
                  "Jeux au pied", "Defenseurs battus", "Soutiens Offensifs"]
    vals_a = [get_stat(stats_a, k) for k in radar_cats]
    vals_b = [get_stat(stats_b, k) for k in radar_cats]

    col_radar, col_donuts = st.columns([11, 12])

    _CHART_H = 420  # hauteur commune radar et possession

    with col_radar:
        st.markdown("<div class='section-title'>Profil global des équipes</div>",
                    unsafe_allow_html=True)
        _fig_radar = radar_chart(radar_cats, vals_a, vals_b)
        _fig_radar.update_layout(height=_CHART_H)
        st.plotly_chart(_fig_radar, use_container_width=True)

    with col_donuts:
        # ── Possession ──────────────────────
        st.markdown("<div class='section-title'>Possession</div>", unsafe_allow_html=True)
        poss_a = get_stat(stats_a, "Possession")
        poss_b = get_stat(stats_b, "Possession")
        _fig_poss = pie_chart([team_a, team_b], [poss_a, poss_b], "Possession", [COLOR_TA, COLOR_TB])
        _fig_poss.update_layout(height=_CHART_H)
        st.plotly_chart(_fig_poss, use_container_width=True)


    # ═══════════════════════════════════════════════════════
    #  POSSESSION PAR QUART-TEMPS — pleine largeur
    # ═══════════════════════════════════════════════════════
    _p_order  = ["1er quart-temps", "2nd quart-temps", "3e quart-temps", "4e quart-temps"]
    _p_col    = "Periode de jeu"
    _p_labels = ["Q1", "Q2", "Q3", "Q4"]
    _poss_rows = []
    for _pi, _period in enumerate(_p_order):
        _pa = len(df[(df["Nom de la ligne"] == f"{team_a} - Possession") & (df[_p_col] == _period)])
        _pb = len(df[(df["Nom de la ligne"] == f"{team_b} - Possession") & (df[_p_col] == _period)])
        _tot = _pa + _pb
        _poss_rows.append({
            "Q": _p_labels[_pi],
            "pct_a": round(_pa / _tot * 100, 1) if _tot else 50.0,
            "pct_b": round(_pb / _tot * 100, 1) if _tot else 50.0,
        })

    _fig_poss_qt = go.Figure()
    _fig_poss_qt.add_trace(go.Bar(
        name=team_a,
        y=[r["Q"] for r in _poss_rows],
        x=[r["pct_a"] for r in _poss_rows],
        orientation="h",
        marker=dict(color=COLOR_A, line=dict(width=0)),
        text=[f"<b>{r['pct_a']:.0f}%</b>" for r in _poss_rows],
        textposition="inside",
        textfont=dict(color="#fff", size=16, family="Arial Black"),
        hovertemplate=f"<b>{team_a}</b><br>%{{y}} : %{{x:.1f}}%<extra></extra>",
    ))
    _fig_poss_qt.add_trace(go.Bar(
        name=team_b,
        y=[r["Q"] for r in _poss_rows],
        x=[r["pct_b"] for r in _poss_rows],
        orientation="h",
        marker=dict(color=COLOR_B, line=dict(width=0)),
        text=[f"<b>{r['pct_b']:.0f}%</b>" for r in _poss_rows],
        textposition="inside",
        textfont=dict(color="#fff", size=16, family="Arial Black"),
        hovertemplate=f"<b>{team_b}</b><br>%{{y}} : %{{x:.1f}}%<extra></extra>",
    ))
    _fig_poss_qt.update_layout(
        barmode="stack",
        paper_bgcolor="#0d1117",
        plot_bgcolor="#0d1117",
        height=265,
        margin=dict(l=50, r=30, t=50, b=20),
        title=dict(
            text="Possession par quart-temps",
            font=dict(size=14, color="#a0aec0"),
            x=0.5, xanchor="center",
        ),
        xaxis=dict(
            range=[0, 100], ticksuffix="%",
            tickfont=dict(size=11, color="#4a5568"),
            gridcolor="#1f2937", showline=False, zeroline=False,
        ),
        yaxis=dict(
            tickfont=dict(size=15, color="#e2e8f0", family="Arial Black"),
            autorange="reversed", showgrid=False,
        ),
        legend=dict(
            orientation="h", y=1.18, x=0.5, xanchor="center",
            font=dict(size=12, color="#e2e8f0"),
            bgcolor="rgba(0,0,0,0)",
        ),
        bargap=0.22,
    )
    _fig_poss_qt.add_vline(x=50, line_dash="dot", line_color="#374151", line_width=1.5)
    st.plotly_chart(_fig_poss_qt, use_container_width=True)

    # ═══════════════════════════════════════════════════════
    #  JEUX AU PIED | FRANCHISSEMENTS — pleine largeur, grands donuts
    # ═══════════════════════════════════════════════════════
    _d2a, _d2b = st.columns(2)
    with _d2a:
        st.markdown("<div class='section-title'>Jeux au pied</div>", unsafe_allow_html=True)
        jop_a = get_stat(stats_a, "Jeux au pied")
        jop_b = get_stat(stats_b, "Jeux au pied")
        _fig_jop = pie_chart([team_a, team_b], [jop_a, jop_b], "Jeux au pied", [COLOR_A, COLOR_B])
        _fig_jop.update_layout(height=460)
        st.plotly_chart(_fig_jop, use_container_width=True)
    with _d2b:
        st.markdown("<div class='section-title'>Franchissements</div>", unsafe_allow_html=True)
        fr_a = get_stat(stats_a, "Franchissements")
        fr_b = get_stat(stats_b, "Franchissements")
        _fig_fr = pie_chart([team_a, team_b], [fr_a, fr_b], "Franchissements", [COLOR_A, COLOR_B])
        _fig_fr.update_layout(height=460)
        st.plotly_chart(_fig_fr, use_container_width=True)

    # ═══════════════════════════════════════════════════════
    #  TABLEAU RÉCAPITULATIF (masqué par défaut)
    # ═══════════════════════════════════════════════════════
    with st.expander("Tableau récapitulatif complet", expanded=False):
        all_actions = sorted(set(list(stats_a.keys()) + list(stats_b.keys())))
        table_data = []
        for action in all_actions:
            va = get_stat(stats_a, action)
            vb = get_stat(stats_b, action)
            diff = va - vb
            table_data.append({
                "Action": action,
                team_a: va,
                team_b: vb,
                "Différence": f"+{diff}" if diff > 0 else str(diff),
                "Avantage": team_a if va > vb else (team_b if vb > va else "Égalité"),
            })
        df_table = pd.DataFrame(table_data)
        st.dataframe(df_table, use_container_width=True, hide_index=True,
                     column_config={
                         team_a: st.column_config.NumberColumn(team_a),
                         team_b: st.column_config.NumberColumn(team_b),
                     })

    st.markdown("---")

    # ── Cumul des séquences ──────────────────────────────────────────────
    st.markdown("<div class='section-title'>Cumul des séquences de jeu</div>",
                unsafe_allow_html=True)

    seq_df = df[df["Nom de la ligne"].str.strip().str.upper() == "SEQUENCE"].copy()

    # Detect column names (accent variants)
    def _find_col(dataframe, *candidates):
        for c in candidates:
            for col in dataframe.columns:
                if col.strip().lower() == c.strip().lower():
                    return col
        return None

    col_debut  = _find_col(seq_df, "Temps de début", "Temps de debut", "Début", "Debut")
    col_duree  = _find_col(seq_df, "Durée", "Duree", "Durée ")

    if seq_df.empty or col_debut is None or col_duree is None:
        st.info("Aucune donnée de séquence disponible dans ce fichier.")
    else:
        seq_df[col_debut] = pd.to_numeric(seq_df[col_debut], errors="coerce")
        seq_df[col_duree] = pd.to_numeric(seq_df[col_duree], errors="coerce")
        seq_df = seq_df.dropna(subset=[col_debut, col_duree])
        seq_df = seq_df.sort_values(col_debut).reset_index(drop=True)
        seq_df["Séquence #"] = seq_df.index + 1
        seq_df["Durée cumulée (s)"] = seq_df[col_duree].cumsum()

        total_seq  = len(seq_df)
        total_dur  = seq_df[col_duree].sum()
        avg_dur    = seq_df[col_duree].mean()
        max_dur    = seq_df[col_duree].max()

        # KPI row
        kpi_cols = st.columns(4)
        kpi_data = [
            ("Nb séquences",  f"{total_seq}"),
            ("Durée totale",  f"{int(total_dur // 60)}min {int(total_dur % 60)}s"),
            ("Durée moyenne", f"{avg_dur:.1f}s"),
            ("Plus longue",   f"{max_dur:.1f}s"),
        ]
        for col_kpi, (lbl, val) in zip(kpi_cols, kpi_data):
            with col_kpi:
                st.markdown(f"""
                <div style="background:#1a202c;border-radius:12px;padding:14px 10px;
                            text-align:center;border:1px solid #2d3748;">
                  <div style="font-size:0.75rem;color:#a0aec0;text-transform:uppercase;
                              letter-spacing:1px;margin-bottom:4px;">{lbl}</div>
                  <div style="font-size:1.6rem;font-weight:900;color:#63b3ed;">{val}</div>
                </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # Step chart — cumulative duration vs sequence number
        fig_seq = go.Figure()

        # Individual bar durations (secondary, light)
        fig_seq.add_trace(go.Bar(
            x=seq_df["Séquence #"],
            y=seq_df[col_duree],
            name="Durée séquence",
            marker_color="rgba(99,179,237,0.35)",
            yaxis="y2",
            hovertemplate="Séq %{x} — durée : %{y:.1f}s<extra></extra>",
        ))

        # Cumulative step line
        fig_seq.add_trace(go.Scatter(
            x=seq_df["Séquence #"],
            y=seq_df["Durée cumulée (s)"],
            mode="lines+markers",
            name="Durée cumulée",
            line=dict(color="#68d391", width=3, shape="hv"),
            marker=dict(size=5, color="#68d391"),
            hovertemplate="Séq %{x} — cumulé : %{y:.0f}s<extra></extra>",
        ))

        fig_seq.update_layout(
            template="plotly_dark",
            paper_bgcolor="#0d1117",
            plot_bgcolor="#0d1117",
            height=380,
            margin=dict(l=40, r=20, t=30, b=40),
            legend=dict(orientation="h", y=1.08, x=0),
            xaxis=dict(
                title="N° de séquence",
                showgrid=False,
                color="#e2e8f0",
                tickfont=dict(size=11),
            ),
            yaxis=dict(
                title="Durée cumulée (s)",
                color="#68d391",
                gridcolor="#2d3748",
                tickfont=dict(size=11),
            ),
            yaxis2=dict(
                title="Durée séquence (s)",
                overlaying="y",
                side="right",
                color="rgba(99,179,237,0.7)",
                showgrid=False,
                tickfont=dict(size=11),
            ),
        )
        st.plotly_chart(fig_seq, use_container_width=True)

        # Timeline strip — sequences on the match clock
        col_fin = _find_col(seq_df, "Temps de fin", "Temps de fin ", "Fin")
        if col_fin:
            seq_df[col_fin] = pd.to_numeric(seq_df[col_fin], errors="coerce")
            fig_tl = go.Figure()
            for _, row in seq_df.iterrows():
                d = float(row[col_duree]) if not pd.isna(row[col_duree]) else 1
                color_intensity = min(1.0, d / 120)  # 120s = full green
                r = int(104 * (1 - color_intensity) + 68 * color_intensity)
                g = int(211 * (1 - color_intensity) + 211 * color_intensity)
                b = int(145 * (1 - color_intensity) + 57 * color_intensity)
                fig_tl.add_trace(go.Bar(
                    x=[d],
                    base=[float(row[col_debut])],
                    y=["Séquences"],
                    orientation="h",
                    marker_color=f"rgba({r},{g},{b},0.85)",
                    showlegend=False,
                    hovertemplate=(
                        f"Séq {int(row['Séquence #'])}<br>"
                        f"Début : {row[col_debut]:.1f}s<br>"
                        f"Durée : {d:.1f}s<extra></extra>"
                    ),
                ))

            match_len = seq_df[col_fin].max() if col_fin else seq_df[col_debut].max() + 60
            fig_tl.update_layout(
                template="plotly_dark",
                paper_bgcolor="#0d1117",
                plot_bgcolor="#0d1117",
                height=110,
                margin=dict(l=40, r=20, t=10, b=30),
                barmode="overlay",
                xaxis=dict(
                    title="Temps de match (s)",
                    range=[0, match_len + 30],
                    color="#e2e8f0",
                    gridcolor="#2d3748",
                ),
                yaxis=dict(color="#e2e8f0", tickfont=dict(size=11)),
            )
            st.plotly_chart(fig_tl, use_container_width=True)


    # ═══════════════════════════════════════════════════════
    #  ANALYSE FATIGUE — 1re mi-temps vs 2e mi-temps
    # ═══════════════════════════════════════════════════════
    st.markdown("<div class='section-title'>Analyse Fatigue — 1re vs 2e mi-temps</div>",
                unsafe_allow_html=True)

    _fat_periods_1 = ["1er quart-temps", "2nd quart-temps"]
    _fat_periods_2 = ["3e quart-temps",  "4e quart-temps"]
    _fat_col       = "Periode de jeu"

    if _fat_col in df.columns:
        _df_h1 = df[df[_fat_col].isin(_fat_periods_1)]
        _df_h2 = df[df[_fat_col].isin(_fat_periods_2)]

        _fat_metrics = [
            ("Plaquages",       True),
            ("Rucks",           True),
            ("Contacts",        True),
            ("Franchissements", True),
            ("Ballons perdus",  False),
            ("Jeux au pied",    True),
        ]

        _fat_rows = []
        for _m, _hib in _fat_metrics:
            _prefix_a = f"{team_a} - "
            _h1_a = int(_df_h1["Nom de la ligne"].str.startswith(_prefix_a + _m).sum())
            _h2_a = int(_df_h2["Nom de la ligne"].str.startswith(_prefix_a + _m).sum())
            _fat_rows.append({"Métrique": _m, "Mi-temps": "1re mi-temps", "Valeur": _h1_a})
            _fat_rows.append({"Métrique": _m, "Mi-temps": "2e mi-temps",  "Valeur": _h2_a})

        _df_fat = pd.DataFrame(_fat_rows)

        _fig_fat = go.Figure()
        _m_list  = [m for m, _ in _fat_metrics]
        _h1_vals = [r["Valeur"] for r in _fat_rows if r["Mi-temps"] == "1re mi-temps"]
        _h2_vals = [r["Valeur"] for r in _fat_rows if r["Mi-temps"] == "2e mi-temps"]

        _fig_fat.add_trace(go.Bar(
            name="1re mi-temps", x=_m_list, y=_h1_vals,
            marker_color="#63b3ed",
            text=_h1_vals, textposition="outside",
            hovertemplate="%{x}<br>1re MT : %{y}<extra></extra>",
        ))
        _fig_fat.add_trace(go.Bar(
            name="2e mi-temps", x=_m_list, y=_h2_vals,
            marker_color="#f6ad55",
            text=_h2_vals, textposition="outside",
            hovertemplate="%{x}<br>2e MT : %{y}<extra></extra>",
        ))
        _fig_fat.update_layout(
            barmode="group",
            template="plotly_dark",
            paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
            height=340,
            margin=dict(l=20, r=20, t=30, b=40),
            legend=dict(orientation="h", y=1.08, x=0.5, xanchor="center",
                        font=dict(size=12, color="#e2e8f0")),
            xaxis=dict(tickfont=dict(size=11, color="#e2e8f0"), gridcolor="#1f2937"),
            yaxis=dict(gridcolor="#1f2937", tickfont=dict(size=11, color="#a0aec0")),
            bargap=0.2, bargroupgap=0.08,
        )
        st.plotly_chart(_fig_fat, use_container_width=True)

        # Tableau de variation
        _fat_table_rows = []
        for _m, _hib in _fat_metrics:
            _prefix_a = f"{team_a} - "
            _h1 = int(_df_h1["Nom de la ligne"].str.startswith(_prefix_a + _m).sum())
            _h2 = int(_df_h2["Nom de la ligne"].str.startswith(_prefix_a + _m).sum())
            _delta = _h2 - _h1
            _pct   = round((_delta / _h1 * 100), 1) if _h1 > 0 else 0.0
            if _hib:
                _trend = "↑ Meilleur" if _delta > 0 else ("↓ Fatigue" if _delta < 0 else "= Stable")
            else:
                _trend = "↑ Meilleur" if _delta < 0 else ("↓ Pire" if _delta > 0 else "= Stable")
            _fat_table_rows.append({
                "Métrique":       _m,
                "1re MT":         _h1,
                "2e MT":          _h2,
                "Variation":      f"{'+ ' if _delta > 0 else ''}{_delta}",
                "% variation":    f"{'+ ' if _pct > 0 else ''}{_pct}%",
                "Signal":         _trend,
            })
        st.dataframe(pd.DataFrame(_fat_table_rows), use_container_width=True, hide_index=True)
    else:
        st.info("Colonne 'Periode de jeu' introuvable pour l'analyse fatigue.")


# ══════════════════════════════════════════════
#  PAGE : COMPARAISON ÉQUIPES
# ══════════════════════════════════════════════


elif page == "Comparaison équipes":
    st.markdown("<div class='section-title'>Jeu offensif</div>", unsafe_allow_html=True)
    off_keys = ["Contacts", "Rucks", "Passes", "Soutiens Offensifs",
                "Franchissements", "Defenseurs battus", "Jeux a la main"]
    fig_off = bar_comparison(
        off_keys,
        [get_stat(stats_a, k) for k in off_keys],
        [get_stat(stats_b, k) for k in off_keys],
        "Actions offensives"
    )
    st.plotly_chart(fig_off, use_container_width=True)

    st.markdown("<div class='section-title'>Jeu défensif</div>", unsafe_allow_html=True)
    def_keys = ["Plaquages", "Assistant plaqueur", "Contre Ruck",
                "Contest", "Ballons perdus", "Turn over et contre attaque"]
    fig_def = bar_comparison(
        def_keys,
        [get_stat(stats_a, k) for k in def_keys],
        [get_stat(stats_b, k) for k in def_keys],
        "Actions défensives"
    )
    st.plotly_chart(fig_def, use_container_width=True)

    st.markdown("<div class='section-title'>Coups de pied & Phases statiques</div>", unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        static_keys = ["Melees", "Touches", "Lancements sur melees", "Lancements sur touches"]
        fig_static = bar_comparison(
            static_keys,
            [get_stat(stats_a, k) for k in static_keys],
            [get_stat(stats_b, k) for k in static_keys],
            "Phases statiques"
        )
        st.plotly_chart(fig_static, use_container_width=True)
    with col2:
        kick_keys = ["Jeux au pied", "Coups d'envoi", "Receptions aeriennes",
                     "Renvois ligne de but", "Botteur"]
        fig_kick = bar_comparison(
            kick_keys,
            [get_stat(stats_a, k) for k in kick_keys],
            [get_stat(stats_b, k) for k in kick_keys],
            "Jeu au pied"
        )
        st.plotly_chart(fig_kick, use_container_width=True)

    # Efficacité plaquages
    st.markdown("<div class='section-title'>Efficacité des plaquages</div>", unsafe_allow_html=True)
    plaq_a = df[df["Nom de la ligne"] == f"{team_a} - Plaquages"]
    plaq_b = df[df["Nom de la ligne"] == f"{team_b} - Plaquages"]
    col1, col2 = st.columns(2)
    with col1:
        eff_a = plaq_a["Efficacite"].dropna().value_counts()
        if not eff_a.empty:
            st.plotly_chart(
                pie_chart(eff_a.index.tolist(), eff_a.values.tolist(),
                          f"{team_a} — Efficacité plaquages",
                          ["#68d391", "#fc8181", "#fbd38d", "#b794f4"]),
                use_container_width=True
            )
    with col2:
        eff_b = plaq_b["Efficacite"].dropna().value_counts()
        if not eff_b.empty:
            st.plotly_chart(
                pie_chart(eff_b.index.tolist(), eff_b.values.tolist(),
                          f"{team_b} — Efficacité plaquages",
                          ["#68d391", "#fc8181", "#fbd38d", "#b794f4"]),
                use_container_width=True
            )


# ══════════════════════════════════════════════
#  PAGE : ANALYSE JOUEURS
# ══════════════════════════════════════════════
elif page == "Analyse joueurs":
    st.markdown("<div class='section-title'>Classement des joueurs par volume d'actions</div>", unsafe_allow_html=True)

    player_rows = df[
        ~df["Nom de la ligne"].str.startswith("Stade") &
        ~df["Nom de la ligne"].str.startswith("Sequence") &
        ~df["Nom de la ligne"].str.startswith("Aurelien") &
        (df["Nom de la ligne"] != "nan") &
        (df["Nom de la ligne"] != "")
    ]

    player_counts = player_rows["Nom de la ligne"].value_counts().reset_index()
    player_counts.columns = ["Joueur", "Actions"]

    top_n = st.slider("Nombre de joueurs à afficher", 5, 30, 20)
    player_counts_top = player_counts.head(top_n)

    fig = px.bar(
        player_counts_top, x="Actions", y="Joueur", orientation="h",
        color="Actions", color_continuous_scale=["#2b6cb0", "#63b3ed", "#bee3f8"],
        title=f"Top {top_n} joueurs — Volume d'actions"
    )
    fig.update_layout(
        plot_bgcolor="#1a1f2e", paper_bgcolor="#1a1f2e",
        font=dict(color="#e2e8f0"),
        yaxis=dict(autorange="reversed", gridcolor="#2d3748"),
        xaxis=dict(gridcolor="#2d3748"),
        coloraxis_showscale=False,
        margin=dict(l=180, r=20, t=40, b=40),
        height=max(400, top_n * 28),
    )
    st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")
    st.markdown("<div class='section-title'>Fiche individuelle & Zone d'activité</div>", unsafe_allow_html=True)

    all_players = player_counts["Joueur"].tolist()
    selected_player = st.selectbox("Choisir un joueur", all_players)

    player_df = player_rows[player_rows["Nom de la ligne"] == selected_player]
    total = len(player_df)

    # ── Terrain de rugby : zone d'activité ─────────────────────────
    st.plotly_chart(
        rugby_field_activity(player_df, selected_player),
        use_container_width=True,
    )

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(f"""<div class='kpi-card'>
          <div class='kpi-value team-a-color'>{total}</div>
          <div class='kpi-label'>Actions totales</div>
        </div>""", unsafe_allow_html=True)

    # Period breakdown for this player
    period_col = "Periode de jeu"
    if period_col in player_df.columns:
        period_vals = player_df[period_col].replace("nan", pd.NA).dropna().value_counts()
        with col2:
            most_active = period_vals.index[0] if not period_vals.empty else "N/A"
            st.markdown(f"""<div class='kpi-card'>
              <div class='kpi-value' style='font-size:1.4rem;color:#fbd38d;'>{most_active}</div>
              <div class='kpi-label'>Période la + active</div>
            </div>""", unsafe_allow_html=True)

    # Code
    code_vals = player_df["code"].replace("nan", pd.NA).dropna().value_counts()
    with col3:
        st.markdown(f"""<div class='kpi-card'>
          <div class='kpi-value team-b-color'>{len(code_vals)}</div>
          <div class='kpi-label'>Catégories distinctes</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        if not code_vals.empty:
            st.plotly_chart(
                pie_chart(code_vals.index.tolist(), code_vals.values.tolist(),
                          "Répartition des actions",
                          px.colors.qualitative.Set3),
                use_container_width=True
            )
    with col2:
        if not period_vals.empty:
            _pv = period_vals.values.tolist()
            _pt = period_vals.sum()
            _ptexts = [f"{v}\n({v/_pt*100:.0f}%)" if _pt > 0 else str(v) for v in _pv]
            fig_p = go.Figure(go.Bar(
                x=period_vals.index.tolist(),
                y=_pv,
                text=_ptexts, textposition="outside",
                textfont=dict(size=11, color="#e2e8f0"),
                cliponaxis=False,
                marker=dict(
                    color=_pv,
                    colorscale=[[0, "#2b6cb0"], [1, "#63b3ed"]],
                    line_width=0,
                ),
            ))
            fig_p.update_layout(
                title="Actions par quart-temps",
                plot_bgcolor="#1a1f2e", paper_bgcolor="#1a1f2e",
                font=dict(color="#e2e8f0"),
                xaxis=dict(gridcolor="#2d3748"),
                yaxis=dict(gridcolor="#2d3748", range=[0, max(_pv) * 1.4]),
                showlegend=False,
                margin=dict(l=40, r=20, t=60, b=40),
            )
            st.plotly_chart(fig_p, use_container_width=True)

    # Raw player data
    with st.expander("Données brutes du joueur"):
        non_empty = player_df.dropna(axis=1, how="all")
        display_cols = [c for c in non_empty.columns if non_empty[c].replace("nan", pd.NA).notna().any()]
        st.dataframe(non_empty[display_cols[:20]], use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════
#  PAGE : RUCKS & CONTACTS
# ══════════════════════════════════════════════
elif page == "Rucks & Contacts":
    st.markdown("<div class='section-title'>Vitesse de libération au ruck</div>", unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    for col, team, color in [(col1, team_a, COLOR_A), (col2, team_b, COLOR_B)]:
        ruck_df = df[df["Nom de la ligne"] == f"{team} - Rucks"]
        speed = ruck_df["Vitesse de liberation"].replace("nan", pd.NA).dropna().value_counts()
        with col:
            if not speed.empty:
                ORDER = ["0 - 3 sec", "3 - 6 sec", "+ 6 sec", "Arret de l'action"]
                speed = speed.reindex([o for o in ORDER if o in speed.index]).fillna(0)
                COLORS_SPEED = ["#68d391", "#fbd38d", "#fc8181", "#b794f4"]
                _sv = speed.values.tolist()
                _st = sum(_sv)
                _stexts = [f"{v}\n({v/_st*100:.0f}%)" if _st > 0 else str(v) for v in _sv]
                fig = go.Figure(go.Bar(
                    x=speed.index.tolist(), y=_sv,
                    text=_stexts, textposition="outside",
                    textfont=dict(size=11, color="#e2e8f0"),
                    cliponaxis=False,
                    marker=dict(color=COLORS_SPEED[:len(speed)], line_width=0),
                ))
                fig.update_layout(
                    title=f"{team} — Vitesse de libération",
                    plot_bgcolor="#1a1f2e", paper_bgcolor="#1a1f2e",
                    font=dict(color="#e2e8f0"),
                    xaxis=dict(gridcolor="#2d3748"),
                    yaxis=dict(gridcolor="#2d3748", range=[0, max(_sv) * 1.4]),
                    showlegend=False,
                    margin=dict(l=40, r=20, t=60, b=40),
                )
                st.plotly_chart(fig, use_container_width=True)

    st.markdown("<div class='section-title'>Contre-rucks & Contest</div>", unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        cr = {"Contre Ruck": [get_stat(stats_a, "Contre Ruck"), get_stat(stats_b, "Contre Ruck")],
              "Contest": [get_stat(stats_a, "Contest"), get_stat(stats_b, "Contest")]}
        fig_cr = go.Figure()
        _cr_max = 1
        for label, (va, vb) in cr.items():
            _tot = va + vb
            _ta = f"{va}\n({va/_tot*100:.0f}%)" if _tot > 0 else str(va)
            _tb = f"{vb}\n({vb/_tot*100:.0f}%)" if _tot > 0 else str(vb)
            _cr_max = max(_cr_max, va, vb)
            fig_cr.add_trace(go.Bar(
                name=label, x=[team_a, team_b], y=[va, vb],
                text=[_ta, _tb], textposition="outside",
                textfont=dict(size=11, color="#e2e8f0"), cliponaxis=False,
            ))
        fig_cr.update_layout(
            title="Contre-ruck & Contest",
            barmode="group",
            plot_bgcolor="#1a1f2e", paper_bgcolor="#1a1f2e",
            font=dict(color="#e2e8f0"),
            xaxis=dict(gridcolor="#2d3748"),
            yaxis=dict(gridcolor="#2d3748", range=[0, _cr_max * 1.45]),
            legend=dict(bgcolor="#2d3748"),
            margin=dict(l=40, r=20, t=60, b=40),
        )
        st.plotly_chart(fig_cr, use_container_width=True)
    with col2:
        sout = {
            "Soutiens Off.": [get_stat(stats_a, "Soutiens Offensifs"), get_stat(stats_b, "Soutiens Offensifs")],
            "Jeux à la main": [get_stat(stats_a, "Jeux a la main"), get_stat(stats_b, "Jeux a la main")],
        }
        fig_so = go.Figure()
        _so_max = 1
        for i, (label, (va, vb)) in enumerate(sout.items()):
            _tot = va + vb
            _ta = f"{va}\n({va/_tot*100:.0f}%)" if _tot > 0 else str(va)
            _tb = f"{vb}\n({vb/_tot*100:.0f}%)" if _tot > 0 else str(vb)
            _so_max = max(_so_max, va, vb)
            fig_so.add_trace(go.Bar(
                name=label, x=[team_a, team_b], y=[va, vb],
                text=[_ta, _tb], textposition="outside",
                textfont=dict(size=11, color="#e2e8f0"), cliponaxis=False,
            ))
        fig_so.update_layout(
            title="Soutien offensif",
            barmode="group",
            plot_bgcolor="#1a1f2e", paper_bgcolor="#1a1f2e",
            font=dict(color="#e2e8f0"),
            xaxis=dict(gridcolor="#2d3748"),
            yaxis=dict(gridcolor="#2d3748", range=[0, _so_max * 1.45]),
            legend=dict(bgcolor="#2d3748"),
            margin=dict(l=40, r=20, t=60, b=40),
        )
        st.plotly_chart(fig_so, use_container_width=True)

    # Contact zone analysis
    st.markdown("<div class='section-title'>Zones de contact</div>", unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    for col, team, color in [(col1, team_a, COLOR_A), (col2, team_b, COLOR_B)]:
        cont_df = df[df["Nom de la ligne"] == f"{team} - Contacts"]
        zones = cont_df["Zone terrain"].replace("nan", pd.NA).dropna().value_counts()
        with col:
            if not zones.empty:
                _zv = zones.values.tolist()
                _zt = sum(_zv)
                _ztexts = [f"{v}\n({v/_zt*100:.0f}%)" if _zt > 0 else str(v) for v in _zv]
                fig_z = px.bar(
                    x=zones.index, y=_zv,
                    color_discrete_sequence=[color],
                    title=f"{team} — Contacts par zone",
                    labels={"x": "Zone", "y": "Nombre"},
                    text=_ztexts,
                )
                fig_z.update_traces(textposition="outside", textfont=dict(size=11, color="#e2e8f0"),
                                    cliponaxis=False)
                fig_z.update_layout(
                    plot_bgcolor="#1a1f2e", paper_bgcolor="#1a1f2e",
                    font=dict(color="#e2e8f0"),
                    xaxis=dict(gridcolor="#2d3748"),
                    yaxis=dict(gridcolor="#2d3748", range=[0, max(_zv) * 1.4]),
                    showlegend=False,
                    margin=dict(l=40, r=20, t=60, b=40),
                )
                st.plotly_chart(fig_z, use_container_width=True)


# ══════════════════════════════════════════════
#  PAGE : JEU AU PIED & PASSES
# ══════════════════════════════════════════════
elif page == "Jeu au pied & Passes":
    st.markdown("<div class='section-title'>Passes — Type & Résultat</div>", unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    for col, team, color in [(col1, team_a, COLOR_A), (col2, team_b, COLOR_B)]:
        pass_df = df[df["Nom de la ligne"] == f"{team} - Passes"]
        with col:
            st.markdown(f"**{team}**")
            tab1, tab2 = st.tabs(["Type de passe", "Résultat"])
            with tab1:
                types = pass_df["Type de passe"].replace("nan", pd.NA).dropna().value_counts()
                if not types.empty:
                    st.plotly_chart(
                        pie_chart(types.index.tolist(), types.values.tolist(),
                                  "Types de passes", px.colors.qualitative.Set2),
                        use_container_width=True
                    )
            with tab2:
                res = pass_df["Resultat"].replace("nan", pd.NA).dropna().value_counts()
                if not res.empty:
                    COLORS_RES = {"Positif": "#68d391", "Neutre": "#fbd38d", "Negatif": "#fc8181"}
                    colors_list = [COLORS_RES.get(r, "#b794f4") for r in res.index]
                    st.plotly_chart(
                        pie_chart(res.index.tolist(), res.values.tolist(),
                                  "Résultats des passes", colors_list),
                        use_container_width=True
                    )

    st.markdown("---")
    st.markdown("<div class='section-title'>Jeu au pied — Analyse</div>", unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    for col, team, color in [(col1, team_a, COLOR_A), (col2, team_b, COLOR_B)]:
        jap_df = df[df["Nom de la ligne"] == f"{team} - Jeux au pied"]
        with col:
            st.markdown(f"**{team}** — {len(jap_df)} coups de pied")
            tab1, tab2 = st.tabs(["Type", "Résultat"])
            with tab1:
                jap_types = jap_df["Type de jeu au pied"].replace("nan", pd.NA).dropna().value_counts()
                if not jap_types.empty:
                    _jtv = jap_types.values.tolist()
                    _jtt = sum(_jtv)
                    _jttexts = [f"{v} ({v/_jtt*100:.0f}%)" if _jtt > 0 else str(v) for v in _jtv]
                    fig_jt = px.bar(
                        x=_jtv, y=jap_types.index, orientation="h",
                        color_discrete_sequence=[color],
                        labels={"x": "Nombre", "y": "Type"},
                        text=_jttexts,
                    )
                    fig_jt.update_traces(textposition="outside", textfont=dict(size=11, color="#e2e8f0"),
                                         cliponaxis=False)
                    fig_jt.update_layout(
                        plot_bgcolor="#1a1f2e", paper_bgcolor="#1a1f2e",
                        font=dict(color="#e2e8f0"), showlegend=False,
                        yaxis=dict(autorange="reversed", gridcolor="#2d3748"),
                        xaxis=dict(gridcolor="#2d3748", range=[0, max(_jtv) * 1.35]),
                        margin=dict(l=20, r=80, t=20, b=20),
                    )
                    st.plotly_chart(fig_jt, use_container_width=True)
            with tab2:
                jap_res = jap_df["Efficacite"].replace("nan", pd.NA).dropna().value_counts()
                if not jap_res.empty:
                    COLORS_EFF = {"EFFICACE": "#68d391", "NON EFFICACE": "#fc8181", "NEUTRE": "#fbd38d"}
                    colors_e = [COLORS_EFF.get(r, "#b794f4") for r in jap_res.index]
                    st.plotly_chart(
                        pie_chart(jap_res.index.tolist(), jap_res.values.tolist(),
                                  "Efficacité", colors_e),
                        use_container_width=True
                    )


    st.markdown("---")
    st.markdown("<div class='section-title'>📊 Tableau Statistiques Coups de Pied</div>", unsafe_allow_html=True)

    # ── Colonnes (id, libellé, couleur_header, source, valeur_filtre) ──
    _KCOLS = [
        ("BOMBE",    "BOMBE",              "#b45fbf", "pied",    "Pression"),
        ("TORPILLE", "TORPILLE",           "#9c27b0", "pied",    "Degagement"),
        ("OCCUP",    "OCCUP",              "#2e7d32", "pied",    "Occupation dans le terrain"),
        ("BOX",      "BOX",                "#0277bd", "pied",    "Court"),
        ("SAPIN",    "SAPIN",              "#00838f", "pied",    "Occupation hors terrain"),
        ("PAU",      "PAU",                "#1565c0", "pied",    "Passe au pied"),
        ("BRIVE",    "BRIVE",              "#283593", None,      None),
        ("SUMMIT",   "SUMMIT/\nSMALL/\nSNAKE", "#00695c", None, None),
        ("CE",       "CE",                 "#1976d2", "envoi",   None),
        ("CR22",     "CR22",               "#0097a7", "r22",     None),
        ("CREB",     "CREB",               "#827717", "rlb",     None),
        ("PENAL_T",  "PENAL\nTOUCHE",      "#5d4037", "botteur", None),
        ("PENALITE", "PENALITE",           "#006064", "tir",     "Penalite"),
        ("TRANSFO",  "TRANSFO",            "#004d40", "tir",     "Transformation"),
        ("DROP",     "DROP",               "#006b6b", None,      None),
    ]
    # ── Groupes efficacité bas de tableau ──
    _KGROUPS = [
        ("BOMBE\nTORPILLE", ["BOMBE","TORPILLE"],          "#6a1b9a"),
        ("OCCUP",           ["OCCUP"],                      "#1b5e20"),
        ("BOX\nSAPIN\nPAU\nBRIVE\nSUMMIT", ["BOX","SAPIN","PAU","BRIVE","SUMMIT"], "#01579b"),
        ("CE\nCR22",        ["CE","CR22"],                  "#0d47a1"),
        ("CREB",            ["CREB"],                       "#558b2f"),
        ("PENAL\nTOUCHE",   ["PENAL_T"],                   "#4e342e"),
        ("PENALITE\nTRANSFO\nDROP", ["PENALITE","TRANSFO","DROP"], "#004d40"),
    ]

    # ── Extraction données ──
    def _kick_player(row, src):
        j = str(row.get("Joueur","") or "").strip()
        b = str(row.get("Botteur","") or "").strip()
        ok = 1 if str(row.get("Resultat","")).strip() == "Positif" else 0
        name = (b if b and b.lower() not in ("nan","") else j) if src in ("envoi","r22","rlb") \
               else (j if j and j.lower() not in ("nan","") else b)
        return (name if name and name.lower() not in ("nan","") else None), ok

    _kp = {}
    _auri_prefix = team_a
    for _cid, _, _, _src, _fval in _KCOLS:
        if _src is None:
            continue
        if _src == "pied":
            _rows = df[(df["Nom de la ligne"] == f"{_auri_prefix} - Jeux au pied") &
                       (df["Type de jeu au pied"].astype(str).str.strip() == _fval)]
        elif _src == "envoi":
            _rows = df[df["Nom de la ligne"].astype(str).str.contains(f"{_auri_prefix}.*Coups d", regex=True, na=False)]
        elif _src == "r22":
            _rows = df[df["Nom de la ligne"].astype(str).str.contains(f"{_auri_prefix}.*Renvois aux", regex=True, na=False)]
        elif _src == "rlb":
            _rows = df[df["Nom de la ligne"].astype(str).str.contains(f"{_auri_prefix}.*Renvois ligne", regex=True, na=False)]
        elif _src == "botteur":
            _rows = df[df["Nom de la ligne"].astype(str).str.contains(f"{_auri_prefix}.*Botteur", regex=True, na=False)]
        elif _src == "tir":
            _rows = df[(df["Nom de la ligne"].astype(str).str.contains(f"{_auri_prefix}.*Buteur", regex=True, na=False)) &
                       (df["Type de tir au but"].astype(str).str.strip() == _fval)]
        else:
            continue
        for _, _row in _rows.iterrows():
            _name, _ok = _kick_player(_row, _src)
            if not _name:
                continue
            _kp.setdefault(_name, {}).setdefault(_cid, {"t": 0, "r": 0})
            _kp[_name][_cid]["t"] += 1
            _kp[_name][_cid]["r"] += _ok

    if _kp:
        def _th(txt, cs=1, rs=1, bg="#0d1b2e", fs=10, color="white"):
            return (f'<th colspan="{cs}" rowspan="{rs}" style="padding:7px 3px;text-align:center;'
                    f'font-size:{fs}px;font-weight:bold;background:{bg};color:{color};'
                    f'border:1px solid #374151;white-space:pre-line;line-height:1.2;">{txt}</th>')
        def _td(txt, cs=1, rs=1, bg="#1a1f2e", color="#e2e8f0", fw="normal", fs=11):
            return (f'<td colspan="{cs}" rowspan="{rs}" style="padding:6px 3px;text-align:center;'
                    f'font-size:{fs}px;font-weight:{fw};color:{color};'
                    f'background:{bg};border:1px solid #374151;white-space:nowrap;">{txt}</td>')

        # ── En-tête ligne 1 ──
        _h1 = _th("JOUEUR", rs=2, bg="#0d1b2e", fs=11)
        for _, _lbl, _col, _, _ in _KCOLS:
            _h1 += _th(_lbl, cs=2, bg=_col, fs=9)
        _h1 += _th("TOTAL", cs=2, rs=2, bg="#37474f", fs=11)
        _h1 += _th("EFFICACITE", rs=2, bg="#111827", fs=10)

        # ── En-tête ligne 2 : T. / R. ──
        _h2 = ""
        for _ in _KCOLS:
            _h2 += _th("T.", bg="#1c2333", fs=8, color="#9ca3af")
            _h2 += _th("R.", bg="#1c2333", fs=8, color="#86efac")

        # ── Lignes joueurs ──
        _rows_html = ""
        _sorted_players = sorted(_kp.items(), key=lambda x: x[0].split()[-1] if x[0].split() else x[0])
        for _pi, (_player, _pst) in enumerate(_sorted_players):
            _rb = "#161b27" if _pi % 2 == 0 else "#1e2535"
            _r = _td(f"<b>{_player}</b>", bg=_rb, fw="bold", fs=11, color="#f1f5f9")
            _tt, _tr = 0, 0
            for _cid, _, _col, _, _ in _KCOLS:
                _t = _pst.get(_cid, {}).get("t", 0)
                _rv = _pst.get(_cid, {}).get("r", 0)
                _tt += _t; _tr += _rv
                _tc = "#6b7280" if _t == 0 else "#cbd5e0"
                _rc = "#6b7280" if _t == 0 else "#86efac"
                _r += _td(str(_t), bg=_rb, color=_tc)
                _r += _td(str(_rv), bg=_rb, color=_rc)
            _r += _td(f"<b>{_tt}</b>", bg=_rb, color="white", fw="bold")
            _r += _td(f"<b>{_tr}</b>", bg=_rb, color="#86efac", fw="bold")
            if _tt > 0:
                _pct = _tr / _tt * 100
                _ec = "#4ade80" if _pct >= 70 else ("#fbbf24" if _pct >= 50 else "#f87171")
                _eff = f"<b>{_pct:.1f}%</b>"
            else:
                _ec, _eff = "#6b7280", "0.0%"
            _r += _td(_eff, bg=_rb, color=_ec, fw="bold", fs=12)
            _rows_html += f"<tr>{_r}</tr>"

        # ── Ligne TOTAL ──
        _totals = {}
        for _cid, _, _, _, _ in _KCOLS:
            _t = sum(_ps.get(_cid,{}).get("t",0) for _ps in _kp.values())
            _r2 = sum(_ps.get(_cid,{}).get("r",0) for _ps in _kp.values())
            _totals[_cid] = {"t": _t, "r": _r2}
        _grand_t = sum(v["t"] for v in _totals.values())
        _grand_r = sum(v["r"] for v in _totals.values())

        _frow = _td("<b>TOTAL</b>", bg="#37474f", color="white", fw="bold")
        for _cid, _, _, _, _ in _KCOLS:
            _frow += _td(f"<b>{_totals[_cid]['t']}</b>", bg="#37474f", color="white", fw="bold")
            _frow += _td(f"<b>{_totals[_cid]['r']}</b>", bg="#37474f", color="#86efac", fw="bold")
        _frow += _td(f"<b>{_grand_t}</b>", bg="#37474f", color="white", fw="bold")
        _frow += _td(f"<b>{_grand_r}</b>", bg="#37474f", color="#86efac", fw="bold")
        _frow += _td("", bg="#37474f")

        # ── Ligne % par colonne ──
        _prow = _td("", bg="#111827")
        for _cid, _, _col, _, _ in _KCOLS:
            _t = _totals[_cid]["t"]; _r2 = _totals[_cid]["r"]
            _p = f"{_r2/_t*100:.1f} %" if _t > 0 else "-"
            _prow += _td(_p, cs=2, bg=_col + "55", color="white", fw="bold", fs=10)
        _prow += _td("", cs=2, bg="#111827")
        _prow += _td("", bg="#111827")

        # ── Ligne EFFICACITE par groupe ──
        _grow = _td("<b>EFFICACITE</b>", bg="#111827", color="#9ca3af", fs=9)
        for _gname, _gcols, _gcol in _KGROUPS:
            _gt = sum(_totals.get(c,{}).get("t",0) for c in _gcols)
            _gr = sum(_totals.get(c,{}).get("r",0) for c in _gcols)
            _gp = f"<b>{_gr/_gt*100:.1f} %</b>" if _gt > 0 else "<b>-</b>"
            _cs = sum(2 for _ in _gcols)
            _grow += _td(_gp, cs=_cs, bg=_gcol, color="white", fw="bold", fs=12)
        _geff = f"{_grand_r/_grand_t*100:.2f}%" if _grand_t > 0 else "0%"
        _grow += _td("", cs=2, bg="#111827")
        _grow += _td(
            f'<div style="display:inline-block;background:#111827;color:white;'
            f'font-size:14px;font-weight:bold;padding:6px 10px;'
            f'clip-path:polygon(50% 0%,100% 25%,100% 75%,50% 100%,0% 75%,0% 25%);">'
            f'{_geff}</div>',
            bg="#111827", color="white", fw="bold", fs=13
        )

        _html_kick = f"""
        <div style="overflow-x:auto;margin-top:12px;border-radius:6px;border:1px solid #374151;">
          <table style="border-collapse:collapse;width:100%;font-family:'Segoe UI',Arial,sans-serif;">
            <thead>
              <tr>{_h1}</tr>
              <tr>{_h2}</tr>
            </thead>
            <tbody>
              {_rows_html}
              <tr>{_frow}</tr>
              <tr>{_prow}</tr>
              <tr>{_grow}</tr>
            </tbody>
          </table>
        </div>
        """
        st.markdown(_html_kick, unsafe_allow_html=True)
    else:
        st.info("Aucune donnée de coups de pied disponible dans ce fichier.")


# ══════════════════════════════════════════════
#  PAGE : PAR QUART-TEMPS
# ══════════════════════════════════════════════
elif page == "Par quart-temps":
    st.markdown("<div class='section-title'>Distribution des actions par quart-temps</div>", unsafe_allow_html=True)

    period_col = "Periode de jeu"
    period_order = ["1er quart-temps", "2nd quart-temps", "3e quart-temps", "4e quart-temps"]

    # Global distribution
    period_dist = df[period_col].replace("nan", pd.NA).dropna()
    period_dist = period_dist[period_dist.isin(period_order)].value_counts()
    period_dist = period_dist.reindex(period_order).fillna(0)

    _gv = period_dist.values.tolist()
    _gt = sum(_gv)
    _gtexts = [f"{int(v)}\n({v/_gt*100:.0f}%)" if _gt > 0 else str(int(v)) for v in _gv]
    fig_global = go.Figure(go.Bar(
        x=period_dist.index.tolist(),
        y=_gv,
        text=_gtexts, textposition="outside",
        textfont=dict(size=12, color="#e2e8f0"),
        cliponaxis=False,
        marker=dict(
            color=["#3182ce", "#63b3ed", "#fc8181", "#e53e3e"],
            line_width=0,
        ),
    ))
    fig_global.update_layout(
        title="Volume total d'actions par quart-temps",
        plot_bgcolor="#1a1f2e", paper_bgcolor="#1a1f2e",
        font=dict(color="#e2e8f0"),
        xaxis=dict(gridcolor="#2d3748"),
        yaxis=dict(gridcolor="#2d3748", range=[0, max(_gv) * 1.4]),
        showlegend=False,
        margin=dict(l=40, r=20, t=60, b=40),
    )
    st.plotly_chart(fig_global, use_container_width=True)

    # Per team per period
    st.markdown("<div class='section-title'>Statistiques clés par quart-temps</div>", unsafe_allow_html=True)
    stat_choice = st.selectbox(
        "Statistique à analyser par quart-temps",
        ["Plaquages", "Rucks", "Contacts", "Passes", "Jeux au pied",
         "Soutiens Offensifs", "Ballons perdus"]
    )

    data_period = []
    for period in period_order:
        period_df = df[df[period_col] == period]
        for team in [team_a, team_b]:
            count = len(period_df[period_df["Nom de la ligne"] == f"{team} - {stat_choice}"])
            data_period.append({"Période": period, "Équipe": team, "Valeur": count})

    df_period = pd.DataFrame(data_period)
    if not df_period.empty:
        # Calcule % A vs B pour chaque période
        def _period_pct_text(row, df_p):
            same_period = df_p[df_p["Période"] == row["Période"]]
            total = same_period["Valeur"].sum()
            v = row["Valeur"]
            return f"{v}\n({v/total*100:.0f}%)" if total > 0 else str(v)
        df_period = df_period.copy()
        df_period["_text"] = df_period.apply(lambda r: _period_pct_text(r, df_period), axis=1)
        _pmax = df_period["Valeur"].max() if not df_period.empty else 1
        fig_period = px.bar(
            df_period, x="Période", y="Valeur", color="Équipe",
            color_discrete_map={team_a: COLOR_A, team_b: COLOR_B},
            barmode="group",
            title=f"{stat_choice} par quart-temps",
            text="_text",
        )
        fig_period.update_traces(textposition="outside", textfont=dict(size=11, color="#e2e8f0"),
                                  cliponaxis=False)
        fig_period.update_layout(
            plot_bgcolor="#1a1f2e", paper_bgcolor="#1a1f2e",
            font=dict(color="#e2e8f0"),
            xaxis=dict(gridcolor="#2d3748"),
            yaxis=dict(gridcolor="#2d3748", range=[0, _pmax * 1.45]),
            legend=dict(bgcolor="#2d3748", bordercolor="#4a5568"),
            margin=dict(l=40, r=20, t=60, b=40),
        )
        st.plotly_chart(fig_period, use_container_width=True)

    # Line chart (trend)
    fig_line = px.line(
        df_period, x="Période", y="Valeur", color="Équipe",
        color_discrete_map={team_a: COLOR_A, team_b: COLOR_B},
        markers=True,
        title=f"Tendance — {stat_choice}",
    )
    fig_line.update_traces(line=dict(width=3), marker=dict(size=10))
    fig_line.update_layout(
        plot_bgcolor="#1a1f2e", paper_bgcolor="#1a1f2e",
        font=dict(color="#e2e8f0"),
        xaxis=dict(gridcolor="#2d3748"),
        yaxis=dict(gridcolor="#2d3748"),
        legend=dict(bgcolor="#2d3748", bordercolor="#4a5568"),
        margin=dict(l=40, r=20, t=40, b=40),
    )
    st.plotly_chart(fig_line, use_container_width=True)

    # ── Possession par quart-temps ───────────────────────────────────────
    st.markdown("---")
    st.markdown("<div class='section-title'>Possession par quart-temps</div>",
                unsafe_allow_html=True)

    poss_data = []
    for period in period_order:
        poss_a_qt = len(df[
            (df["Nom de la ligne"] == f"{team_a} - Possession") &
            (df[period_col] == period)
        ])
        poss_b_qt = len(df[
            (df["Nom de la ligne"] == f"{team_b} - Possession") &
            (df[period_col] == period)
        ])
        total_qt = poss_a_qt + poss_b_qt
        poss_data.append({
            "period": period,
            "poss_a": poss_a_qt,
            "poss_b": poss_b_qt,
            "total": total_qt,
            "pct_a": round(poss_a_qt / total_qt * 100, 1) if total_qt > 0 else 50.0,
            "pct_b": round(poss_b_qt / total_qt * 100, 1) if total_qt > 0 else 50.0,
        })

    # 4 mini-donuts côte à côte
    qt_labels = [p["period"].replace("quart-temps", "QT").replace("1er ", "Q1 ")
                              .replace("2nd ", "Q2 ").replace("3e ", "Q3 ").replace("4e ", "Q4 ")
                 for p in poss_data]
    donut_cols = st.columns(4)
    for i, (col_qt, pd_row) in enumerate(zip(donut_cols, poss_data)):
        with col_qt:
            label = period_order[i].replace("quart-temps", "").strip()
            fig_qt = go.Figure(go.Pie(
                labels=[team_a, team_b],
                values=[pd_row["pct_a"], pd_row["pct_b"]],
                hole=0.55,
                marker=dict(colors=[COLOR_A, COLOR_B]),
                textinfo="percent",
                textfont=dict(size=12, color="#fff"),
                hovertemplate="%{label} : %{percent}<extra></extra>",
            ))
            fig_qt.update_layout(
                title=dict(text=label, font=dict(size=13, color="#e2e8f0"), x=0.5),
                paper_bgcolor="#0d1117",
                plot_bgcolor="#0d1117",
                showlegend=False,
                margin=dict(l=5, r=5, t=35, b=5),
                height=200,
                annotations=[dict(
                    text=f"<b>{pd_row['pct_a']:.0f}%</b>",
                    x=0.5, y=0.5, showarrow=False,
                    font=dict(size=16, color=COLOR_A),
                )],
            )
            st.plotly_chart(fig_qt, use_container_width=True)

    # Barre horizontale empilée (% par quart)
    fig_stack = go.Figure()
    fig_stack.add_trace(go.Bar(
        name=team_a,
        y=[p["period"] for p in poss_data],
        x=[p["pct_a"] for p in poss_data],
        orientation="h",
        marker_color=COLOR_A,
        text=[f"{p['pct_a']:.1f}%" for p in poss_data],
        textposition="inside",
        textfont=dict(color="#fff", size=13),
        hovertemplate=f"{team_a} : %{{x:.1f}}%<extra></extra>",
    ))
    fig_stack.add_trace(go.Bar(
        name=team_b,
        y=[p["period"] for p in poss_data],
        x=[p["pct_b"] for p in poss_data],
        orientation="h",
        marker_color=COLOR_B,
        text=[f"{p['pct_b']:.1f}%" for p in poss_data],
        textposition="inside",
        textfont=dict(color="#fff", size=13),
        hovertemplate=f"{team_b} : %{{x:.1f}}%<extra></extra>",
    ))
    fig_stack.update_layout(
        barmode="stack",
        template="plotly_dark",
        paper_bgcolor="#0d1117",
        plot_bgcolor="#0d1117",
        height=220,
        margin=dict(l=130, r=20, t=20, b=30),
        xaxis=dict(range=[0, 100], ticksuffix="%", gridcolor="#2d3748"),
        yaxis=dict(autorange="reversed", color="#e2e8f0"),
        legend=dict(orientation="h", y=1.12, x=0),
        bargap=0.3,
    )
    st.plotly_chart(fig_stack, use_container_width=True)


# ══════════════════════════════════════════════
#  PAGE : ANALYSE GPS
# ══════════════════════════════════════════════
elif page == "Analyse GPS":

    # ── Fonction de chargement GPS ────────────────────────────────────────
    _GPS_NUM_COLS = ["m/min", "HI", "%HI", "THI", "Vmax", "Vmax%",
                     "RHIE Efforts Per Bout - Max", "RHIE Total Bouts",
                     "Distance totale", "Courue", "Marchée",
                     "Total Player Load", "Acc 2,5 m/s/s", "Dec -2,5 m/s/s"]

    @st.cache_data(show_spinner=False)
    def _load_gps_raw(path: str) -> pd.DataFrame:
        """Charge HUB DATAS Excel → toutes les lignes Match (Session + Quart-temps)."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            sheet = wb["HUB DATAS 2.0"] if "HUB DATAS 2.0" in wb.sheetnames else wb.active
            data = list(sheet.iter_rows(values_only=True))
            headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(data[0])]
            df = pd.DataFrame(data[1:], columns=headers)
            df.columns = [c.strip() for c in df.columns]
            df = df[df["Entrainement"].astype(str).str.strip() == "Match"].copy()
            for c in _GPS_NUM_COLS:
                if c in df.columns:
                    df[c] = pd.to_numeric(df[c], errors="coerce")
            if "Date" in df.columns:
                df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
            if "Period Number" in df.columns:
                df["Period Number"] = pd.to_numeric(df["Period Number"], errors="coerce")
            return df
        except Exception as e:
            st.error(f"Erreur chargement GPS : {e}")
            return pd.DataFrame()

    @st.cache_data(show_spinner=False)
    def load_gps(path: str) -> pd.DataFrame:
        """Retourne uniquement les lignes Session (total match par joueur)."""
        df = _load_gps_raw(path)
        if df.empty:
            return df
        return df[df["Period Name"].astype(str).str.strip() == "Session"].copy()

    @st.cache_data(show_spinner=False)
    def load_gps_quarters(path: str) -> pd.DataFrame:
        """Retourne les lignes Quart-temps avec colonne 'Quart' (Q1-Q4)."""
        df = _load_gps_raw(path)
        if df.empty:
            return df
        mask = df["Period Name"].astype(str).str.strip().str.startswith("Quart-temps")
        df_qt = df[mask].copy()
        if df_qt.empty:
            return df_qt

        # Assigner Q1-Q4 par match selon le rang du Period Number.
        # On boucle manuellement pour éviter que pandas retire "Activity Name"
        # des colonnes lors du groupby().apply() (comportement pandas >= 2.0).
        def _assign_quarters(grp):
            periods = sorted(grp["Period Number"].dropna().unique())
            n = len(periods)
            if n == 0:
                grp = grp.copy()
                grp["Quart"] = "Q1"
                return grp
            q_map = {p: f"Q{min(4, 1 + int(i * 4) // n)}"
                     for i, p in enumerate(periods)}
            grp = grp.copy()
            grp["Quart"] = grp["Period Number"].map(q_map).fillna("Q1")
            return grp

        _chunks = []
        for _act, _grp in df_qt.groupby("Activity Name", sort=False):
            _chunks.append(_assign_quarters(_grp))
        df_qt = pd.concat(_chunks, ignore_index=True) if _chunks else df_qt
        return df_qt

    # ── Regroupement Avant / Arrière ──────────────────────────────────────
    def groupe_position(pos: str) -> str:
        pos = str(pos).lower()
        avant_kw = ["prop", "hooker", "second row", "lock", "flanker",
                    "number 8", "loose", "tight", "pilier", "talonneur",
                    "deuxième", "troisième", "n°8", "scrum half"]
        if any(k in pos for k in avant_kw):
            return "Avant"
        return "Arrière"

    # ── Couleur %Vmax selon groupe ─────────────────────────────────────────
    def vmax_color(pct, groupe):
        if groupe == "Avant":
            if pct >= 85:   return "#68d391", "🟢"
            elif pct >= 80: return "#f6ad55", "🟠"
            else:           return "#fc8181", "🔴"
        else:  # Arrière
            if pct >= 90:   return "#68d391", "🟢"
            elif pct >= 80: return "#f6ad55", "🟠"
            else:           return "#fc8181", "🔴"

    # ── Chargement ────────────────────────────────────────────────────────
    _gps_path = st.session_state.get("gps_path")

    if not _gps_path:
        st.markdown("""
        <div style="text-align:center;padding:80px 0;">
          <div style="font-size:3.5rem;">📡</div>
          <h3 style="color:#e2e8f0;">Aucun fichier GPS chargé</h3>
          <p style="color:#a0aec0;max-width:420px;margin:12px auto;">
            Charge ton fichier <strong>HUB DATAS .xlsx</strong> via le panneau latéral
            (section "Données GPS") pour accéder à cette page.
          </p>
        </div>
        """, unsafe_allow_html=True)
        st.stop()

    with st.spinner("Chargement des données GPS…"):
        _df_gps = load_gps(_gps_path)

    if _df_gps.empty:
        st.warning("Aucune donnée de match trouvée dans le fichier GPS.")
        st.stop()

    # ── Sélecteur de match ────────────────────────────────────────────────
    _gps_matches = sorted(_df_gps["Activity Name"].dropna().unique().tolist())
    _sel_match = st.selectbox(
        "Match",
        _gps_matches,
        format_func=lambda x: x,
    )
    _df_m = _df_gps[_df_gps["Activity Name"] == _sel_match].copy()
    _match_date = _df_m["Date"].iloc[0] if not _df_m.empty and "Date" in _df_m.columns else None

    if _match_date is not None:
        st.markdown(f"""
        <div style="background:linear-gradient(135deg,#1a1f2e,#2d3748);
                    border-radius:12px;padding:16px 24px;margin-bottom:20px;
                    border:1px solid #4a5568;display:flex;align-items:center;gap:20px;">
          <div>
            <div style="font-size:1.3rem;font-weight:900;color:#e2e8f0;">{_sel_match}</div>
            <div style="font-size:0.85rem;color:#a0aec0;margin-top:4px;">
              {_match_date.strftime("%d/%m/%Y") if hasattr(_match_date,"strftime") else str(_match_date)[:10]}
            </div>
          </div>
          <div style="margin-left:auto;font-size:0.8rem;color:#718096;">
            {len(_df_m)} joueurs · données GPS Catapult
          </div>
        </div>
        """, unsafe_allow_html=True)

    # Ajouter groupe position
    if "Position Name" in _df_m.columns:
        _df_m["Groupe"] = _df_m["Position Name"].apply(groupe_position)
    else:
        _df_m["Groupe"] = "Arrière"

    # ═══════════════════════════════════════════════════════
    #  KPIs ÉQUIPE — Global + Par quart-temps
    # ═══════════════════════════════════════════════════════
    st.markdown("<div class='section-title'>Métriques équipe — ce match</div>",
                unsafe_allow_html=True)

    _kpi_gps_defs = [
        ("m/min moyen",       "m/min",             "{:.1f}",     "#63b3ed", "mean"),
        ("Distance totale",   "Distance totale",   "{:.0f} m",   "#68d391", "mean"),
        ("%HI moyen",         "%HI",               "{:.1f}%",    "#f6ad55", "mean"),
        ("Vmax max",          "Vmax",              "{:.1f} km/h","#b794f4", "max"),
        ("Player Load moyen", "Total Player Load", "{:.0f}",     "#4fd1c5", "mean"),
    ]

    def _render_kpi_cards(df_src, suffix=""):
        """Affiche les 5 cartes KPI depuis df_src."""
        _cols = st.columns(5)
        for _col, (_lbl, _key, _fmt, _clr, _agg) in zip(_cols, _kpi_gps_defs):
            with _col:
                if _key not in df_src.columns or df_src[_key].dropna().empty:
                    _disp = "—"
                elif _agg == "max":
                    _disp = _fmt.format(df_src[_key].max())
                else:
                    _disp = _fmt.format(df_src[_key].mean())
                st.markdown(f"""
                <div style="background:#111827;border-radius:12px;padding:16px 10px;
                            text-align:center;border:1px solid #1f2937;">
                  <div style="font-size:0.62rem;color:#6b7280;text-transform:uppercase;
                              letter-spacing:1.5px;margin-bottom:8px;">{_lbl}{suffix}</div>
                  <div style="font-size:1.5rem;font-weight:900;color:{_clr};">{_disp}</div>
                </div>""", unsafe_allow_html=True)

    # Onglets Global / Quart-temps
    _tab_global, _tab_qt = st.tabs(["🏉 Global (session)", "⏱ Par quart-temps"])

    with _tab_global:
        _render_kpi_cards(_df_m)

    with _tab_qt:
        # Charger les données quart-temps
        with st.spinner("Chargement des quart-temps…"):
            _df_qt_all = load_gps_quarters(_gps_path)
        if not _df_qt_all.empty and "Activity Name" in _df_qt_all.columns:
            _df_qt_m = _df_qt_all[_df_qt_all["Activity Name"] == _sel_match].copy()
        else:
            _df_qt_m = pd.DataFrame()

        if _df_qt_m.empty:
            st.info("Aucune donnée Quart-temps disponible pour ce match.")
        else:
            # ── Cartes KPI par quart ──────────────────────────────────
            _q_colors = {"Q1": "#63b3ed", "Q2": "#68d391",
                         "Q3": "#f6ad55", "Q4": "#fc8181"}
            _q_labels = {"Q1": "Q1 — 1ᵉʳ quart", "Q2": "Q2 — 2ᵉ quart",
                         "Q3": "Q3 — 3ᵉ quart",  "Q4": "Q4 — 4ᵉ quart"}

            for _q in ["Q1", "Q2", "Q3", "Q4"]:
                _df_q = _df_qt_m[_df_qt_m["Quart"] == _q]
                if _df_q.empty:
                    continue
                st.markdown(
                    f"<div style='margin:14px 0 6px 0;font-size:0.85rem;font-weight:700;"
                    f"color:{_q_colors[_q]};letter-spacing:1px;'>{_q_labels[_q]}</div>",
                    unsafe_allow_html=True)
                _render_kpi_cards(_df_q, suffix=f" {_q}")

            # ── Graphique évolution des métriques par quart ───────────
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown("<div class='section-title'>Évolution par quart-temps</div>",
                        unsafe_allow_html=True)

            _qt_metrics = ["m/min", "%HI", "Vmax", "Total Player Load"]
            _qt_metric_labels = {
                "m/min":              "m/min",
                "%HI":                "% Haute Intensité",
                "Vmax":               "Vmax (km/h)",
                "Total Player Load":  "Player Load",
            }
            _qt_colors_lines = ["#63b3ed", "#f6ad55", "#b794f4", "#4fd1c5"]

            _qt_agg = (
                _df_qt_m.groupby("Quart")[_qt_metrics]
                .agg({"m/min": "mean", "%HI": "mean",
                      "Vmax": "max", "Total Player Load": "mean"})
                .reindex(["Q1", "Q2", "Q3", "Q4"])
                .reset_index()
            )

            _fig_qt = go.Figure()
            for _metric, _clr in zip(_qt_metrics, _qt_colors_lines):
                if _metric not in _qt_agg.columns:
                    continue
                _fig_qt.add_trace(go.Scatter(
                    x=_qt_agg["Quart"],
                    y=_qt_agg[_metric],
                    mode="lines+markers+text",
                    name=_qt_metric_labels.get(_metric, _metric),
                    line=dict(color=_clr, width=2.5),
                    marker=dict(size=10, color=_clr),
                    text=[f"{v:.1f}" if pd.notna(v) else "" for v in _qt_agg[_metric]],
                    textposition="top center",
                    textfont=dict(size=11, color=_clr),
                    hovertemplate=f"<b>%{{x}}</b><br>{_qt_metric_labels.get(_metric, _metric)}: %{{y:.1f}}<extra></extra>",
                ))

            _fig_qt.update_layout(
                template="plotly_dark",
                paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
                height=380,
                legend=dict(orientation="h", yanchor="bottom", y=1.02,
                            xanchor="right", x=1, font=dict(size=11)),
                margin=dict(l=40, r=20, t=40, b=40),
                xaxis=dict(title="Quart-temps", gridcolor="#1f2937"),
                yaxis=dict(title="Valeur", gridcolor="#1f2937"),
            )
            st.plotly_chart(_fig_qt, use_container_width=True)

            # ── Distance totale par quart (barres) ────────────────────
            if "Distance totale" in _df_qt_m.columns:
                _qt_dist = (
                    _df_qt_m.groupby("Quart")["Distance totale"]
                    .mean()
                    .reindex(["Q1", "Q2", "Q3", "Q4"])
                    .reset_index()
                )
                _fig_dist = go.Figure(go.Bar(
                    x=_qt_dist["Quart"],
                    y=_qt_dist["Distance totale"],
                    marker_color=["#63b3ed", "#68d391", "#f6ad55", "#fc8181"],
                    text=[f"{v:.0f} m" if pd.notna(v) else "" for v in _qt_dist["Distance totale"]],
                    textposition="outside",
                    hovertemplate="<b>%{x}</b><br>Distance moy. : %{y:.0f} m<extra></extra>",
                ))
                _fig_dist.update_layout(
                    title=dict(text="Distance moyenne par quart-temps (m)", font=dict(size=13)),
                    template="plotly_dark",
                    paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
                    height=320, showlegend=False,
                    margin=dict(l=40, r=20, t=50, b=40),
                    xaxis=dict(title="Quart-temps"),
                    yaxis=dict(title="Distance (m)", gridcolor="#1f2937"),
                )
                st.plotly_chart(_fig_dist, use_container_width=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ═══════════════════════════════════════════════════════
    #  m/min — match vs moyenne saison
    # ═══════════════════════════════════════════════════════
    st.markdown("<div class='section-title'>m/min — match vs moyenne saison</div>",
                unsafe_allow_html=True)

    if "m/min" in _df_gps.columns and "Player Name" in _df_gps.columns:
        _avg_saison = _df_gps.groupby("Player Name")["m/min"].mean().reset_index()
        _avg_saison.columns = ["Player Name", "m/min_saison"]
        _df_mmin = _df_m[["Player Name", "m/min", "Groupe"]].merge(_avg_saison, on="Player Name", how="left")
        _df_mmin = _df_mmin.dropna(subset=["m/min"]).sort_values("m/min", ascending=False)

        _fig_mmin = go.Figure()
        _fig_mmin.add_trace(go.Bar(
            name="Ce match",
            x=_df_mmin["Player Name"], y=_df_mmin["m/min"],
            marker_color="#63b3ed",
            text=[f"{v:.1f}" for v in _df_mmin["m/min"]],
            textposition="outside", textfont=dict(size=10),
            hovertemplate="<b>%{x}</b><br>m/min match : %{y:.1f}<extra></extra>",
        ))
        _fig_mmin.add_trace(go.Scatter(
            name="Moyenne saison",
            x=_df_mmin["Player Name"], y=_df_mmin["m/min_saison"],
            mode="markers",
            marker=dict(color="#f6ad55", size=10, symbol="diamond"),
            hovertemplate="<b>%{x}</b><br>Moy. saison : %{y:.1f}<extra></extra>",
        ))
        _fig_mmin.update_layout(
            template="plotly_dark", paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
            height=360, barmode="group",
            margin=dict(l=20, r=20, t=20, b=80),
            xaxis=dict(tickangle=-35, tickfont=dict(size=10, color="#a0aec0")),
            yaxis=dict(title="m/min", gridcolor="#1f2937"),
            legend=dict(orientation="h", y=1.08, x=0.5, xanchor="center"),
        )
        st.plotly_chart(_fig_mmin, use_container_width=True)

    # ═══════════════════════════════════════════════════════
    #  RHIE — match vs meilleure valeur saison
    # ═══════════════════════════════════════════════════════
    st.markdown("<div class='section-title'>RHIE — efforts répétés haute intensité</div>",
                unsafe_allow_html=True)

    _rhie_col = "RHIE Efforts Per Bout - Max"
    if _rhie_col in _df_gps.columns and "Player Name" in _df_gps.columns:
        _best_rhie = _df_gps.groupby("Player Name")[_rhie_col].max().reset_index()
        _best_rhie.columns = ["Player Name", "RHIE_best"]
        _df_rhie = _df_m[["Player Name", _rhie_col, "Groupe"]].merge(_best_rhie, on="Player Name", how="left")
        _df_rhie = _df_rhie.dropna(subset=[_rhie_col]).sort_values(_rhie_col, ascending=False)

        _fig_rhie = go.Figure()
        _fig_rhie.add_trace(go.Bar(
            name="Ce match",
            x=_df_rhie["Player Name"], y=_df_rhie[_rhie_col],
            marker_color="#b794f4",
            text=[f"{int(v)}" for v in _df_rhie[_rhie_col]],
            textposition="outside",
            hovertemplate="<b>%{x}</b><br>RHIE match : %{y}<extra></extra>",
        ))
        _fig_rhie.add_trace(go.Scatter(
            name="Meilleure valeur saison",
            x=_df_rhie["Player Name"], y=_df_rhie["RHIE_best"],
            mode="markers",
            marker=dict(color="#68d391", size=11, symbol="star"),
            hovertemplate="<b>%{x}</b><br>Meilleur RHIE : %{y}<extra></extra>",
        ))
        _fig_rhie.update_layout(
            template="plotly_dark", paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
            height=360,
            margin=dict(l=20, r=20, t=20, b=80),
            xaxis=dict(tickangle=-35, tickfont=dict(size=10, color="#a0aec0")),
            yaxis=dict(title="RHIE (bouts max)", gridcolor="#1f2937"),
            legend=dict(orientation="h", y=1.08, x=0.5, xanchor="center"),
        )
        st.plotly_chart(_fig_rhie, use_container_width=True)

    # ═══════════════════════════════════════════════════════
    #  Vmax — Avants vs Arrières
    # ═══════════════════════════════════════════════════════
    st.markdown("<div class='section-title'>Vmax — Avants & Arrières</div>",
                unsafe_allow_html=True)

    _col_vmax_a, _col_vmax_b = st.columns(2)
    for _grp, _clr, _col_v in [("Avant", "#63b3ed", _col_vmax_a),
                                ("Arrière", "#fc8181", _col_vmax_b)]:
        with _col_v:
            _dfg = _df_m[_df_m["Groupe"] == _grp].sort_values("Vmax", ascending=False)
            if not _dfg.empty and "Vmax" in _dfg.columns:
                _fig_vx = go.Figure(go.Bar(
                    x=_dfg["Vmax"], y=_dfg["Player Name"],
                    orientation="h",
                    marker=dict(
                        color=_dfg["Vmax"],
                        colorscale=[[0,"#fc8181"],[0.5,"#f6ad55"],[1,"#68d391"]],
                        cmin=_dfg["Vmax"].min() * 0.9,
                        cmax=_dfg["Vmax"].max(),
                    ),
                    text=[f"{v:.1f} km/h" for v in _dfg["Vmax"]],
                    textposition="outside",
                    hovertemplate="<b>%{y}</b><br>Vmax : %{x:.1f} km/h<extra></extra>",
                ))
                _fig_vx.update_layout(
                    template="plotly_dark", paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
                    height=max(280, len(_dfg) * 38),
                    title=dict(text=f"{_grp}s", font=dict(size=13, color=_clr), x=0.5),
                    margin=dict(l=10, r=80, t=40, b=20),
                    xaxis=dict(title="km/h", gridcolor="#1f2937"),
                    yaxis=dict(tickfont=dict(size=10, color="#e2e8f0")),
                )
                st.plotly_chart(_fig_vx, use_container_width=True)

    # ═══════════════════════════════════════════════════════
    #  %Vmax — tableau avec code couleur
    # ═══════════════════════════════════════════════════════
    st.markdown("<div class='section-title'>%Vmax — indicateur d'intensité</div>",
                unsafe_allow_html=True)

    st.markdown("""
    <div style="display:flex;gap:20px;margin-bottom:14px;flex-wrap:wrap;">
      <div style="background:#1a2a1a;border:1px solid #68d391;border-radius:8px;
                  padding:8px 16px;font-size:0.8rem;color:#68d391;">
        🟢 <b>Avant</b> : ≥ 85% &nbsp;|&nbsp; <b>Arrière</b> : ≥ 90%
      </div>
      <div style="background:#2a1e0e;border:1px solid #f6ad55;border-radius:8px;
                  padding:8px 16px;font-size:0.8rem;color:#f6ad55;">
        🟠 <b>Avant</b> : 80–85% &nbsp;|&nbsp; <b>Arrière</b> : 80–90%
      </div>
      <div style="background:#2a1010;border:1px solid #fc8181;border-radius:8px;
                  padding:8px 16px;font-size:0.8rem;color:#fc8181;">
        🔴 <b>Tous</b> : &lt; 80%
      </div>
    </div>
    """, unsafe_allow_html=True)

    if "Vmax%" in _df_m.columns and "Player Name" in _df_m.columns:
        _df_vp = _df_m[["Player Name", "Position Name", "Groupe", "Vmax", "Vmax%"]].dropna(subset=["Vmax%"]).copy()
        _df_vp = _df_vp.sort_values("Vmax%", ascending=False)

        _vmax_cards = ""
        for _, _row in _df_vp.iterrows():
            _pct = float(_row["Vmax%"])
            _grp = str(_row.get("Groupe", "Arrière"))
            _clr, _icon = vmax_color(_pct, _grp)
            _vmax_cards += f"""
            <div style="background:#111827;border-radius:10px;padding:12px 16px;
                        border-left:4px solid {_clr};margin-bottom:8px;
                        display:flex;align-items:center;gap:16px;">
              <div style="font-size:1.3rem;">{_icon}</div>
              <div style="flex:1;">
                <div style="font-size:0.9rem;font-weight:700;color:#e2e8f0;">{_row["Player Name"]}</div>
                <div style="font-size:0.75rem;color:#718096;">{_row.get("Position Name","—")} · {_grp}</div>
              </div>
              <div style="text-align:right;">
                <div style="font-size:1.6rem;font-weight:900;color:{_clr};">{_pct:.1f}%</div>
                <div style="font-size:0.7rem;color:#718096;">Vmax {float(_row["Vmax"]):.1f} km/h</div>
              </div>
            </div>"""

        _col_vp1, _col_vp2 = st.columns(2)
        _rows_list = _df_vp.iterrows()
        _all_rows = list(_df_vp.iterrows())
        _mid = len(_all_rows) // 2 + len(_all_rows) % 2

        def _cards_html(rows_slice):
            html = ""
            for _, _row in rows_slice:
                _pct = float(_row["Vmax%"])
                _grp = str(_row.get("Groupe", "Arrière"))
                _clr2, _icon2 = vmax_color(_pct, _grp)
                html += f"""
                <div style="background:#111827;border-radius:10px;padding:12px 16px;
                            border-left:4px solid {_clr2};margin-bottom:8px;
                            display:flex;align-items:center;gap:16px;">
                  <div style="font-size:1.3rem;">{_icon2}</div>
                  <div style="flex:1;">
                    <div style="font-size:0.9rem;font-weight:700;color:#e2e8f0;">{_row["Player Name"]}</div>
                    <div style="font-size:0.75rem;color:#718096;">{_row.get("Position Name","—")} · {_grp}</div>
                  </div>
                  <div style="text-align:right;">
                    <div style="font-size:1.6rem;font-weight:900;color:{_clr2};">{_pct:.1f}%</div>
                    <div style="font-size:0.7rem;color:#718096;">Vmax {float(_row["Vmax"]):.1f} km/h</div>
                  </div>
                </div>"""
            return html

        with _col_vp1:
            st.markdown(_cards_html(_all_rows[:_mid]), unsafe_allow_html=True)
        with _col_vp2:
            st.markdown(_cards_html(_all_rows[_mid:]), unsafe_allow_html=True)

    # ═══════════════════════════════════════════════════════
    #  %HI — graphe par joueur
    # ═══════════════════════════════════════════════════════
    st.markdown("<div class='section-title'>%HI — part de distance en haute intensité</div>",
                unsafe_allow_html=True)

    if "%HI" in _df_m.columns:
        _df_hi = _df_m[["Player Name", "%HI", "Groupe", "HI", "Distance totale"]].dropna(subset=["%HI"]).sort_values("%HI", ascending=True)
        _hi_colors = [
            "#68d391" if float(v) >= 15 else ("#f6ad55" if float(v) >= 10 else "#fc8181")
            for v in _df_hi["%HI"]
        ]
        _fig_hi = go.Figure(go.Bar(
            x=_df_hi["%HI"], y=_df_hi["Player Name"],
            orientation="h",
            marker_color=_hi_colors,
            text=[f"{v:.1f}%" for v in _df_hi["%HI"]],
            textposition="outside",
            hovertemplate="<b>%{y}</b><br>%HI : %{x:.1f}%<br>HI abs : "
                          + _df_hi["HI"].astype(str) + " m<extra></extra>",
        ))
        _fig_hi.update_layout(
            template="plotly_dark", paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
            height=max(320, len(_df_hi) * 28),
            margin=dict(l=10, r=60, t=20, b=20),
            xaxis=dict(title="%HI", ticksuffix="%", gridcolor="#1f2937"),
            yaxis=dict(tickfont=dict(size=10, color="#e2e8f0")),
        )
        st.plotly_chart(_fig_hi, use_container_width=True)

    # ═══════════════════════════════════════════════════════
    #  DONUT Distance Courue / Marchée — par joueur
    # ═══════════════════════════════════════════════════════
    st.markdown("<div class='section-title'>Distance Courue vs Marchée — par joueur</div>",
                unsafe_allow_html=True)

    if "Courue" in _df_m.columns and "Marchée" in _df_m.columns:
        _df_dist = _df_m[["Player Name", "Courue", "Marchée", "Groupe"]].dropna(subset=["Courue","Marchée"]).copy()
        _df_dist = _df_dist.sort_values("Courue", ascending=False)

        # Donut global équipe
        _total_courue = _df_dist["Courue"].sum()
        _total_marchee = _df_dist["Marchée"].sum()
        _total_dist = _total_courue + _total_marchee

        _col_donut_eq, _col_donut_detail = st.columns([1, 2])
        with _col_donut_eq:
            _fig_donut_eq = go.Figure(go.Pie(
                labels=["Distance courue", "Marchée"],
                values=[_total_courue, _total_marchee],
                hole=0.55,
                marker=dict(colors=["#63b3ed", "#4a5568"]),
                textinfo="label+percent",
                textfont=dict(size=11, color="#e2e8f0"),
                hovertemplate="%{label}<br>%{value:.0f} m (%{percent})<extra></extra>",
            ))
            _fig_donut_eq.update_layout(
                template="plotly_dark", paper_bgcolor="#0d1117",
                height=300, margin=dict(l=10,r=10,t=30,b=10),
                title=dict(text="Équipe entière", font=dict(size=12, color="#a0aec0"), x=0.5),
                annotations=[dict(
                    text=f"<b>{_total_dist:.0f}m</b>",
                    x=0.5, y=0.5, showarrow=False,
                    font=dict(size=14, color="#e2e8f0"),
                )],
                showlegend=True,
                legend=dict(orientation="h", y=-0.1, x=0.5, xanchor="center",
                            font=dict(size=10, color="#a0aec0")),
            )
            st.plotly_chart(_fig_donut_eq, use_container_width=True)

        with _col_donut_detail:
            # Barre empilée par joueur
            _fig_stack_dist = go.Figure()
            _fig_stack_dist.add_trace(go.Bar(
                name="Distance courue",
                y=_df_dist["Player Name"], x=_df_dist["Courue"],
                orientation="h", marker_color="#63b3ed",
                text=[f"{v:.0f}m" for v in _df_dist["Courue"]],
                textposition="inside", textfont=dict(size=9, color="#fff"),
                hovertemplate="<b>%{y}</b><br>Courue : %{x:.0f} m<extra></extra>",
            ))
            _fig_stack_dist.add_trace(go.Bar(
                name="Marchée",
                y=_df_dist["Player Name"], x=_df_dist["Marchée"],
                orientation="h", marker_color="#4a5568",
                text=[f"{v:.0f}m" for v in _df_dist["Marchée"]],
                textposition="inside", textfont=dict(size=9, color="#a0aec0"),
                hovertemplate="<b>%{y}</b><br>Marchée : %{x:.0f} m<extra></extra>",
            ))
            _fig_stack_dist.update_layout(
                barmode="stack",
                template="plotly_dark", paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
                height=max(320, len(_df_dist) * 30),
                margin=dict(l=10, r=20, t=20, b=20),
                xaxis=dict(title="Mètres", gridcolor="#1f2937"),
                yaxis=dict(tickfont=dict(size=9, color="#e2e8f0"), autorange="reversed"),
                legend=dict(orientation="h", y=1.06, x=0.5, xanchor="center",
                            font=dict(size=10, color="#e2e8f0")),
            )
            st.plotly_chart(_fig_stack_dist, use_container_width=True)

    # ═══════════════════════════════════════════════════════
    #  TABLEAU RÉCAPITULATIF GPS
    # ═══════════════════════════════════════════════════════
    with st.expander("Tableau complet GPS — ce match", expanded=False):
        _cols_show = [c for c in [
            "Player Name", "Position Name", "Groupe",
            "m/min", "Distance totale", "Courue", "Marchée",
            "HI", "%HI", "Vmax", "Vmax%",
            "RHIE Efforts Per Bout - Max", "RHIE Total Bouts",
            "Total Player Load"
        ] if c in _df_m.columns]
        _df_show = _df_m[_cols_show].copy()
        _df_show = _df_show.sort_values("m/min", ascending=False) if "m/min" in _df_show.columns else _df_show
        st.dataframe(
            _df_show.round(2),
            use_container_width=True,
            hide_index=True,
        )


# ══════════════════════════════════════════════
#  PAGE : FUSION GPS × VIDÉO
# ══════════════════════════════════════════════
elif page == "Fusion GPS×Vidéo":

    import sys as _sys
    import tempfile as _tmpf
    import io as _io

    # ── Import du moteur de fusion ──────────────────────────────────────
    _merge_path = os.path.join(os.path.dirname(__file__), "merge_gps_video.py")
    if os.path.exists(_merge_path) and _merge_path not in _sys.path:
        _sys.path.insert(0, os.path.dirname(_merge_path))
    try:
        from merge_gps_video import (
            merge_gps_video, load_gps_excel, load_video_csv,
            extract_video_stats, auto_detect_match_mapping,
            normalize_name, detect_auri_prefix
        )
        _merge_ok = True
    except ImportError as _e:
        _merge_ok = False
        st.error(f"Impossible d'importer merge_gps_video.py : {_e}")

    # ── En-tête ─────────────────────────────────────────────────────────
    st.markdown("""
    <div style="background:linear-gradient(135deg,#1a1f2e,#2d3748);
                border-radius:16px;padding:20px 28px;margin-bottom:20px;
                border:1px solid #4a5568;">
      <h2 style="color:#e2e8f0;margin:0;font-size:1.4rem;">
        🔗 Fusion GPS × Vidéo — Saison 2025-2026
      </h2>
      <p style="color:#a0aec0;margin:6px 0 0;font-size:0.85rem;">
        Fusionne automatiquement tes données Catapult et Dartfish
        pour des analyses croisant charge physique et performance tactique.
      </p>
    </div>
    """, unsafe_allow_html=True)

    if not _merge_ok:
        st.stop()

    # ── Initialisation session state ────────────────────────────────────
    if "fusion_df" not in st.session_state:
        st.session_state["fusion_df"] = None
    if "fusion_mapping" not in st.session_state:
        st.session_state["fusion_mapping"] = {}

    # ════════════════════════════════════════════════════════════
    #  ÉTAPE 1 — Upload des fichiers
    # ════════════════════════════════════════════════════════════
    st.markdown("<div class='section-title'>Étape 1 — Charger les fichiers</div>",
                unsafe_allow_html=True)

    _up_col1, _up_col2 = st.columns([1, 2])

    with _up_col1:
        st.markdown("""
        <div style="background:#0d1117;border:1px dashed #4a5568;border-radius:12px;
                    padding:16px;margin-bottom:8px;">
          <div style="font-size:0.75rem;color:#63b3ed;font-weight:700;
                      text-transform:uppercase;letter-spacing:1px;margin-bottom:8px;">
            📡 Données GPS
          </div>
        """, unsafe_allow_html=True)
        _gps_up = st.file_uploader(
            "HUB DATAS .xlsx",
            type=["xlsx"],
            key="fusion_gps_up",
            help="Fichier HUB DATAS exporté depuis Catapult",
            label_visibility="collapsed",
        )
        if _gps_up:
            st.success(f"✅ {_gps_up.name}")
        st.markdown("</div>", unsafe_allow_html=True)

    with _up_col2:
        st.markdown("""
        <div style="background:#0d1117;border:1px dashed #4a5568;border-radius:12px;
                    padding:16px;margin-bottom:8px;">
          <div style="font-size:0.75rem;color:#fc8181;font-weight:700;
                      text-transform:uppercase;letter-spacing:1px;margin-bottom:8px;">
            🎬 Données Vidéo (plusieurs matchs possibles)
          </div>
        """, unsafe_allow_html=True)
        _vid_ups = st.file_uploader(
            "Fichiers CSV Dartfish",
            type=["csv"],
            accept_multiple_files=True,
            key="fusion_vid_up",
            help="Un ou plusieurs fichiers CSV Dartfish (ex: AURI-CAB-J13.csv)",
            label_visibility="collapsed",
        )
        if _vid_ups:
            for _f in _vid_ups:
                st.success(f"✅ {_f.name}")
        st.markdown("</div>", unsafe_allow_html=True)

    # ════════════════════════════════════════════════════════════
    #  ÉTAPE 2 — Mapping des matchs
    # ════════════════════════════════════════════════════════════
    if _gps_up and _vid_ups:
        st.markdown("<div class='section-title'>Étape 2 — Correspondance des matchs</div>",
                    unsafe_allow_html=True)

        # Charger GPS temporairement pour avoir la liste des matchs
        _gps_tmp = _tmpf.NamedTemporaryFile(delete=False, suffix=".xlsx")
        _gps_tmp.write(_gps_up.read())
        _gps_tmp.flush()
        _gps_up.seek(0)

        with st.spinner("Analyse des matchs GPS…"):
            try:
                _df_gps_tmp = load_gps_excel(_gps_tmp.name)
                _gps_match_list = sorted(
                    _df_gps_tmp["Activity Name"].dropna().unique().tolist()
                )
            except Exception as _ex:
                st.error(f"Erreur lecture GPS : {_ex}")
                st.stop()

        # Sauvegarder les fichiers vidéo temporairement
        _vid_tmp_paths = {}
        for _vf in _vid_ups:
            _t = _tmpf.NamedTemporaryFile(delete=False, suffix=".csv")
            _t.write(_vf.read())
            _t.flush()
            _vid_tmp_paths[_vf.name] = _t.name
            _vf.seek(0)

        # Auto-détection
        _auto_map = auto_detect_match_mapping(
            list(_vid_tmp_paths.keys()), _gps_match_list
        )

        st.markdown("""
        <p style="color:#a0aec0;font-size:0.85rem;margin-bottom:12px;">
          L'algorithme a détecté automatiquement les correspondances.
          Vérifie et corrige si nécessaire.
        </p>
        """, unsafe_allow_html=True)

        _manual_map = {}
        for _fname, _info in _auto_map.items():
            _col_a, _col_b, _col_c = st.columns([3, 3, 1])
            with _col_a:
                st.markdown(
                    f"<div style='padding:8px 0;color:#e2e8f0;font-size:0.85rem;'>"
                    f"🎬 <b>{_fname[:40]}</b></div>",
                    unsafe_allow_html=True
                )
            with _col_b:
                _default_idx = 0
                if _info["gps_match"] and _info["gps_match"] in _gps_match_list:
                    _default_idx = _gps_match_list.index(_info["gps_match"])
                _selected = st.selectbox(
                    "Match GPS",
                    options=_gps_match_list,
                    index=_default_idx,
                    key=f"map_sel_{_fname}",
                    label_visibility="collapsed",
                )
                _manual_map[_fname] = _selected
            with _col_c:
                _badge = "🟢 AUTO" if _info["confiance"] == "AUTO" else "🔵 MANUEL"
                st.markdown(
                    f"<div style='padding:10px 0;font-size:0.7rem;color:#718096;'>"
                    f"{_badge}</div>",
                    unsafe_allow_html=True
                )

        # ════════════════════════════════════════════════════════
        #  ÉTAPE 3 — Lancer la fusion
        # ════════════════════════════════════════════════════════
        st.markdown("<div class='section-title'>Étape 3 — Lancer la fusion</div>",
                    unsafe_allow_html=True)

        _btn_col, _info_col = st.columns([1, 3])
        with _btn_col:
            _run_fusion = st.button(
                "🔗 Fusionner les données",
                use_container_width=True,
                type="primary",
            )
        with _info_col:
            st.markdown(
                f"<div style='padding:10px;color:#a0aec0;font-size:0.85rem;'>"
                f"<b>{len(_vid_ups)}</b> fichier(s) vidéo · "
                f"<b>{len(_gps_match_list)}</b> matchs GPS disponibles"
                f"</div>",
                unsafe_allow_html=True
            )

        if _run_fusion:
            with st.spinner("Fusion en cours…"):
                try:
                    _out_tmp = _tmpf.NamedTemporaryFile(
                        delete=False, suffix=".xlsx"
                    )
                    _out_path = _out_tmp.name
                    _out_tmp.close()

                    _df_result = merge_gps_video(
                        gps_path=_gps_tmp.name,
                        video_files=list(_vid_tmp_paths.values()),
                        output_path=_out_path,
                        manual_mapping={
                            os.path.basename(v): _manual_map.get(k, "")
                            for k, v in _vid_tmp_paths.items()
                        },
                    )

                    # Stocker dans la session
                    if st.session_state["fusion_df"] is not None:
                        # Accumuler avec les données existantes
                        _existing = st.session_state["fusion_df"]
                        _combined = pd.concat(
                            [_existing, _df_result], ignore_index=True
                        ).drop_duplicates(
                            subset=["joueur_norm", "gps_match"], keep="last"
                        )
                        st.session_state["fusion_df"] = _combined
                    else:
                        st.session_state["fusion_df"] = _df_result

                    st.session_state["fusion_xlsx_path"] = _out_path
                    st.success(
                        f"✅ Fusion réussie ! "
                        f"{len(_df_result)} lignes · "
                        f"{_df_result['joueur_norm'].nunique()} joueurs · "
                        f"{_df_result['gps_match'].nunique()} matchs"
                    )
                    st.rerun()

                except Exception as _ex:
                    st.error(f"Erreur lors de la fusion : {_ex}")
                    import traceback
                    st.code(traceback.format_exc())

    # ════════════════════════════════════════════════════════════
    #  RÉSULTATS — si des données sont déjà fusionnées
    # ════════════════════════════════════════════════════════════
    _df_fus = st.session_state.get("fusion_df")

    if _df_fus is not None and not _df_fus.empty:

        st.markdown("---")
        st.markdown("<div class='section-title'>Données fusionnées</div>",
                    unsafe_allow_html=True)

        # ── KPIs résumé ───────────────────────────────────────────────────
        _fus_k1, _fus_k2, _fus_k3, _fus_k4 = st.columns(4)
        for _col_fus, (_lbl_f, _val_f, _clr_f) in zip(
            [_fus_k1, _fus_k2, _fus_k3, _fus_k4],
            [
                ("Matchs fusionnés",  _df_fus["gps_match"].nunique(),   "#63b3ed"),
                ("Joueurs uniques",   _df_fus["joueur_norm"].nunique(),  "#68d391"),
                ("Lignes totales",    len(_df_fus),                      "#f6ad55"),
                ("Colonnes",          len(_df_fus.columns),              "#b794f4"),
            ]
        ):
            with _col_fus:
                st.markdown(f"""
                <div style="background:#111827;border-radius:12px;padding:14px 10px;
                            text-align:center;border:1px solid #1f2937;">
                  <div style="font-size:0.65rem;color:#6b7280;text-transform:uppercase;
                              letter-spacing:1.5px;margin-bottom:6px;">{_lbl_f}</div>
                  <div style="font-size:1.6rem;font-weight:900;color:{_clr_f};">{_val_f}</div>
                </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Filtres ───────────────────────────────────────────────────────
        _f_col1, _f_col2, _f_col3 = st.columns(3)
        with _f_col1:
            _fus_match_opts = ["Tous"] + sorted(_df_fus["gps_match"].dropna().unique().tolist())
            _sel_fus_match = st.selectbox("Filtrer par match", _fus_match_opts, key="fus_match_filter")
        with _f_col2:
            _fus_joueur_opts = ["Tous"] + sorted(_df_fus["joueur_norm"].dropna().unique().tolist())
            _sel_fus_joueur = st.selectbox("Filtrer par joueur", _fus_joueur_opts, key="fus_joueur_filter")
        with _f_col3:
            if "Position Name" in _df_fus.columns:
                _fus_pos_opts = ["Tous"] + sorted(_df_fus["Position Name"].dropna().unique().tolist())
                _sel_fus_pos = st.selectbox("Filtrer par poste", _fus_pos_opts, key="fus_pos_filter")
            else:
                _sel_fus_pos = "Tous"

        _df_fus_view = _df_fus.copy()
        if _sel_fus_match != "Tous":
            _df_fus_view = _df_fus_view[_df_fus_view["gps_match"] == _sel_fus_match]
        if _sel_fus_joueur != "Tous":
            _df_fus_view = _df_fus_view[_df_fus_view["joueur_norm"] == _sel_fus_joueur]
        if _sel_fus_pos != "Tous" and "Position Name" in _df_fus_view.columns:
            _df_fus_view = _df_fus_view[_df_fus_view["Position Name"] == _sel_fus_pos]

        # ── Graphe croisé : charge GPS × performance vidéo ────────────────
        if "gps_m/min" in _df_fus_view.columns and "vid_plq_pct" in _df_fus_view.columns:
            _df_scatter = _df_fus_view.dropna(subset=["gps_m/min", "vid_plq_pct"])
            if not _df_scatter.empty:
                st.markdown(
                    "<div class='section-title'>Charge physique (m/min) × Réussite plaquage</div>",
                    unsafe_allow_html=True
                )
                _fig_scatter = go.Figure()
                for _match_s in _df_scatter["gps_match"].unique():
                    _ds = _df_scatter[_df_scatter["gps_match"] == _match_s]
                    _fig_scatter.add_trace(go.Scatter(
                        x=_ds["gps_m/min"], y=_ds["vid_plq_pct"],
                        mode="markers+text",
                        name=str(_match_s),
                        text=_ds["joueur_norm"].str.split().str[-1],
                        textposition="top center",
                        textfont=dict(size=9),
                        marker=dict(size=11),
                        hovertemplate=(
                            "<b>%{text}</b><br>"
                            "m/min : %{x:.1f}<br>"
                            "% Plaquages : %{y:.1f}%"
                            "<extra>" + str(_match_s) + "</extra>"
                        ),
                    ))
                _fig_scatter.update_layout(
                    template="plotly_dark",
                    paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
                    height=420,
                    margin=dict(l=50, r=20, t=30, b=50),
                    xaxis=dict(title="m/min (charge GPS)", gridcolor="#1f2937"),
                    yaxis=dict(title="% Plaquages réussis (vidéo)",
                               ticksuffix="%", gridcolor="#1f2937"),
                    legend=dict(orientation="h", y=1.06, x=0.5,
                                xanchor="center", font=dict(size=10)),
                )
                # Ligne médiane
                _med_mmin = _df_scatter["gps_m/min"].median()
                _med_plq  = _df_scatter["vid_plq_pct"].median()
                _fig_scatter.add_vline(x=_med_mmin, line_dash="dot",
                                       line_color="#4a5568", line_width=1.5)
                _fig_scatter.add_hline(y=_med_plq, line_dash="dot",
                                       line_color="#4a5568", line_width=1.5)
                st.plotly_chart(_fig_scatter, use_container_width=True)

        # ── Graphe signal fatigue ──────────────────────────────────────────
        if "signal_fatigue" in _df_fus_view.columns and "joueur_norm" in _df_fus_view.columns:
            _surcharge = _df_fus_view[_df_fus_view["signal_fatigue"] == "⚠ SURCHARGE"]
            if not _surcharge.empty:
                st.markdown(
                    "<div class='section-title'>⚠ Joueurs en signal de surcharge</div>",
                    unsafe_allow_html=True
                )
                _sur_count = _surcharge.groupby("joueur_norm").size().reset_index(name="Nb matchs surcharge")
                _sur_count = _sur_count.sort_values("Nb matchs surcharge", ascending=False)
                _fig_sur = go.Figure(go.Bar(
                    y=_sur_count["joueur_norm"], x=_sur_count["Nb matchs surcharge"],
                    orientation="h",
                    marker_color="#fc8181",
                    text=_sur_count["Nb matchs surcharge"],
                    textposition="outside",
                    hovertemplate="<b>%{y}</b><br>Matchs surcharge : %{x}<extra></extra>",
                ))
                _fig_sur.update_layout(
                    template="plotly_dark",
                    paper_bgcolor="#0d1117", plot_bgcolor="#0d1117",
                    height=max(250, len(_sur_count) * 30),
                    margin=dict(l=10, r=40, t=20, b=20),
                    xaxis=dict(title="Nb matchs ⚠", gridcolor="#1f2937",
                               tickformat="d"),
                    yaxis=dict(tickfont=dict(size=10, color="#e2e8f0")),
                )
                st.plotly_chart(_fig_sur, use_container_width=True)

        # ── Tableau complet filtré ─────────────────────────────────────────
        st.markdown("<div class='section-title'>Tableau fusionné complet</div>",
                    unsafe_allow_html=True)

        # Colonnes à afficher en priorité
        _priority_cols = [
            "gps_match", "joueur_norm", "Position Name",
            "gps_m/min", "gps_Vmax%", "gps_RHIE Efforts Per Bout - Max",
            "gps_%HI", "gps_Total Player Load",
            "vid_Plaquages", "vid_plq_pct",
            "vid_Rucks", "vid_Contacts", "vid_Franchissements",
            "vid_actions_Q1", "vid_actions_Q3",
            "signal_fatigue",
        ]
        _show_cols = [c for c in _priority_cols if c in _df_fus_view.columns]
        _remaining = [c for c in _df_fus_view.columns if c not in _show_cols]

        _all_shown = st.checkbox("Afficher toutes les colonnes", value=False,
                                 key="fus_show_all")
        _cols_to_show = (_show_cols + _remaining) if _all_shown else _show_cols

        st.dataframe(
            _df_fus_view[_cols_to_show].round(2),
            use_container_width=True,
            hide_index=True,
            column_config={
                "vid_plq_pct":  st.column_config.NumberColumn("% Plaquages", format="%.1f%%"),
                "gps_m/min":    st.column_config.NumberColumn("m/min", format="%.1f"),
                "gps_Vmax%":    st.column_config.NumberColumn("Vmax%", format="%.1f%%"),
                "gps_%HI":      st.column_config.NumberColumn("%HI", format="%.1f%%"),
                "signal_fatigue": st.column_config.TextColumn("Signal"),
            }
        )

        # ── Export ────────────────────────────────────────────────────────
        st.markdown("<br>", unsafe_allow_html=True)
        _exp_col1, _exp_col2, _exp_col3 = st.columns([1, 1, 2])

        with _exp_col1:
            # Export Excel du dataset complet
            _xlsx_buf = _io.BytesIO()
            with pd.ExcelWriter(_xlsx_buf, engine="openpyxl") as _wr:
                _df_fus.to_excel(_wr, sheet_name="Fusion", index=False)
                if "gps_match" in _df_fus.columns:
                    _map_log = _df_fus[["gps_match", "joueur_norm"]].drop_duplicates()
                    _map_log.to_excel(_wr, sheet_name="Mapping", index=False)
            _xlsx_buf.seek(0)
            st.download_button(
                label="📥 Télécharger Excel (saison complète)",
                data=_xlsx_buf,
                file_name=f"fusion_gps_video_saison.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        with _exp_col2:
            # Export CSV filtré
            _csv_buf = _df_fus_view[_cols_to_show].to_csv(index=False).encode("utf-8-sig")
            st.download_button(
                label="📥 Télécharger CSV (vue filtrée)",
                data=_csv_buf,
                file_name="fusion_vue_filtree.csv",
                mime="text/csv",
                use_container_width=True,
            )

        with _exp_col3:
            if st.button("🗑 Effacer les données fusionnées", use_container_width=True):
                st.session_state["fusion_df"] = None
                st.rerun()

    elif not (_gps_up and _vid_ups):
        st.markdown("""
        <div style="text-align:center;padding:60px 0;">
          <div style="font-size:3rem;">🔗</div>
          <h3 style="color:#e2e8f0;">Aucune donnée chargée</h3>
          <p style="color:#a0aec0;max-width:420px;margin:12px auto;">
            Charge ton fichier <strong>HUB DATAS .xlsx</strong> (GPS)
            et un ou plusieurs <strong>CSV Dartfish</strong> (vidéo)
            pour démarrer la fusion.
          </p>
          <p style="color:#4a5568;font-size:0.8rem;">
            Les données fusionnées restent disponibles tout au long
            de la session — tu peux ajouter des matchs au fur et à mesure.
          </p>
        </div>
        """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  PAGE : SUIVI DE SAISON — multi-matchs + progression joueur
# ══════════════════════════════════════════════════════════════════════════════
if page == "Suivi de Saison":

    st.markdown(
        "<h2 style='color:#e2e8f0;font-weight:900;margin-bottom:4px;'>"
        "📅 Suivi de Saison</h2>"
        "<p style='color:#6b7280;font-size:0.9rem;margin-bottom:20px;'>"
        "Comparez plusieurs matchs et visualisez l'évolution des joueurs dans le temps.</p>",
        unsafe_allow_html=True,
    )

    # ── Charger tous les CSV disponibles ────────────────────────────────────
    _all_csv_files = sorted(glob.glob(os.path.join(_DATA_DIR, "*.csv")))
    _match_names   = [os.path.splitext(os.path.basename(f))[0] for f in _all_csv_files]
    _name_to_path  = dict(zip(_match_names, _all_csv_files))

    if not _all_csv_files:
        st.warning("Aucun fichier CSV trouvé dans le dossier data/. Charge des matchs d'abord.")
        st.stop()

    # ── Sélecteur multi-matchs ───────────────────────────────────────────────
    st.markdown("### 🗂 Sélection des matchs")
    _sel_matches = st.multiselect(
        "Choisir les matchs à analyser",
        options=_match_names,
        default=_match_names[:min(5, len(_match_names))],
        help="Sélectionne plusieurs matchs pour les comparer",
    )

    if not _sel_matches:
        st.info("Sélectionne au moins un match pour commencer.")
        st.stop()

    # ── Chargement des DataFrames sélectionnés ──────────────────────────────
    @st.cache_data(show_spinner="Chargement des matchs…")
    def _load_multi(paths_tuple):
        frames = {}
        for name, path in paths_tuple:
            try:
                _df = pd.read_csv(path, sep=None, engine="python", encoding="utf-8-sig")
                _df.columns = [c.strip() for c in _df.columns]
                frames[name] = _df
            except Exception:
                pass
        return frames

    _loaded = _load_multi(tuple(
        (nm, _name_to_path[nm]) for nm in _sel_matches
    ))

    if not _loaded:
        st.error("Impossible de charger les fichiers sélectionnés.")
        st.stop()

    # ── Clés d'actions disponibles ───────────────────────────────────────────
    _ACTION_KEYS = [
        "Plaquages", "Passes", "Rucks", "Courses", "Melees",
        "Touches", "Gratteurs", "Porteurs",
    ]

    # Déterminer les équipes Aurillac dans chaque match
    def _get_auri_team(frame):
        """Retourne le nom de l'équipe Aurillacois dans ce dataframe."""
        for _t in frame["Nom de la ligne"].dropna().unique():
            _t = str(_t).strip()
            if "aurillac" in _t.lower() and " - " in _t:
                return _t.split(" - ")[0].strip()
        return None

    # ────────────────────────────────────────────────────────────────────────
    #  ONGLETS
    # ────────────────────────────────────────────────────────────────────────
    _tab1, _tab2, _tab3 = st.tabs([
        "📊 Comparaison matchs",
        "📈 Progression joueurs",
        "🗺️ Carte cumulée",
    ])

    # ════════════════════════════════════════════════════════════════════════
    #  ONGLET 1 — Comparaison stats équipe par match
    # ════════════════════════════════════════════════════════════════════════
    with _tab1:
        st.markdown("#### Statistiques équipe match par match")

        # Sélecteur d'action
        _cmp_action = st.selectbox(
            "Action à comparer",
            _ACTION_KEYS,
            key="cmp_action",
        )
        _cmp_team_filter = st.radio(
            "Équipe",
            ["Aurillacois", "Adversaire", "Les deux"],
            horizontal=True,
            key="cmp_team_filter",
        )

        # Construire un DataFrame de comparaison
        _cmp_rows = []
        for _nm, _fr in _loaded.items():
            _auri = _get_auri_team(_fr)
            _all_teams = set()
            for _line in _fr["Nom de la ligne"].dropna().unique():
                _line = str(_line).strip()
                if f" - {_cmp_action}" in _line:
                    _all_teams.add(_line.replace(f" - {_cmp_action}", "").strip())
            for _team in _all_teams:
                _cnt = len(_fr[_fr["Nom de la ligne"].str.strip() == f"{_team} - {_cmp_action}"])
                _result_col = None
                for _c in _fr.columns:
                    if "resultat" in _c.lower():
                        _result_col = _c
                        break
                _pos = 0
                if _result_col:
                    _pos = len(_fr[
                        (_fr["Nom de la ligne"].str.strip() == f"{_team} - {_cmp_action}") &
                        (_fr[_result_col].astype(str).str.lower().str.strip() == "positif")
                    ])
                _is_auri = (_auri and _team == _auri)
                _cmp_rows.append({
                    "Match":    _nm,
                    "Équipe":   _team,
                    "Type":     "Aurillacois" if _is_auri else "Adversaire",
                    "Total":    _cnt,
                    "Réussis":  _pos,
                    "Taux %":   round(_pos / _cnt * 100, 1) if _cnt > 0 else 0,
                })

        if not _cmp_rows:
            st.info(f"Aucune donnée trouvée pour « {_cmp_action} » dans les matchs sélectionnés.")
        else:
            _cmp_df = pd.DataFrame(_cmp_rows)
            if _cmp_team_filter != "Les deux":
                _cmp_df = _cmp_df[_cmp_df["Type"] == _cmp_team_filter]

            import plotly.express as px

            # Graphique Total
            _fig_cmp = px.bar(
                _cmp_df, x="Match", y="Total", color="Type",
                barmode="group",
                color_discrete_map={"Aurillacois": COLOR_A, "Adversaire": COLOR_B},
                title=f"{_cmp_action} — Total par match",
                labels={"Total": "Nombre d'actions", "Match": ""},
                text="Total",
            )
            _fig_cmp.update_traces(textposition="outside", textfont_size=11)
            _fig_cmp.update_layout(
                paper_bgcolor="#0e1117", plot_bgcolor="#0e1117",
                font=dict(color="#e2e8f0"),
                xaxis=dict(tickangle=-30, gridcolor="#1f2937"),
                yaxis=dict(gridcolor="#1f2937"),
                legend=dict(bgcolor="rgba(0,0,0,0.4)"),
                margin=dict(t=50, b=80),
                height=400,
            )
            st.plotly_chart(_fig_cmp, use_container_width=True)

            # Graphique Taux de réussite (si données)
            if _cmp_df["Réussis"].sum() > 0:
                _fig_taux = px.line(
                    _cmp_df.sort_values("Match"), x="Match", y="Taux %",
                    color="Type",
                    color_discrete_map={"Aurillacois": COLOR_A, "Adversaire": COLOR_B},
                    title=f"{_cmp_action} — Taux de réussite (%)",
                    markers=True,
                    labels={"Taux %": "Taux (%)", "Match": ""},
                )
                _fig_taux.update_traces(line_width=2.5, marker_size=8)
                _fig_taux.add_hline(y=70, line_dash="dot", line_color="#68d391",
                                    annotation_text="Seuil 70 %", annotation_position="right")
                _fig_taux.update_layout(
                    paper_bgcolor="#0e1117", plot_bgcolor="#0e1117",
                    font=dict(color="#e2e8f0"),
                    xaxis=dict(tickangle=-30, gridcolor="#1f2937"),
                    yaxis=dict(gridcolor="#1f2937", range=[0, 105]),
                    legend=dict(bgcolor="rgba(0,0,0,0.4)"),
                    margin=dict(t=50, b=80),
                    height=350,
                )
                st.plotly_chart(_fig_taux, use_container_width=True)

            # Tableau récap
            st.markdown("**Tableau récapitulatif**")
            _cmp_show = _cmp_df[["Match", "Équipe", "Type", "Total", "Réussis", "Taux %"]].copy()
            _cmp_show = _cmp_show.sort_values(["Match", "Type"])
            st.dataframe(
                _cmp_show.style.background_gradient(subset=["Total"], cmap="Blues")
                               .background_gradient(subset=["Taux %"], cmap="RdYlGn", vmin=0, vmax=100),
                hide_index=True,
                use_container_width=True,
            )

    # ════════════════════════════════════════════════════════════════════════
    #  ONGLET 2 — Progression joueur
    # ════════════════════════════════════════════════════════════════════════
    with _tab2:
        st.markdown("#### Évolution d'un joueur au fil des matchs")

        _c1, _c2 = st.columns([1, 1])
        with _c1:
            _prog_action = st.selectbox(
                "Action", _ACTION_KEYS, key="prog_action"
            )
        with _c2:
            _prog_metric = st.radio(
                "Métrique", ["Nombre total", "Taux de réussite (%)"],
                horizontal=True, key="prog_metric"
            )

        # Collecter tous les joueurs disponibles dans l'action sélectionnée
        _all_players = set()
        for _nm, _fr in _loaded.items():
            _pcol = None
            for _c in _fr.columns:
                if "joueur" in _c.lower():
                    _pcol = _c
                    break
            if not _pcol:
                continue
            for _line in _fr["Nom de la ligne"].dropna().unique():
                _line = str(_line).strip()
                if _line.endswith(f" - {_prog_action}"):
                    _mask = _fr["Nom de la ligne"].str.strip() == _line
                    _ps = _fr[_mask][_pcol].dropna().astype(str).str.strip()
                    _ps = _ps[~_ps.isin(["nan", "", "None"])]
                    _all_players.update(_ps.unique())

        _all_players = sorted(_all_players)

        if not _all_players:
            st.info(f"Aucun joueur trouvé pour « {_prog_action} ».")
        else:
            # Sélecteur joueurs
            _sel_players = st.multiselect(
                "Joueurs à suivre",
                _all_players,
                default=_all_players[:min(4, len(_all_players))],
                key="prog_players",
            )
            _prog_team_filter = st.radio(
                "Équipe", ["Aurillacois", "Adversaire", "Toutes"],
                horizontal=True, key="prog_team"
            )

            if not _sel_players:
                st.info("Sélectionne au moins un joueur.")
            else:
                # Construire le DataFrame de progression
                _prog_rows = []
                for _nm, _fr in _loaded.items():
                    _auri = _get_auri_team(_fr)
                    _pcol = None
                    for _c in _fr.columns:
                        if "joueur" in _c.lower():
                            _pcol = _c
                            break
                    if not _pcol:
                        continue
                    _rcol = None
                    for _c in _fr.columns:
                        if "resultat" in _c.lower():
                            _rcol = _c
                            break
                    for _line in _fr["Nom de la ligne"].dropna().unique():
                        _line = str(_line).strip()
                        if not _line.endswith(f" - {_prog_action}"):
                            continue
                        _team = _line.replace(f" - {_prog_action}", "").strip()
                        _is_auri = (_auri and _team == _auri)
                        _type = "Aurillacois" if _is_auri else "Adversaire"
                        if _prog_team_filter != "Toutes" and _type != _prog_team_filter:
                            continue
                        _rows_act = _fr[_fr["Nom de la ligne"].str.strip() == _line].copy()
                        _rows_act[_pcol] = _rows_act[_pcol].astype(str).str.strip()
                        for _pl in _sel_players:
                            _pr = _rows_act[_rows_act[_pcol] == _pl]
                            _cnt = len(_pr)
                            if _cnt == 0:
                                continue
                            _pos = 0
                            if _rcol:
                                _pos = len(_pr[_pr[_rcol].astype(str).str.lower().str.strip() == "positif"])
                            _prog_rows.append({
                                "Match":   _nm,
                                "Joueur":  _pl,
                                "Équipe":  _type,
                                "Total":   _cnt,
                                "Réussis": _pos,
                                "Taux %":  round(_pos / _cnt * 100, 1) if _cnt > 0 else 0,
                            })

                if not _prog_rows:
                    st.info("Aucune donnée disponible pour la sélection.")
                else:
                    _prog_df = pd.DataFrame(_prog_rows)
                    _y_col = "Total" if _prog_metric == "Nombre total" else "Taux %"
                    _y_label = "Nombre d'actions" if _y_col == "Total" else "Taux de réussite (%)"

                    # ── Courbe de progression ──────────────────────────────
                    _fig_prog = px.line(
                        _prog_df.sort_values("Match"),
                        x="Match", y=_y_col, color="Joueur",
                        markers=True,
                        title=f"Progression — {_prog_action} ({_y_label})",
                        labels={_y_col: _y_label, "Match": ""},
                    )
                    _fig_prog.update_traces(line_width=2.5, marker_size=9)
                    if _y_col == "Taux %":
                        _fig_prog.add_hline(y=70, line_dash="dot", line_color="#68d391",
                                            annotation_text="70 %", annotation_position="right")
                        _fig_prog.update_yaxes(range=[0, 105])
                    _fig_prog.update_layout(
                        paper_bgcolor="#0e1117", plot_bgcolor="#0e1117",
                        font=dict(color="#e2e8f0"),
                        xaxis=dict(tickangle=-30, gridcolor="#1f2937"),
                        yaxis=dict(gridcolor="#1f2937"),
                        legend=dict(bgcolor="rgba(0,0,0,0.4)", orientation="h",
                                    yanchor="bottom", y=1.02, xanchor="left", x=0),
                        margin=dict(t=60, b=80),
                        height=420,
                    )
                    st.plotly_chart(_fig_prog, use_container_width=True)

                    # ── Heatmap joueurs × matchs ───────────────────────────
                    st.markdown("**Carte de chaleur — intensité par joueur et par match**")
                    _heat_pivot = _prog_df.pivot_table(
                        index="Joueur", columns="Match",
                        values=_y_col, aggfunc="sum", fill_value=0
                    )
                    _fig_heat = px.imshow(
                        _heat_pivot,
                        color_continuous_scale="YlOrRd",
                        title=f"Heatmap — {_prog_action} ({_y_label})",
                        labels=dict(x="Match", y="Joueur", color=_y_label),
                        aspect="auto",
                        text_auto=True,
                    )
                    _fig_heat.update_layout(
                        paper_bgcolor="#0e1117",
                        font=dict(color="#e2e8f0"),
                        xaxis=dict(tickangle=-30),
                        coloraxis_colorbar=dict(
                            title=_y_label,
                            tickfont=dict(color="#e2e8f0"),
                        ),
                        margin=dict(t=50, b=80),
                        height=max(250, 60 * len(_sel_players) + 100),
                    )
                    st.plotly_chart(_fig_heat, use_container_width=True)

                    # ── Classement cumulé ──────────────────────────────────
                    st.markdown("**Classement cumulé sur les matchs sélectionnés**")
                    _rank_df = (_prog_df.groupby("Joueur")[[_y_col, "Total"]]
                                .agg({_y_col: "mean", "Total": "sum"})
                                .reset_index()
                                .rename(columns={_y_col: f"Moy. {_y_label}",
                                                 "Total": "Total cumulé"}))
                    _rank_df = _rank_df.sort_values(f"Moy. {_y_label}", ascending=False)
                    _rank_df.index = range(1, len(_rank_df) + 1)
                    _rank_df[f"Moy. {_y_label}"] = _rank_df[f"Moy. {_y_label}"].round(1)

                    _fig_rank = px.bar(
                        _rank_df.reset_index().rename(columns={"index": "Rang"}),
                        x="Joueur", y=f"Moy. {_y_label}",
                        color=f"Moy. {_y_label}",
                        color_continuous_scale="teal",
                        title=f"Classement — Moyenne {_y_label} sur la période",
                        text=f"Moy. {_y_label}",
                    )
                    _fig_rank.update_traces(textposition="outside")
                    _fig_rank.update_layout(
                        paper_bgcolor="#0e1117", plot_bgcolor="#0e1117",
                        font=dict(color="#e2e8f0"),
                        xaxis=dict(gridcolor="#1f2937"),
                        yaxis=dict(gridcolor="#1f2937"),
                        showlegend=False,
                        margin=dict(t=50, b=60),
                        height=380,
                    )
                    st.plotly_chart(_fig_rank, use_container_width=True)

    # ════════════════════════════════════════════════════════════════════════
    #  ONGLET 3 — Carte terrain cumulée (plusieurs matchs)
    # ════════════════════════════════════════════════════════════════════════
    with _tab3:
        st.markdown("#### Carte terrain cumulée sur plusieurs matchs")
        st.caption("Superpose les actions d'un joueur sur tous les matchs sélectionnés.")

        _mc1, _mc2 = st.columns([1, 1])
        with _mc1:
            _map_action = st.selectbox("Action", _ACTION_KEYS, key="map_multi_action")
        with _mc2:
            _map_team_type = st.radio(
                "Équipe", ["Aurillacois", "Adversaire"],
                horizontal=True, key="map_multi_team"
            )

        # Collecter tous les joueurs de cette action
        _map_players = set()
        for _nm, _fr in _loaded.items():
            _pcol_m = None
            for _c in _fr.columns:
                if "joueur" in _c.lower():
                    _pcol_m = _c
                    break
            if not _pcol_m:
                continue
            _auri = _get_auri_team(_fr)
            for _line in _fr["Nom de la ligne"].dropna().unique():
                _line = str(_line).strip()
                if not _line.endswith(f" - {_map_action}"):
                    continue
                _team = _line.replace(f" - {_map_action}", "").strip()
                _is_auri = (_auri and _team == _auri)
                _type = "Aurillacois" if _is_auri else "Adversaire"
                if _type != _map_team_type:
                    continue
                _rows_m = _fr[_fr["Nom de la ligne"].str.strip() == _line]
                _ps_m = _rows_m[_pcol_m].dropna().astype(str).str.strip()
                _ps_m = _ps_m[~_ps_m.isin(["nan", "", "None"])]
                _map_players.update(_ps_m.unique())

        _map_players = sorted(_map_players)

        if not _map_players:
            st.info(f"Aucun joueur trouvé pour « {_map_action} » ({_map_team_type}).")
        else:
            _sel_map_player = st.selectbox(
                "Joueur", _map_players, key="map_multi_player"
            )
            _color_map_match = st.toggle(
                "🎨 Colorier par match", value=True, key="map_color_match"
            )
            _show_dens_multi = st.toggle(
                "🔥 Carte de densité", value=False, key="map_dens_multi"
            )

            # Collecter tous les points
            _all_pts = []
            for _nm, _fr in _loaded.items():
                _pcol_m = None
                _cx_m = _cy_m = None
                for _c in _fr.columns:
                    if "joueur" in _c.lower():
                        _pcol_m = _c
                    if "coordonnee x" in _c.lower() or "coordonnée x" in _c.lower():
                        _cx_m = _c
                    if "coordonnee y" in _c.lower() or "coordonnée y" in _c.lower():
                        _cy_m = _c
                if not (_pcol_m and _cx_m and _cy_m):
                    continue
                _auri = _get_auri_team(_fr)
                for _line in _fr["Nom de la ligne"].dropna().unique():
                    _line = str(_line).strip()
                    if not _line.endswith(f" - {_map_action}"):
                        continue
                    _team_m = _line.replace(f" - {_map_action}", "").strip()
                    _is_auri_m = (_auri and _team_m == _auri)
                    if (_map_team_type == "Aurillacois") != _is_auri_m:
                        continue
                    _rows_m = _fr[_fr["Nom de la ligne"].str.strip() == _line].copy()
                    _rows_m[_pcol_m] = _rows_m[_pcol_m].astype(str).str.strip()
                    _rows_m = _rows_m[_rows_m[_pcol_m] == _sel_map_player]
                    _rows_m["_px"] = pd.to_numeric(_rows_m[_cy_m], errors="coerce")
                    _rows_m["_py"] = pd.to_numeric(_rows_m[_cx_m], errors="coerce")
                    _rows_m = _rows_m.dropna(subset=["_px", "_py"])
                    _rows_m["Match"] = _nm
                    _all_pts.append(_rows_m[["_px", "_py", "Match"]])

            if not _all_pts:
                st.info("Aucune coordonnée disponible pour ce joueur / cette action.")
            else:
                _pts_df = pd.concat(_all_pts, ignore_index=True)

                import plotly.graph_objects as _pgo2

                _fig_multi = _pgo2.Figure()

                # Terrain
                _fig_multi.add_shape(type="rect", x0=0, y0=0, x1=100, y1=68,
                                     fillcolor="#2d5016", line=dict(color="white", width=0), layer="below")
                for _ex0, _ex1 in [(0, 10), (90, 100)]:
                    _fig_multi.add_shape(type="rect", x0=_ex0, y0=0, x1=_ex1, y1=68,
                                         fillcolor="#1e3a0c", line=dict(color="white", width=2), layer="below")
                for _lx, _lw, _dash in [(10,1.5,"solid"),(22,2,"solid"),(50,2.5,"solid"),
                                          (78,2,"solid"),(88,1.5,"solid"),(5,1,"dot"),(95,1,"dot")]:
                    _fig_multi.add_shape(type="line", x0=_lx, y0=0, x1=_lx, y1=68,
                                         line=dict(color="white", width=_lw, dash=_dash), layer="below")
                _fig_multi.add_shape(type="rect", x0=0, y0=0, x1=100, y1=68,
                                     fillcolor="rgba(0,0,0,0)", line=dict(color="white", width=2))
                for _lx, _lbl in [(5,"EN-BUT"),(16,"10m"),(36,"22m→50"),(50,"50m"),
                                   (64,"50→22m"),(84,"10m"),(95,"EN-BUT")]:
                    _fig_multi.add_annotation(x=_lx, y=66, text=_lbl, showarrow=False,
                                              font=dict(color="rgba(255,255,255,0.35)", size=8))

                # Densité optionnelle
                if _show_dens_multi and len(_pts_df) > 2:
                    _fig_multi.add_trace(_pgo2.Histogram2dContour(
                        x=_pts_df["_px"], y=_pts_df["_py"],
                        colorscale=[[0,"rgba(0,0,0,0)"],[0.3,"rgba(255,200,0,0.2)"],
                                    [0.7,"rgba(255,100,0,0.5)"],[1,"rgba(220,0,0,0.8)"]],
                        showscale=False, ncontours=14,
                        line=dict(width=0), contours=dict(showlabels=False),
                        hoverinfo="skip",
                    ))

                # Points par match
                _MATCH_COLORS = [
                    "#63b3ed","#68d391","#f6ad55","#fc8181","#b794f4",
                    "#76e4f7","#fbd38d","#fc8181","#48bb78","#ed8936",
                ]
                if _color_map_match:
                    for _mi, _match_nm in enumerate(_sel_matches):
                        _mpts = _pts_df[_pts_df["Match"] == _match_nm]
                        if _mpts.empty:
                            continue
                        _fig_multi.add_trace(_pgo2.Scatter(
                            x=_mpts["_px"], y=_mpts["_py"], mode="markers",
                            name=_match_nm,
                            marker=dict(size=10,
                                        color=_MATCH_COLORS[_mi % len(_MATCH_COLORS)],
                                        opacity=0.85, line=dict(color="white", width=1)),
                            hovertemplate=f"<b>{_match_nm}</b><br>X: %{{x:.0f}} | Y: %{{y:.0f}}<extra></extra>",
                        ))
                else:
                    _fig_multi.add_trace(_pgo2.Scatter(
                        x=_pts_df["_px"], y=_pts_df["_py"], mode="markers",
                        name=_sel_map_player,
                        marker=dict(size=10, color=COLOR_A, opacity=0.75,
                                    line=dict(color="white", width=1)),
                        hovertemplate="X: %{x:.0f} | Y: %{y:.0f}<extra></extra>",
                    ))

                _fig_multi.update_layout(
                    paper_bgcolor="#0e1117", plot_bgcolor="#2d5016",
                    font=dict(color="#e2e8f0"),
                    xaxis=dict(range=[0,100], showgrid=False, zeroline=False,
                               showticklabels=False, fixedrange=True),
                    yaxis=dict(range=[0,68], showgrid=False, zeroline=False,
                               showticklabels=False, fixedrange=True,
                               scaleanchor="x", scaleratio=0.68),
                    legend=dict(bgcolor="rgba(0,0,0,0.6)", bordercolor="#374151",
                                font=dict(color="#e2e8f0", size=11),
                                orientation="h", yanchor="bottom", y=1.01,
                                xanchor="left", x=0),
                    margin=dict(l=5, r=5, t=50, b=5),
                    height=460,
                    title=dict(
                        text=f"📍 {len(_pts_df)} actions — {_sel_map_player} "
                             f"({_map_action}) sur {len(_sel_matches)} matchs",
                        font=dict(color="#e2e8f0", size=13), x=0.5,
                    ),
                    hovermode="closest",
                )
                st.plotly_chart(_fig_multi, use_container_width=True)

                # Stats rapides
                _match_counts = _pts_df["Match"].value_counts().reset_index()
                _match_counts.columns = ["Match", "Nb actions"]
                _match_counts = _match_counts.sort_values("Match")
                st.dataframe(_match_counts, hide_index=True, use_container_width=True)

