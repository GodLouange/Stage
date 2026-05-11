"""
╔══════════════════════════════════════════════════════════════════════╗
║         ALGORITHME DE FUSION GPS × VIDÉO — Stade Aurillacois        ║
║         merge_gps_video.py                                           ║
╚══════════════════════════════════════════════════════════════════════╝

Fusionne automatiquement :
  • Les données GPS Catapult (HUB DATAS .xlsx)
  • Les données vidéo Dartfish (fichiers CSV par match)

Sortie : fichier Excel multi-feuilles avec :
  • Feuille "Fusion" : 1 ligne = 1 joueur × 1 match (GPS + vidéo)
  • Feuille "Stats_vidéo" : détail des actions par joueur/match
  • Feuille "Mapping_matchs" : correspondance vidéo ↔ GPS
  • Feuille "Mapping_joueurs" : normalisation des noms

Usage :
  python3 merge_gps_video.py \
      --gps   "HUB DATAS 25-26.xlsx" \
      --video "AURI-CAB-J13 Clip CSV.csv" "AURI-BO-J24 Clip CSV.csv" \
      --out   "fusion_saison.xlsx"

  Ou sans arguments : le script cherche automatiquement dans le
  dossier courant et dans data/.
"""

import os
import re
import glob
import argparse
import difflib
import warnings
import pandas as pd
import numpy as np
from datetime import datetime

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────
#  1. TABLE D'ABRÉVIATIONS — nom d'équipe dans le fichier vidéo → GPS
# ─────────────────────────────────────────────────────────────────────
TEAM_ALIASES = {
    # Abbréviations dans les noms de fichiers CSV
    "ASBH": "BEZIERS",
    "BO":   "BIARRITZ",
    "CAB":  "BRIVE",
    "CAR":  "CARCASSONNE",
    "COL":  "COLOMIERS",
    "AGE":  "AGEN",
    "DAX":  "DAX",
    "NEV":  "NEVERS",
    "SUA":  "SUA",       # Seyssins/Valence → à préciser
    "MDM":  "MONT DE MARSAN",
    "OYO":  "OYONNAX",
    "PRO":  "PROVENCE",
    "GRE":  "GRENOBLE",
    "VAN":  "VANNES",
    "BIA":  "BIARRITZ",
    "ANG":  "ANGOULEME",
    "BRI":  "BRIVE",
}

# Nom complet des équipes dans les CSV vidéo → slug normalisé
TEAM_FULLNAME = {
    "stade aurillacois":    "AURILLAC",
    "as beziers herault":   "BEZIERS",
    "biarritz olympique":   "BIARRITZ",
    "ca brive":             "BRIVE",
    "us colomiers":         "COLOMIERS",
    "mont de marsan":       "MONT DE MARSAN",
    "carcassonne":          "CARCASSONNE",
    "agen":                 "AGEN",
    "oyonnax":              "OYONNAX",
    "provence rugby":       "PROVENCE",
    "rc grenoble":          "GRENOBLE",
    "vannes":               "VANNES",
    "valence romans":       "VALENCE",
    "dax":                  "DAX",
    "nevers":               "NEVERS",
    "angouleme":            "ANGOULEME",
}


# ─────────────────────────────────────────────────────────────────────
#  2. NORMALISATION DES NOMS
# ─────────────────────────────────────────────────────────────────────
def normalize_name(name: str) -> str:
    """
    Convertit n'importe quelle casse en NOM_NORMALISÉ.
    'Hugo BASTARD' → 'HUGO BASTARD'
    'hugo bastard' → 'HUGO BASTARD'
    """
    if not isinstance(name, str):
        return ""
    # Supprimer accents courants
    replacements = {
        "é": "e", "è": "e", "ê": "e", "à": "a", "â": "a",
        "ù": "u", "û": "u", "î": "i", "ô": "o", "ç": "c",
        "É": "E", "È": "E", "Ê": "E", "À": "A", "Â": "A",
    }
    for src, dst in replacements.items():
        name = name.replace(src, dst)
    return name.strip().upper()


def fuzzy_match_name(name: str, candidates: list, threshold: float = 0.75) -> str | None:
    """
    Trouve le meilleur candidat pour un nom donné.
    Retourne None si aucune correspondance ≥ threshold.
    """
    name_norm = normalize_name(name)
    candidates_norm = {normalize_name(c): c for c in candidates}
    matches = difflib.get_close_matches(
        name_norm, list(candidates_norm.keys()),
        n=1, cutoff=threshold
    )
    if matches:
        return candidates_norm[matches[0]]
    return None


# ─────────────────────────────────────────────────────────────────────
#  3. AUTO-DÉTECTION DE LA CORRESPONDANCE MATCH VIDÉO ↔ GPS
# ─────────────────────────────────────────────────────────────────────
def extract_match_info_from_filename(filename: str) -> dict:
    """
    Extrait l'adversaire et le numéro de journée depuis le nom de fichier.

    Exemples :
      'AURI-CAB-J13 Clip CSV.csv' → {jouee: 'dom', adversaire: 'BRIVE', journee: 13}
      'ASBH-AURI Clip CSV.csv'    → {jouee: 'ext', adversaire: 'BEZIERS', journee: None}
      'AURI-BO-J24 Clip CSV.csv'  → {jouee: 'dom', adversaire: 'BIARRITZ', journee: 24}
    """
    name = os.path.basename(filename).upper()
    # Supprimer extension et suffixes
    name = re.sub(r'\s*CLIP\s*CSV.*$', '', name).strip()

    result = {"fichier": os.path.basename(filename), "adversaire_slug": None,
              "journee": None, "domicile": None}

    # Extraire numéro de journée
    j_match = re.search(r'J(\d+)', name)
    if j_match:
        result["journee"] = int(j_match.group(1))

    # Identifier domicile/extérieur
    parts = re.split(r'[-_]', name)
    parts = [p.strip() for p in parts if p.strip() and not re.match(r'J\d+', p)]

    if not parts:
        return result

    # Premier bloc = équipe qui reçoit (ou équipe à domicile)
    # Si AURI est en premier → domicile, sinon extérieur
    auri_kw = ["AURI", "SA", "AURILLAC", "STADE"]
    if any(k in parts[0] for k in auri_kw):
        result["domicile"] = True
        adv_parts = parts[1:]
    else:
        result["domicile"] = False
        adv_parts = parts[:1]

    # Mapper l'abréviation → slug GPS
    for p in adv_parts:
        p_clean = re.sub(r'J\d+', '', p).strip()
        if p_clean in TEAM_ALIASES:
            result["adversaire_slug"] = TEAM_ALIASES[p_clean]
            break
        # Essai direct
        for alias, slug in TEAM_ALIASES.items():
            if alias in p_clean:
                result["adversaire_slug"] = slug
                break
        if result["adversaire_slug"]:
            break

    return result


def find_gps_match(info: dict, gps_matches: list) -> str | None:
    """
    Trouve le code match GPS correspondant à partir des infos extraites.
    Stratégie : numéro de journée d'abord, puis nom adversaire.
    """
    # Stratégie 1 : journée + adversaire
    if info["journee"] and info["adversaire_slug"]:
        pattern = f"J{info['journee']} {info['adversaire_slug']}"
        for m in gps_matches:
            if pattern.upper() in m.upper():
                return m
        # Essai sans espace
        pattern2 = f"J{info['journee']}{info['adversaire_slug']}"
        for m in gps_matches:
            if pattern2.upper() in m.upper().replace(" ", ""):
                return m

    # Stratégie 2 : journée seule
    if info["journee"]:
        candidates = []
        for m in gps_matches:
            j_match = re.search(r'J(\d+)\b', m.upper())
            if j_match and int(j_match.group(1)) == info["journee"]:
                candidates.append(m)
        if len(candidates) == 1:
            return candidates[0]

    # Stratégie 3 : adversaire seul (fuzzy)
    if info["adversaire_slug"]:
        best = None
        best_score = 0
        for m in gps_matches:
            score = difflib.SequenceMatcher(
                None, info["adversaire_slug"], m.upper()
            ).ratio()
            if score > best_score:
                best_score = score
                best = m
        if best_score >= 0.5:
            return best

    return None


def auto_detect_match_mapping(video_files: list, gps_matches: list) -> dict:
    """
    Construit automatiquement le dict {video_file: gps_match_code}.
    """
    mapping = {}
    for vf in video_files:
        info = extract_match_info_from_filename(vf)
        gps_code = find_gps_match(info, gps_matches)
        mapping[os.path.basename(vf)] = {
            "gps_match":  gps_code,
            "info":       info,
            "confiance":  "AUTO" if gps_code else "NON_TROUVÉ",
        }
    return mapping


# ─────────────────────────────────────────────────────────────────────
#  4. CHARGEMENT DES DONNÉES
# ─────────────────────────────────────────────────────────────────────
def load_video_csv(filepath: str) -> pd.DataFrame | None:
    """Charge un CSV Dartfish avec gestion des encodages."""
    for enc in ["utf-16", "utf-16-le", "utf-8-sig", "latin-1", "utf-8"]:
        try:
            df = pd.read_csv(filepath, encoding=enc, sep=",",
                             low_memory=False, on_bad_lines="skip")
            df.columns = [c.strip().lstrip("\ufeff") for c in df.columns]
            if "Nom de la ligne" in df.columns:
                for col in df.select_dtypes("object").columns:
                    df[col] = df[col].astype(str).str.strip().replace("nan", pd.NA)
                return df
        except Exception:
            continue
    print(f"  ⚠ Impossible de charger : {filepath}")
    return None


def load_gps_excel(filepath: str) -> pd.DataFrame:
    """Charge le HUB DATAS Excel, filtre les sessions de match."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    sheet_name = "HUB DATAS 2.0" if "HUB DATAS 2.0" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    data = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(data[0])]
    df = pd.DataFrame(data[1:], columns=headers)
    df.columns = [c.strip() for c in df.columns]

    # Filtre : matchs uniquement, session globale
    df_match = df[
        (df["Entrainement"].astype(str).str.strip() == "Match") &
        (df["Period Name"].astype(str).str.strip() == "Session")
    ].copy()

    # Colonnes numériques GPS
    GPS_NUM_COLS = [
        "m/min", "HI", "THI", "%HI",
        "Distance totale", "Courue", "Marchée",
        "Vmax", "Vmax%",
        "RHIE Efforts Per Bout - Max", "RHIE Total Bouts",
        "Total Player Load",
        "Acc 2,5 m/s/s", "Dec -2,5 m/s/s", "Acc + Dec",
        "Acc Max", "Dec Max",
        "Contact Involvement Total Count",
        "DT 85%", "DT 90%",
    ]
    for col in GPS_NUM_COLS:
        if col in df_match.columns:
            df_match[col] = pd.to_numeric(df_match[col], errors="coerce")

    if "Date" in df_match.columns:
        df_match["Date"] = pd.to_datetime(df_match["Date"], errors="coerce")

    # Normaliser nom joueur
    df_match["joueur_norm"] = df_match["Player Name"].apply(normalize_name)

    print(f"  ✅ GPS chargé : {len(df_match)} lignes de match | {df_match['Activity Name'].nunique()} matchs")
    return df_match


# ─────────────────────────────────────────────────────────────────────
#  5. EXTRACTION DES STATS VIDÉO PAR JOUEUR
# ─────────────────────────────────────────────────────────────────────
AURI_PREFIXES = ["stade aurillacois", "aurillac", "sa -", "sa-"]

def detect_auri_prefix(df: pd.DataFrame) -> str | None:
    """
    Détecte le préfixe Aurillacois qui a le plus de lignes
    avec un joueur renseigné dans la colonne Joueur.
    Priorité : "Stade Aurillacois" > "SA" > autres.
    """
    candidates = {}
    has_joueur = "Joueur" in df.columns

    for line in df["Nom de la ligne"].dropna().unique():
        line_low = str(line).lower()
        for pref in AURI_PREFIXES:
            if line_low.startswith(pref) and " - " in str(line):
                prefix_val = str(line).split(" - ")[0].strip()
                if prefix_val not in candidates:
                    # Compter les lignes avec Joueur renseigné pour ce préfixe
                    rows = df[df["Nom de la ligne"].str.startswith(prefix_val, na=False)]
                    if has_joueur:
                        count = rows[rows["Joueur"].notna() &
                                     (rows["Joueur"].astype(str).str.strip() != "")
                                    ].shape[0]
                    else:
                        count = len(rows)
                    candidates[prefix_val] = count

    if not candidates:
        return None
    # Retourner le préfixe avec le plus de lignes joueur
    return max(candidates, key=candidates.get)


def extract_video_stats(df: pd.DataFrame, match_id: str) -> pd.DataFrame:
    """
    Agrège les stats vidéo par joueur pour un match donné.
    Retourne un DataFrame avec 1 ligne par joueur.
    """
    auri_prefix = detect_auri_prefix(df)
    if not auri_prefix:
        print(f"  ⚠ Préfixe Aurillac non détecté dans {match_id}")
        return pd.DataFrame()

    # ── Actions AURI → comptages globaux ──────────────────────────────
    df_auri = df[df["Nom de la ligne"].str.startswith(auri_prefix, na=False)].copy()
    df_auri["action"] = df_auri["Nom de la ligne"].str.replace(
        f"^{re.escape(auri_prefix)}\\s*-\\s*", "", regex=True
    ).str.strip()

    # ── Stats par joueur (colonne Joueur) ─────────────────────────────
    if "Joueur" not in df.columns:
        return pd.DataFrame()

    # Garder uniquement les lignes avec un joueur renseigné + action AURI
    df_joueur = df_auri[df_auri["Joueur"].notna() & (df_auri["Joueur"] != "")].copy()

    # Résultat normalisé
    if "Resultat" in df_joueur.columns:
        df_joueur["res_norm"] = df_joueur["Resultat"].str.strip().str.lower()
    else:
        df_joueur["res_norm"] = "inconnu"

    # Actions à agréger
    ACTIONS = {
        "Plaquages":        "Plaquages",
        "Contacts":         "Contacts",
        "Rucks":            "Rucks",
        "Passes":           "Passes",
        "Soutiens Offensifs": "Soutiens Offensifs",
        "Franchissements":  "Franchissements",
        "Ballons perdus":   "Ballons perdus",
        "Jeux au pied":     "Jeux au pied",
        "Assistant plaqueur": "Assistant plaqueur",
        "Melees":           "Melees",
        "Touches":          "Touches",
    }

    rows = []
    for joueur, grp in df_joueur.groupby("Joueur"):
        row = {
            "match_id":    match_id,
            "joueur_video": str(joueur),
            "joueur_norm":  normalize_name(str(joueur)),
        }
        # Comptage par type d'action
        action_counts = grp["action"].value_counts()
        for key, label in ACTIONS.items():
            row[f"vid_{label}"] = int(action_counts.get(key, 0))

        # Taux de réussite plaquages
        plq = grp[grp["action"] == "Plaquages"]
        if len(plq):
            row["vid_plq_reussis"]  = int((plq["res_norm"] == "positif").sum())
            row["vid_plq_total"]    = len(plq)
            row["vid_plq_pct"]      = round(row["vid_plq_reussis"] / len(plq) * 100, 1)
        else:
            row["vid_plq_reussis"] = 0
            row["vid_plq_total"]   = 0
            row["vid_plq_pct"]     = np.nan

        # Vitesse de libération rucks (si disponible)
        if "Vitesse de liberation" in grp.columns:
            ruck_rows = grp[grp["action"] == "Rucks"]
            if not ruck_rows.empty:
                speed_map = {"rapide": 3, "moyen": 2, "lent": 1}
                speeds = ruck_rows["Vitesse de liberation"].str.lower().map(speed_map)
                row["vid_ruck_vitesse_moy"] = round(speeds.mean(), 2)
            else:
                row["vid_ruck_vitesse_moy"] = np.nan

        # Stats par période (fatigue)
        if "Periode de jeu" in grp.columns:
            periods = {"1er quart-temps": "Q1", "2nd quart-temps": "Q2",
                       "3e quart-temps": "Q3", "4e quart-temps": "Q4"}
            for period_name, period_key in periods.items():
                p_grp = grp[grp["Periode de jeu"] == period_name]
                row[f"vid_actions_{period_key}"] = len(p_grp)
                plq_p = p_grp[p_grp["action"] == "Plaquages"]
                if len(plq_p):
                    row[f"vid_plq_pct_{period_key}"] = round(
                        (plq_p["res_norm"] == "positif").sum() / len(plq_p) * 100, 1
                    )
                else:
                    row[f"vid_plq_pct_{period_key}"] = np.nan

        rows.append(row)

    return pd.DataFrame(rows) if rows else pd.DataFrame()


# ─────────────────────────────────────────────────────────────────────
#  6. ALGORITHME PRINCIPAL DE FUSION
# ─────────────────────────────────────────────────────────────────────
def merge_gps_video(
    gps_path: str,
    video_files: list,
    output_path: str = "fusion_gps_video.xlsx",
    manual_mapping: dict = None,
) -> pd.DataFrame:
    """
    Fusionne les données GPS et vidéo.

    Paramètres :
    -----------
    gps_path      : chemin vers le fichier HUB DATAS .xlsx
    video_files   : liste de chemins vers les CSV vidéo
    output_path   : chemin du fichier Excel de sortie
    manual_mapping: dict optionnel {nom_fichier_csv: code_gps_match}
                   ex: {"AURI-COL Clip CSV.csv": "J16 COLOMIERS"}

    Retourne :
    ----------
    DataFrame fusionné (1 ligne = 1 joueur × 1 match)
    """
    print("\n" + "="*60)
    print("   FUSION GPS × VIDÉO — Stade Aurillacois")
    print("="*60)

    # ── Chargement GPS ──────────────────────────────────────────────
    print("\n📡 Chargement GPS...")
    df_gps = load_gps_excel(gps_path)
    gps_matches = sorted(df_gps["Activity Name"].dropna().unique().tolist())

    # ── Auto-détection correspondance matchs ────────────────────────
    print("\n🔍 Détection des correspondances matchs...")
    auto_map = auto_detect_match_mapping(video_files, gps_matches)

    # Appliquer corrections manuelles
    if manual_mapping:
        for fname, gps_code in manual_mapping.items():
            for key in auto_map:
                if os.path.basename(fname) in key or key in os.path.basename(fname):
                    auto_map[key]["gps_match"] = gps_code
                    auto_map[key]["confiance"] = "MANUEL"

    # Afficher le mapping
    print("\n┌─────────────────────────────────────────────────────────┐")
    print("│  Fichier vidéo             →  Match GPS         Statut  │")
    print("├─────────────────────────────────────────────────────────┤")
    for fname, info in auto_map.items():
        gps = info["gps_match"] or "❌ NON TROUVÉ"
        conf = info["confiance"]
        short_fname = fname[:28].ljust(28)
        short_gps   = str(gps)[:20].ljust(20)
        print(f"│  {short_fname} →  {short_gps} {conf:8} │")
    print("└─────────────────────────────────────────────────────────┘")

    # ── Traitement de chaque match vidéo ────────────────────────────
    all_video_stats = []
    all_fusion_rows = []
    mapping_log     = []

    for vf in video_files:
        fname = os.path.basename(vf)
        match_info = auto_map.get(fname, {})
        gps_code   = match_info.get("gps_match")

        print(f"\n🎬 Traitement vidéo : {fname}")
        df_vid = load_video_csv(vf)
        if df_vid is None:
            continue

        # Stats vidéo par joueur
        vid_stats = extract_video_stats(df_vid, fname)
        if vid_stats.empty:
            print(f"  ⚠ Aucune stat vidéo extraite")
            continue

        vid_stats["fichier_video"] = fname
        all_video_stats.append(vid_stats)
        print(f"  ✅ Vidéo : {len(vid_stats)} joueurs avec données")

        # Logs mapping
        mapping_log.append({
            "fichier_video": fname,
            "gps_match":     gps_code or "NON_TROUVÉ",
            "confiance":     match_info.get("confiance", "?"),
            "adversaire":    match_info.get("info", {}).get("adversaire_slug", "?"),
            "journee":       match_info.get("info", {}).get("journee", "?"),
        })

        # GPS pour ce match
        if not gps_code:
            print(f"  ⚠ Pas de match GPS correspondant → stats GPS absentes")
            # Ajouter quand même les stats vidéo sans GPS
            vid_stats["gps_match"] = None
            all_fusion_rows.append(vid_stats)
            continue

        df_gps_match = df_gps[df_gps["Activity Name"] == gps_code].copy()
        print(f"  📡 GPS {gps_code} : {len(df_gps_match)} joueurs GPS")

        if df_gps_match.empty:
            vid_stats["gps_match"] = gps_code
            all_fusion_rows.append(vid_stats)
            continue

        # ── Normalisation et jointure ────────────────────────────────
        # Colonnes GPS à garder
        GPS_COLS = [
            "joueur_norm", "Player Name", "Position Name", "Date",
            "m/min", "Distance totale", "Courue", "Marchée",
            "HI", "THI", "%HI",
            "Vmax", "Vmax%",
            "RHIE Efforts Per Bout - Max", "RHIE Total Bouts",
            "Total Player Load",
            "Acc 2,5 m/s/s", "Dec -2,5 m/s/s",
            "Contact Involvement Total Count",
        ]
        gps_keep = [c for c in GPS_COLS if c in df_gps_match.columns]
        df_gps_sub = df_gps_match[gps_keep].copy()
        df_gps_sub = df_gps_sub.rename(columns={
            c: f"gps_{c}" if c not in ["joueur_norm", "Player Name",
                                        "Position Name", "Date"]
            else c
            for c in df_gps_sub.columns
        })

        # Jointure sur joueur_norm
        merged = vid_stats.merge(df_gps_sub, on="joueur_norm", how="outer")
        merged["gps_match"] = gps_code

        # Compléter joueur_video si absent (joueur GPS sans vidéo)
        merged["joueur_video"] = merged["joueur_video"].fillna(
            merged["Player Name"] if "Player Name" in merged.columns else ""
        )

        # Indicateurs fusionnés
        if "gps_m/min" in merged.columns and "vid_Plaquages" in merged.columns:
            # Charge physique haute + plaquage bas → signal de fatigue
            mmin_med  = merged["gps_m/min"].median()
            plq_med   = merged["vid_plq_pct"].median()
            merged["signal_fatigue"] = (
                (merged["gps_m/min"] > mmin_med) &
                (merged["vid_plq_pct"] < plq_med)
            ).map({True: "⚠ SURCHARGE", False: "OK"})

        all_fusion_rows.append(merged)
        print(f"  🔗 Fusion : {len(merged)} joueurs fusionnés")

    # ── Assemblage final ─────────────────────────────────────────────
    if not all_fusion_rows:
        print("\n❌ Aucune donnée fusionnée.")
        return pd.DataFrame()

    df_fusion = pd.concat(all_fusion_rows, ignore_index=True)
    df_fusion_sorted = df_fusion.sort_values(
        ["gps_match", "joueur_norm"], na_position="last"
    )

    # ── Joueurs mapping (pour débogage) ─────────────────────────────
    player_map_rows = []
    for _, row in df_fusion.drop_duplicates(subset=["joueur_norm"]).iterrows():
        player_map_rows.append({
            "joueur_video": row.get("joueur_video", ""),
            "joueur_norm":  row.get("joueur_norm", ""),
            "joueur_gps":   row.get("Player Name", ""),
            "position_gps": row.get("Position Name", ""),
        })
    df_player_map = pd.DataFrame(player_map_rows)

    # ── Export Excel ─────────────────────────────────────────────────
    print(f"\n💾 Export vers : {output_path}")
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Feuille principale
        df_fusion_sorted.to_excel(writer, sheet_name="Fusion", index=False)

        # Stats vidéo brutes
        if all_video_stats:
            df_all_vid = pd.concat(all_video_stats, ignore_index=True)
            df_all_vid.to_excel(writer, sheet_name="Stats_video", index=False)

        # Mapping matchs
        pd.DataFrame(mapping_log).to_excel(
            writer, sheet_name="Mapping_matchs", index=False
        )

        # Mapping joueurs
        df_player_map.to_excel(writer, sheet_name="Mapping_joueurs", index=False)

    print(f"✅ Fichier Excel créé : {output_path}")
    print(f"   Lignes totales   : {len(df_fusion_sorted)}")
    print(f"   Joueurs uniques  : {df_fusion_sorted['joueur_norm'].nunique()}")
    print(f"   Matchs couverts  : {df_fusion_sorted['gps_match'].nunique()}")
    print("="*60 + "\n")

    return df_fusion_sorted


# ─────────────────────────────────────────────────────────────────────
#  7. POINT D'ENTRÉE CLI
# ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Fusion GPS × Vidéo — Stade Aurillacois"
    )
    parser.add_argument("--gps",   nargs="?",  help="Chemin HUB DATAS .xlsx")
    parser.add_argument("--video", nargs="*",  help="Fichiers CSV vidéo")
    parser.add_argument("--out",   default="fusion_gps_video.xlsx",
                        help="Fichier de sortie .xlsx")
    # Mapping manuel : paires fichier_csv=code_gps
    parser.add_argument("--map",   nargs="*",
                        help="Corrections manuelles ex: AURI-COL.csv=J16 COLOMIERS")
    args = parser.parse_args()

    # Auto-découverte si pas d'arguments
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    DATA_DIR = os.path.join(BASE_DIR, "data")

    gps_path = args.gps
    if not gps_path:
        # Chercher le fichier GPS
        for ext in ["*.xlsx", "*.xls"]:
            found = glob.glob(os.path.join(BASE_DIR, "**", ext), recursive=True)
            found = [f for f in found if "HUB" in os.path.basename(f).upper()
                     or "GPS" in os.path.basename(f).upper()
                     or "DATAS" in os.path.basename(f).upper()]
            if found:
                gps_path = found[0]
                print(f"📡 GPS auto-détecté : {gps_path}")
                break

    video_files = args.video
    if not video_files:
        video_files = (
            glob.glob(os.path.join(DATA_DIR, "*.csv")) +
            glob.glob(os.path.join(BASE_DIR, "*.csv"))
        )
        video_files = [f for f in video_files
                       if "sync" not in f.lower() and "merge" not in f.lower()
                       and "fusion" not in f.lower()]
        print(f"🎬 CSV vidéo auto-détectés : {len(video_files)} fichiers")

    if not gps_path or not video_files:
        print("❌ Fichiers non trouvés. Utilise --gps et --video.")
        exit(1)

    # Parsing du mapping manuel
    manual_map = {}
    if args.map:
        for pair in args.map:
            if "=" in pair:
                fname, gcode = pair.split("=", 1)
                manual_map[fname.strip()] = gcode.strip()

    out_path = os.path.join(BASE_DIR, args.out)
    merge_gps_video(gps_path, video_files, out_path, manual_map)
